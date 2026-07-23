"""Excel workbook for the Bursar weekly report (openpyxl)."""
from __future__ import annotations

import io
from typing import Any

from django.utils import timezone


def render_bursar_weekly_excel(metrics: dict[str, Any]) -> tuple[bytes, str]:
    """Return (xlsx_bytes, filename)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to render the bursar report Excel.") from exc

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    title_font = Font(bold=True, size=14, color="1F3A5F")
    good_font = Font(bold=True, color="1B7A3D")
    bad_font = Font(bold=True, color="A11C1C")

    def style_header(ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

    def autosize(ws, max_width: int = 40):
        from openpyxl.cell.cell import MergedCell

        for column_cells in ws.columns:
            letter = None
            length = 0
            for cell in column_cells:
                if isinstance(cell, MergedCell):
                    continue
                if letter is None:
                    letter = cell.column_letter
                val = "" if cell.value is None else str(cell.value)
                length = max(length, min(len(val), max_width))
            if letter:
                ws.column_dimensions[letter].width = max(12, length + 2)

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = metrics.get("report_title") or "Bursar Weekly Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    rows = [
        ("University", metrics.get("university_name")),
        ("Prepared for", metrics.get("prepared_for")),
        ("Intake / cohort", metrics.get("intake_label")),
        ("Report date", metrics.get("report_date")),
        ("Data as of", metrics.get("data_as_of")),
        ("Week", f"{metrics.get('week_start')} – {metrics.get('week_end')}"),
        ("", ""),
        ("Admitted", metrics.get("admitted_total")),
        ("Paid (strict)", metrics.get("paid_total")),
        ("Not paid", metrics.get("not_paid_total")),
        ("Collection rate %", metrics.get("collection_rate")),
        ("Total collected", metrics.get("total_collected_display")),
        ("Revenue at risk", metrics.get("revenue_at_risk_display")),
        ("Commitment threshold", metrics.get("threshold_display")),
        ("Flag-only paid (admission_fee_paid)", metrics.get("flag_paid_total")),
        ("Applications this week", metrics.get("applications_received_week")),
        ("Pipeline pending", metrics.get("applications_pending")),
        ("Programme enrolled", metrics.get("enrolled_count")),
        ("Enrolment pending", metrics.get("enrolment_pending")),
        ("Local", metrics.get("local_count")),
        ("International", metrics.get("international_count")),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=value)
        if label == "Paid (strict)":
            cell.font = good_font
        if label in ("Not paid", "Revenue at risk"):
            cell.font = bad_font
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Risk statement")
    ws.cell(row=r, column=2, value=metrics.get("risk_statement"))
    r += 2
    ws.cell(row=r, column=1, value="Reconciliation")
    ws.cell(row=r, column=2, value=metrics.get("reconciliation_note"))
    r += 2
    ws.cell(row=r, column=1, value="Source")
    ws.cell(row=r, column=2, value=metrics.get("source_note"))
    autosize(ws)

    # --- By Faculty ---
    ws_f = wb.create_sheet("By Faculty")
    headers = ["Faculty", "Admitted", "Paid", "Not paid", "Rate %", "Amount", "At risk"]
    ws_f.append(headers)
    style_header(ws_f, 1, len(headers))
    for row in metrics.get("by_faculty") or []:
        ws_f.append(
            [
                row.get("name"),
                row.get("admitted"),
                row.get("paid"),
                row.get("not_paid"),
                row.get("collection_rate"),
                row.get("amount_display") or row.get("amount"),
                row.get("revenue_at_risk_display") or row.get("revenue_at_risk"),
            ]
        )
    autosize(ws_f)

    # --- By Campus ---
    ws_c = wb.create_sheet("By Campus")
    headers = ["Campus", "Admitted", "Paid", "Not paid", "Rate %", "Amount"]
    ws_c.append(headers)
    style_header(ws_c, 1, len(headers))
    for row in metrics.get("by_campus") or []:
        ws_c.append(
            [
                row.get("name"),
                row.get("admitted"),
                row.get("paid"),
                row.get("not_paid"),
                row.get("collection_rate"),
                row.get("amount_display") or row.get("amount"),
            ]
        )
    autosize(ws_c)

    # --- Demographics ---
    ws_d = wb.create_sheet("Demographics")
    ws_d.append(["Dimension", "Name", "Count", "Pct %"])
    style_header(ws_d, 1, 4)
    for row in metrics.get("by_level") or []:
        ws_d.append(["Academic level", row.get("name"), row.get("count"), row.get("pct")])
    for row in metrics.get("by_gender") or []:
        ws_d.append(["Gender", row.get("name"), row.get("count"), row.get("pct")])
    autosize(ws_d)

    # --- Monthly applications ---
    ws_ma = wb.create_sheet("Monthly Applications")
    ws_ma.append(["Month", "Applications"])
    style_header(ws_ma, 1, 2)
    for row in metrics.get("monthly_applications") or []:
        ws_ma.append([row.get("month"), row.get("count")])
    autosize(ws_ma)

    # --- Monthly collections ---
    ws_mc = wb.create_sheet("Monthly Collections")
    ws_mc.append(["Month", "Transactions", "Amount"])
    style_header(ws_mc, 1, 3)
    for row in metrics.get("monthly_collections") or []:
        ws_mc.append(
            [row.get("month"), row.get("transactions"), row.get("amount_display") or row.get("amount")]
        )
    autosize(ws_mc)

    # --- Observations ---
    ws_o = wb.create_sheet("Observations")
    ws_o["A1"] = "Key observations"
    ws_o["A1"].font = title_font
    r = 3
    for item in metrics.get("observations") or []:
        ws_o.cell(row=r, column=1, value=item)
        r += 1
    r += 1
    ws_o.cell(row=r, column=1, value="Recommendations").font = title_font
    r += 2
    for item in metrics.get("recommendations") or []:
        ws_o.cell(row=r, column=1, value=item)
        r += 1
    ws_o.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    date_stamp = timezone.localdate().isoformat()
    filename = f"Bursar_Weekly_Report_{date_stamp}.xlsx"
    return buf.getvalue(), filename
