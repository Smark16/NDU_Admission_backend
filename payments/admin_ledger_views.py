"""Staff-facing tuition payment ledger for admitted students."""
from __future__ import annotations

import logging
import csv
from datetime import datetime

from decimal import Decimal

from django.db.models import Count, DecimalField, F, Max, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_date
from accounts.erp_drf_permissions import (
    AccountsClearedReportPermission,
    FinanceModuleAdminPermission,
    IsSuperAdminOnly,
)
from accounts.super_admin import user_is_super_admin
from rest_framework import status
from rest_framework.permissions import BasePermission
from rest_framework.response import Response
from rest_framework.views import APIView

from admissions.models import AdmittedStudent, Batch, Campus
from Programs.models import Program, ProgramBatch

from .models import StudentTuitionPayment, TuitionLedger
from .student_portal_finance import (
    COMMITMENT_FEE_THRESHOLD,
    payment_status_dict,
    student_billing_lines,
    student_finance_totals,
)

from .tasks import queue_bulk_commitment_reminders

logger = logging.getLogger(__name__)

from .commitment_queryset import annotate_commitment_ugx_paid, filter_by_commitment_met

def _parse_page(value, default: int = 1) -> int:
    try:
        page = int(value)
    except (TypeError, ValueError):
        return default
    return max(page, 1)


def _parse_page_size(value, default: int = 25, maximum: int = 100) -> int:
    try:
        size = int(value)
    except (TypeError, ValueError):
        return default
    return min(max(size, 1), maximum)


def _parse_bool(value: str | None) -> bool | None:
    if value is None or value == "":
        return None
    return value.lower() in {"1", "true", "yes"}


def _parse_int(value: str | None) -> int | None:
    if value is None or value == "":
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _batch_intake_label(batch: Batch | None) -> str | None:
    if batch is None:
        return None
    if batch.academic_year:
        return f"{batch.name} ({batch.academic_year})"
    return batch.name


def _ledger_cohort_params(request) -> dict[str, int | str | None]:
    return {
        "batch_id": _parse_int(request.query_params.get("batch_id") or request.query_params.get("intake_id")),
        "program_id": _parse_int(request.query_params.get("program_id")),
        "campus_id": _parse_int(request.query_params.get("campus_id")),
        "program_batch_id": _parse_int(request.query_params.get("program_batch_id")),
        "academic_year": (request.query_params.get("academic_year") or "").strip() or None,
        "intake": (request.query_params.get("intake") or "").strip() or None,
    }


def _apply_student_cohort_filters(qs, cohort: dict[str, int | str | None]):
    if cohort["batch_id"]:
        qs = qs.filter(admitted_batch_id=cohort["batch_id"])
    if cohort["program_id"]:
        qs = qs.filter(admitted_program_id=cohort["program_id"])
    if cohort["campus_id"]:
        qs = qs.filter(admitted_campus_id=cohort["campus_id"])
    if cohort["academic_year"]:
        qs = qs.filter(admitted_batch__academic_year=cohort["academic_year"])
    if cohort["program_batch_id"]:
        qs = qs.filter(programme_enrollment__program_batch_id=cohort["program_batch_id"])
    if cohort["intake"]:
        intake = cohort["intake"]
        if intake.isdigit():
            qs = qs.filter(admitted_batch_id=int(intake))
        else:
            qs = qs.filter(
                Q(admitted_batch__name=cohort["intake"])
                | Q(admitted_batch__code=cohort["intake"])
            )
    return qs


def _apply_transaction_cohort_filters(qs, cohort: dict[str, int | str | None]):
    if cohort["batch_id"]:
        qs = qs.filter(student__admitted_batch_id=cohort["batch_id"])
    if cohort["program_id"]:
        qs = qs.filter(student__admitted_program_id=cohort["program_id"])
    if cohort["campus_id"]:
        qs = qs.filter(student__admitted_campus_id=cohort["campus_id"])
    if cohort["academic_year"]:
        qs = qs.filter(student__admitted_batch__academic_year=cohort["academic_year"])
    if cohort["program_batch_id"]:
        qs = qs.filter(student__programme_enrollment__program_batch_id=cohort["program_batch_id"])
    if cohort["intake"]:
        intake = cohort["intake"]
        if intake.isdigit():
            qs = qs.filter(student__admitted_batch_id=int(intake))
        else:
            qs = qs.filter(
                Q(student__admitted_batch__name=cohort["intake"])
                | Q(student__admitted_batch__code=cohort["intake"])
            )
    return qs


def _student_search_filter(search: str) -> Q:
    term = (search or "").strip()
    if not term:
        return Q()
    return (
        Q(student_id__icontains=term)
        | Q(reg_no__icontains=term)
        | Q(schoolpay_code__icontains=term)
        | Q(application__first_name__icontains=term)
        | Q(application__last_name__icontains=term)
    )


def _transaction_search_filter(search: str) -> Q:
    term = (search or "").strip()
    if not term:
        return Q()
    return (
        Q(student__student_id__icontains=term)
        | Q(student__reg_no__icontains=term)
        | Q(student__schoolpay_code__icontains=term)
        | Q(student__application__first_name__icontains=term)
        | Q(student__application__last_name__icontains=term)
        | Q(receipt_number__icontains=term)
        | Q(payment_reference__icontains=term)
        | Q(label__icontains=term)
    )


def _student_payment_counts(student: AdmittedStudent) -> tuple[int, int]:
    if "completed_payment_count" in student.__dict__:
        return (
            int(student.completed_payment_count or 0),
            int(student.pending_payment_count or 0),
        )
    completed = StudentTuitionPayment.objects.filter(student=student, status="completed").count()
    pending = StudentTuitionPayment.objects.filter(student=student, status="pending").count()
    return completed, pending


def _student_last_paid_at(student: AdmittedStudent):
    if "last_paid_at" in student.__dict__:
        return student.last_paid_at
    return (
        StudentTuitionPayment.objects.filter(student=student, status="completed")
        .order_by("-paid_at", "-created_at")
        .values_list("paid_at", flat=True)
        .first()
    )


def _cohort_finance_summary(students_qs) -> dict[str, float]:
    billed = Decimal("0")
    paid = Decimal("0")
    balance = Decimal("0")
    for student in students_qs.iterator(chunk_size=100):
        totals = student_finance_totals(student)
        billed += Decimal(str(totals["total_required"]))
        paid += Decimal(str(totals["total_paid"]))
        balance += Decimal(str(totals["balance"]))
    return {
        "total_billed": float(billed),
        "total_paid": float(paid),
        "total_balance": float(balance),
    }


def _student_display_name(student: AdmittedStudent) -> str:
    try:
        return student.full_name or ""
    except Exception:
        if student.application_id and student.application:
            return getattr(student.application, "full_name", "") or ""
    return ""


def _payment_code_fields(student: AdmittedStudent) -> dict:
    """SchoolPay wallet code vs reg. no. (students not synced to SchoolPay have no wallet yet)."""
    wallet = (student.student_id or student.schoolpay_code or "").strip()
    reg_no = (student.reg_no or "").strip()
    return {
        "student_id": student.student_id,
        "schoolpay_code": wallet or None,
        "payment_code": wallet or reg_no or None,
        "schoolpay_registered": bool(wallet),
        "payment_code_is_reg_no": not wallet and bool(reg_no),
    }


def _commitment_student_row(student: AdmittedStudent) -> dict:
    """Lightweight list row — uses commitment annotations when present."""
    threshold = float(COMMITMENT_FEE_THRESHOLD)
    paid_raw = getattr(student, "commitment_paid_ugx", None)
    if paid_raw is None:
        finance = student_finance_totals(student)
        paid = float(finance["commitment_paid_ugx"])
        met = bool(finance["commitment_met"])
        balance = float(finance["commitment_balance"])
    else:
        paid = float(paid_raw or 0)
        admission_paid = bool(student.admission_fee_paid)
        met = admission_paid or paid >= threshold
        balance = 0.0 if met else max(threshold - paid, 0.0)

    enrollment_status = None
    try:
        enrollment_status = student.programme_enrollment.status
    except Exception:
        enrollment_status = None

    has_temp = getattr(student, "has_temporary_access_pass", None)
    temp_sponsor = getattr(student, "temporary_access_sponsor", None)
    temp_until = getattr(student, "temporary_access_valid_until", None)
    if has_temp is None:
        from admissions.temporary_access import get_active_pass

        active_pass = get_active_pass(student)
        has_temp = active_pass is not None
        temp_sponsor = active_pass.sponsor_label if active_pass else None
        temp_until = active_pass.valid_until if active_pass else None

    return {
        "id": student.id,
        "reg_no": student.reg_no,
        "student_name": _student_display_name(student),
        **_payment_code_fields(student),
        "program": student.admitted_program.name if student.admitted_program_id else None,
        "campus": student.admitted_campus.name if student.admitted_campus_id else None,
        "batch_id": student.admitted_batch_id,
        "batch_name": student.admitted_batch.name if student.admitted_batch_id else None,
        "academic_year": student.admitted_batch.academic_year if student.admitted_batch_id else None,
        "intake": _batch_intake_label(student.admitted_batch if student.admitted_batch_id else None),
        "commitment_threshold": threshold,
        "commitment_paid_ugx": paid,
        "commitment_met": met,
        "commitment_balance": balance,
        "total_paid": paid,
        "balance": balance,
        "enrollment_status": enrollment_status,
        "has_temporary_access_pass": bool(has_temp),
        "temporary_access_sponsor": temp_sponsor or None,
        "temporary_access_valid_until": (
            temp_until.isoformat() if hasattr(temp_until, "isoformat") else temp_until
        ),
    }


def _commitment_students_queryset(
    *,
    search: str = "",
    cohort: dict[str, int | str | None] | None = None,
    commitment_met: bool | None = None,
):
    from admissions.temporary_access import annotate_temporary_access

    qs = annotate_temporary_access(
        AdmittedStudent.objects.filter(is_admitted=True)
        .select_related(
            "admitted_program",
            "admitted_campus",
            "admitted_batch",
            "application",
            "programme_enrollment",
        )
        .filter(_student_search_filter(search))
        .order_by(F("student_id").asc(nulls_last=True), "reg_no", "-id")
    )
    if cohort:
        qs = _apply_student_cohort_filters(qs, cohort)
    if commitment_met is not None:
        qs = filter_by_commitment_met(qs, commitment_met, strict=True)
    else:
        qs = annotate_commitment_ugx_paid(qs)
    return qs


class _CsvEcho:
    """Write CSV rows for StreamingHttpResponse."""

    def write(self, value: str) -> str:
        return value


def _student_row(student: AdmittedStudent) -> dict:
    finance = student_finance_totals(student)
    completed_count, pending_count = _student_payment_counts(student)
    paid_at = _student_last_paid_at(student)
    enrollment_status = None
    try:
        enrollment_status = student.programme_enrollment.status
    except Exception:
        enrollment_status = None

    return {
        "id": student.id,
        "student_id": student.student_id,
        "reg_no": student.reg_no,
        "student_name": _student_display_name(student),
        "program": student.admitted_program.name if student.admitted_program_id else None,
        "campus": student.admitted_campus.name if student.admitted_campus_id else None,
        "batch_id": student.admitted_batch_id,
        "batch_name": student.admitted_batch.name if student.admitted_batch_id else None,
        "academic_year": student.admitted_batch.academic_year if student.admitted_batch_id else None,
        "intake": _batch_intake_label(student.admitted_batch if student.admitted_batch_id else None),
        "schoolpay_code": student.effective_schoolpay_code,
        "total_required": finance["total_required"],
        "total_paid": finance["total_paid"],
        "balance": finance["balance"],
        "percentage_paid": finance["percentage_paid"],
        "display_currency": finance["display_currency"],
        "tuition_structure_total": finance["tuition_structure_total"],
        "ad_hoc_total": finance["ad_hoc_total"],
        "ad_hoc_not_yet_due_total": finance.get("ad_hoc_not_yet_due_total", 0),
        "scheduled_other_fees_due": finance["scheduled_other_fees_due"],
        "commitment_threshold": finance["commitment_threshold"],
        "commitment_paid_ugx": finance["commitment_paid_ugx"],
        "commitment_met": finance["commitment_met"],
        "commitment_balance": finance["commitment_balance"],
        "completed_payment_count": completed_count,
        "pending_payment_count": pending_count,
        "last_paid_at": paid_at.isoformat() if paid_at else None,
        "enrollment_status": enrollment_status,
    }


def _transaction_row(payment: StudentTuitionPayment) -> dict:
    student = payment.student
    if payment.source == "ad_hoc":
        fee_label = payment.label or (payment.fee_head.name if payment.fee_head_id else "Ad-hoc charge")
    else:
        fee_label = payment.label or (
            payment.fee_plan_rule.fee_head.name
            if payment.fee_plan_rule_id and payment.fee_plan_rule.fee_head_id
            else "Tuition"
        )

    return {
        "id": payment.id,
        "student_pk": student.id,
        "student_id": student.student_id,
        "reg_no": student.reg_no,
        "student_name": _student_display_name(student),
        "program": student.admitted_program.name if student.admitted_program_id else None,
        "intake": _batch_intake_label(student.admitted_batch if student.admitted_batch_id else None),
        "amount": float(payment.amount),
        "currency": payment.currency or "UGX",
        "status": payment.status,
        "source": payment.source,
        "label": fee_label,
        "payment_method": payment.payment_method or "",
        "receipt_number": payment.receipt_number or "",
        "payment_reference": payment.payment_reference or "",
        "paid_at": payment.paid_at.isoformat() if payment.paid_at else None,
        "created_at": payment.created_at.isoformat(),
        "is_waived": payment.is_waived,
    }


class AdminTuitionLedgerFiltersView(APIView):
    """GET /api/payments/admin/tuition_ledger/filters"""

    permission_classes = [FinanceModuleAdminPermission]

    def get(self, request):
        admitted = AdmittedStudent.objects.filter(is_admitted=True)
        batch_ids = admitted.values_list("admitted_batch_id", flat=True).distinct()
        program_ids = admitted.values_list("admitted_program_id", flat=True).distinct()
        campus_ids = admitted.values_list("admitted_campus_id", flat=True).distinct()
        program_batch_ids = (
            admitted.filter(programme_enrollment__program_batch_id__isnull=False)
            .values_list("programme_enrollment__program_batch_id", flat=True)
            .distinct()
        )

        intakes = [
            {
                "id": batch.id,
                "name": batch.name,
                "academic_year": batch.academic_year,
                "label": _batch_intake_label(batch),
            }
            for batch in Batch.objects.filter(id__in=batch_ids).order_by("-created_at")
        ]
        academic_years = list(
            Batch.objects.filter(id__in=batch_ids)
            .exclude(academic_year="")
            .order_by("-academic_year")
            .values_list("academic_year", flat=True)
            .distinct()
        )
        programs = list(
            Program.objects.filter(id__in=program_ids).order_by("name").values("id", "name")
        )
        campuses = list(
            Campus.objects.filter(id__in=campus_ids).order_by("name").values("id", "name")
        )
        program_batches = [
            {
                "id": batch.id,
                "name": batch.name,
                "program_id": batch.program_id,
                "program_name": batch.program.name if batch.program_id else None,
                "academic_year": batch.academic_year,
            }
            for batch in ProgramBatch.objects.filter(id__in=program_batch_ids)
            .select_related("program")
            .order_by("program__name", "name")
        ]

        return Response(
            {
                "intakes": intakes,
                "academic_years": academic_years,
                "programs": programs,
                "campuses": campuses,
                "program_batches": program_batches,
            }
        )


class AdminTuitionLedgerStudentsView(APIView):
    """GET /api/payments/admin/tuition_ledger/students"""

    permission_classes = [FinanceModuleAdminPermission]

    def get(self, request):
        page = _parse_page(request.query_params.get("page"))
        page_size = _parse_page_size(request.query_params.get("page_size"))
        search = request.query_params.get("search", "")
        commitment_met = _parse_bool(request.query_params.get("commitment_met"))
        cohort = _ledger_cohort_params(request)
        include_finance_summary = _parse_bool(
            request.query_params.get("include_finance_summary")
        )
        skip_summary = _parse_bool(request.query_params.get("skip_summary"))

        qs = _commitment_students_queryset(
            search=search,
            cohort=cohort,
            commitment_met=commitment_met,
        )

        total = qs.count()
        offset = (page - 1) * page_size
        page_qs = list(qs[offset : offset + page_size])
        rows = []
        for student in page_qs:
            try:
                rows.append(_commitment_student_row(student))
            except Exception:
                rows.append(
                    {
                        "id": student.id,
                        "student_id": student.student_id,
                        "reg_no": student.reg_no,
                        "student_name": _student_display_name(student),
                        "program": student.admitted_program.name
                        if student.admitted_program_id
                        else None,
                        "campus": student.admitted_campus.name
                        if student.admitted_campus_id
                        else None,
                        "commitment_threshold": float(COMMITMENT_FEE_THRESHOLD),
                        "commitment_paid_ugx": 0.0,
                        "commitment_met": bool(student.admission_fee_paid),
                        "commitment_balance": float(COMMITMENT_FEE_THRESHOLD),
                        "total_paid": 0.0,
                        "balance": float(COMMITMENT_FEE_THRESHOLD),
                        "enrollment_status": None,
                    }
                )

        summary = {
            "students_count": 0,
            "commitment_met_count": 0,
            "commitment_threshold": float(COMMITMENT_FEE_THRESHOLD),
            "completed_payments_count": 0,
            "completed_amount_ugx": 0.0,
        }
        if not skip_summary:
            summary_qs = _apply_student_cohort_filters(
                AdmittedStudent.objects.filter(is_admitted=True),
                cohort,
            )
            commitment_met_count = filter_by_commitment_met(
                summary_qs, True, strict=True
            ).count()
            payment_totals = _apply_transaction_cohort_filters(
                StudentTuitionPayment.objects.filter(status="completed"),
                cohort,
            ).aggregate(
                completed_count=Count("id"),
                completed_amount_ugx=Coalesce(
                    Sum("amount", filter=Q(currency="UGX")),
                    Value(0),
                    output_field=DecimalField(max_digits=14, decimal_places=2),
                ),
            )
            summary = {
                "students_count": summary_qs.count(),
                "commitment_met_count": commitment_met_count,
                "commitment_threshold": float(COMMITMENT_FEE_THRESHOLD),
                "completed_payments_count": int(payment_totals["completed_count"] or 0),
                "completed_amount_ugx": float(payment_totals["completed_amount_ugx"] or 0),
            }
            if include_finance_summary:
                try:
                    summary.update(
                        _cohort_finance_summary(
                            summary_qs.select_related(
                                "admitted_program",
                                "admitted_campus",
                                "admitted_batch",
                                "application",
                                "programme_enrollment",
                            )
                        )
                    )
                except Exception:
                    summary.update(
                        {
                            "total_billed": 0.0,
                            "total_paid": 0.0,
                            "total_balance": 0.0,
                        }
                    )

        return Response(
            {
                "summary": summary,
                "filters": cohort,
                "results": rows,
                "page": page,
                "page_size": page_size,
                "total": total,
            }
        )


class AdminTuitionLedgerStudentsExportView(APIView):
    """GET /api/payments/admin/tuition_ledger/students/export — CSV download."""

    permission_classes = [FinanceModuleAdminPermission]

    def get(self, request):
        search = request.query_params.get("search", "")
        commitment_met = _parse_bool(request.query_params.get("commitment_met"))
        if commitment_met is None:
            commitment_met = False
        cohort = _ledger_cohort_params(request)
        qs = _commitment_students_queryset(
            search=search,
            cohort=cohort,
            commitment_met=commitment_met,
        )

        label = "paid" if commitment_met else "unpaid"
        filename = f"commitment_{label}_{datetime.now().strftime('%Y-%m-%d')}.csv"

        def stream_rows():
            pseudo_buffer = _CsvEcho()
            writer = csv.writer(pseudo_buffer)
            yield writer.writerow(
                [
                    "Pay code (SchoolPay wallet)",
                    "Reg No",
                    "SchoolPay synced",
                    "Name",
                    "Program",
                    "Campus",
                    "Intake",
                    "Commitment Paid (UGX)",
                    "Commitment Balance (UGX)",
                    "Status",
                ]
            )
            for student in qs.iterator(chunk_size=500):
                try:
                    row = _commitment_student_row(student)
                except Exception:
                    pay = _payment_code_fields(student)
                    row = {
                        **pay,
                        "reg_no": student.reg_no,
                        "student_name": _student_display_name(student),
                        "program": None,
                        "campus": None,
                        "intake": None,
                        "commitment_paid_ugx": 0.0,
                        "commitment_balance": float(COMMITMENT_FEE_THRESHOLD),
                        "commitment_met": bool(student.admission_fee_paid),
                    }
                yield writer.writerow(
                    [
                        row.get("schoolpay_code") or "",
                        row.get("reg_no") or "",
                        "Yes" if row.get("schoolpay_registered") else "No",
                        row.get("student_name") or "",
                        row.get("program") or "",
                        row.get("campus") or "",
                        row.get("intake") or "",
                        row.get("commitment_paid_ugx") or 0,
                        row.get("commitment_balance") or 0,
                        "Paid" if row.get("commitment_met") else "Not paid",
                    ]
                )

        response = StreamingHttpResponse(stream_rows(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class AdminTuitionLedgerStudentDetailView(APIView):
    """GET /api/payments/admin/tuition_ledger/students/<student_id>"""

    permission_classes = [FinanceModuleAdminPermission]

    def get(self, request, student_id):
        student = get_object_or_404(
            AdmittedStudent.objects.select_related(
                "admitted_program",
                "admitted_campus",
                "admitted_batch",
                "application",
                "programme_enrollment",
            ),
            pk=student_id,
            is_admitted=True,
        )
        finance = payment_status_dict(student, request)
        return Response(
            {
                "student": _student_row(student),
                "finance": finance,
                "billing_lines": student_billing_lines(student),
            }
        )


class CanApproveManualBankPayment(BasePermission):
    message = "Only Bursar, Finance Manager, or Super Admin can approve bank payment changes."

    def has_permission(self, request, view):
        from payments.manual_bank_approval import user_can_approve_manual_bank

        return bool(
            request.user
            and request.user.is_authenticated
            and user_can_approve_manual_bank(request.user)
        )


class CanPostManualBankPayment(BasePermission):
    message = "Only Bursar, Finance Manager, or Super Admin can post a bank payment."

    def has_permission(self, request, view):
        from payments.manual_bank_approval import user_can_post_manual_bank

        return bool(
            request.user
            and request.user.is_authenticated
            and user_can_post_manual_bank(request.user)
        )


class AdminPostManualBankPaymentView(APIView):
    """
    POST — record a bank payment (Bursar / Finance Manager / Super Admin).

    Approver posts apply immediately (no dual-control wait).
    """

    permission_classes = [CanPostManualBankPayment]

    def post(self, request, student_id):
        from audit.utils import log_audit_event
        from payments.manual_bank_approval import serialize_change_request, submit_change_request
        from payments.models import ManualBankPaymentChangeRequest

        student = get_object_or_404(
            AdmittedStudent.objects.select_related("application"),
            pk=student_id,
            is_admitted=True,
        )
        payload = {
            "amount": request.data.get("amount"),
            "bank_reference": request.data.get("bank_reference")
            or request.data.get("reference")
            or "",
            "payment_date": request.data.get("payment_date")
            or request.data.get("paid_at"),
            "notes": request.data.get("notes") or request.data.get("narration") or "",
            "bank_name": request.data.get("bank_name") or "",
        }
        try:
            change, result, applied = submit_change_request(
                request_type=ManualBankPaymentChangeRequest.RequestType.POST,
                student=student,
                requested_by=request.user,
                payload=payload,
                reason=(request.data.get("reason") or "").strip(),
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        actor = request.user.get_full_name() or request.user.email or str(request.user.pk)
        log_audit_event(
            request.user,
            "manual_bank_payment_applied" if applied else "manual_bank_payment_request",
            student,
            (
                f"Bank payment POST {'applied' if applied else 'requested'} amount={payload.get('amount')} "
                f"ref={payload.get('bank_reference')} by={actor} request_id={change.id}"
            ),
            request,
        )
        return Response(
            {
                "message": (
                    "Bank payment credited to the student."
                    if applied
                    else (
                        "Bank payment submitted for Bursar / Finance Manager approval. "
                        "It will credit the student only after approval."
                    )
                ),
                "pending": not applied,
                "request": serialize_change_request(change),
                "ledger_id": getattr(result, "id", None) if result is not None else change.ledger_id,
            },
            status=status.HTTP_201_CREATED,
        )


class AdminManualBankPaymentDetailView(APIView):
    """
    PATCH / DELETE — System Admin only (accountability). Bursar may post, not alter.
    """

    permission_classes = [IsSuperAdminOnly]

    def patch(self, request, ledger_id):
        from audit.utils import log_audit_event
        from payments.manual_bank_approval import serialize_change_request, submit_change_request
        from payments.manual_bank_payment import get_manual_bank_ledger
        from payments.models import ManualBankPaymentChangeRequest

        try:
            ledger = get_manual_bank_ledger(ledger_id)
        except TuitionLedger.DoesNotExist:
            return Response({"detail": "Bank payment not found."}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        student = ledger.student
        if student is None:
            return Response(
                {"detail": "Bank payment is not linked to a student."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = request.data
        payload = {}
        if "amount" in data:
            payload["amount"] = data.get("amount")
        if "bank_reference" in data or "reference" in data:
            payload["bank_reference"] = data.get("bank_reference") or data.get("reference") or ""
        if "payment_date" in data or "paid_at" in data:
            payload["payment_date"] = data.get("payment_date") or data.get("paid_at")
        if "notes" in data or "narration" in data:
            payload["notes"] = data.get("notes") if "notes" in data else data.get("narration")
        if "bank_name" in data:
            payload["bank_name"] = data.get("bank_name") or ""

        reason = (data.get("reason") or "").strip()
        try:
            change, result, applied = submit_change_request(
                request_type=ManualBankPaymentChangeRequest.RequestType.UPDATE,
                student=student,
                ledger=ledger,
                requested_by=request.user,
                payload=payload,
                reason=reason,
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        actor = request.user.get_full_name() or request.user.email or str(request.user.pk)
        log_audit_event(
            request.user,
            "manual_bank_payment_edit_applied" if applied else "manual_bank_payment_edit_request",
            student,
            (
                f"Bank payment EDIT {'applied' if applied else 'requested'} ledger={ledger.id} "
                f"reason={reason!r} by={actor} request_id={change.id}"
            ),
            request,
        )
        return Response(
            {
                "message": (
                    "Bank payment updated."
                    if applied
                    else "Edit submitted for Bursar / Finance Manager approval."
                ),
                "pending": not applied,
                "request": serialize_change_request(change),
                "ledger_id": getattr(result, "id", None) if result is not None else change.ledger_id,
            },
            status=status.HTTP_201_CREATED,
        )

    def delete(self, request, ledger_id):
        from audit.utils import log_audit_event
        from payments.manual_bank_approval import serialize_change_request, submit_change_request
        from payments.manual_bank_payment import get_manual_bank_ledger
        from payments.models import ManualBankPaymentChangeRequest

        try:
            ledger = get_manual_bank_ledger(ledger_id)
        except TuitionLedger.DoesNotExist:
            return Response({"detail": "Bank payment not found."}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        student = ledger.student
        if student is None:
            return Response(
                {"detail": "Bank payment is not linked to a student."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        reason = (
            request.data.get("reason")
            if hasattr(request, "data") and request.data is not None
            else None
        )
        if reason is None:
            reason = request.query_params.get("reason") or ""
        reason = str(reason).strip()

        try:
            change, _result, applied = submit_change_request(
                request_type=ManualBankPaymentChangeRequest.RequestType.DELETE,
                student=student,
                ledger=ledger,
                requested_by=request.user,
                payload={
                    "amount": str(ledger.amount),
                    "bank_reference": ledger.source_channel_transaction_id,
                    "receipt": ledger.schoolpay_receipt_number,
                },
                reason=reason,
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        actor = request.user.get_full_name() or request.user.email or str(request.user.pk)
        log_audit_event(
            request.user,
            "manual_bank_payment_delete_applied" if applied else "manual_bank_payment_delete_request",
            student,
            (
                f"Bank payment DELETE {'applied' if applied else 'requested'} ledger={ledger.id} "
                f"receipt={ledger.schoolpay_receipt_number} reason={reason!r} "
                f"by={actor} request_id={change.id}"
            ),
            request,
        )
        return Response(
            {
                "message": (
                    "Bank payment deleted."
                    if applied
                    else "Delete submitted for Bursar / Finance Manager approval."
                ),
                "pending": not applied,
                "request": serialize_change_request(change),
            },
            status=status.HTTP_201_CREATED,
        )


class CanViewOrApproveManualBankRequests(BasePermission):
    message = "Only Super Admin, Bursar, or Finance Manager can view bank payment approval requests."

    def has_permission(self, request, view):
        from payments.manual_bank_approval import user_can_approve_manual_bank

        u = request.user
        if not u or not u.is_authenticated:
            return False
        return user_is_super_admin(u) or user_can_approve_manual_bank(u)


class AdminManualBankChangeRequestListView(APIView):
    """List pending/recent bank payment change requests."""

    permission_classes = [CanViewOrApproveManualBankRequests]

    def get(self, request):
        from payments.manual_bank_approval import (
            serialize_change_request,
            user_can_approve_manual_bank,
        )
        from payments.models import ManualBankPaymentChangeRequest

        status_filter = (request.query_params.get("status") or "pending").strip().lower()
        qs = ManualBankPaymentChangeRequest.objects.select_related(
            "student",
            "student__application",
            "ledger",
            "requested_by",
            "reviewed_by",
        )
        if status_filter and status_filter != "all":
            qs = qs.filter(status=status_filter)
        qs = qs.order_by("-requested_at")[:100]
        return Response(
            {
                "results": [serialize_change_request(r) for r in qs],
                "can_approve": user_can_approve_manual_bank(request.user),
            }
        )


class AdminManualBankChangeRequestApproveView(APIView):
    permission_classes = [CanApproveManualBankPayment]

    def post(self, request, request_id):
        from audit.utils import log_audit_event
        from payments.manual_bank_approval import apply_change_request, serialize_change_request
        from payments.models import ManualBankPaymentChangeRequest
        from payments.student_portal_finance import payment_status_dict

        change = get_object_or_404(
            ManualBankPaymentChangeRequest.objects.select_related(
                "student", "student__application", "ledger", "requested_by"
            ),
            pk=request_id,
        )
        try:
            change, result = apply_change_request(
                change,
                reviewed_by=request.user,
                review_notes=request.data.get("review_notes") or "",
            )
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        actor = request.user.get_full_name() or request.user.email or str(request.user.pk)
        log_audit_event(
            request.user,
            "manual_bank_payment_approved",
            change.student,
            (
                f"Approved {change.request_type} request_id={change.id} "
                f"reason={change.reason!r} by={actor}"
            ),
            request,
        )
        finance = (
            payment_status_dict(change.student, request) if change.student_id else None
        )
        return Response(
            {
                "message": f"{change.get_request_type_display()} approved and applied.",
                "request": serialize_change_request(change),
                "finance": finance,
            }
        )


class AdminManualBankChangeRequestRejectView(APIView):
    permission_classes = [CanApproveManualBankPayment]

    def post(self, request, request_id):
        from audit.utils import log_audit_event
        from payments.manual_bank_approval import reject_change_request, serialize_change_request
        from payments.models import ManualBankPaymentChangeRequest

        change = get_object_or_404(
            ManualBankPaymentChangeRequest.objects.select_related(
                "student", "student__application", "requested_by"
            ),
            pk=request_id,
        )
        try:
            change = reject_change_request(
                change,
                reviewed_by=request.user,
                review_notes=request.data.get("review_notes")
                or request.data.get("reason")
                or "",
            )
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        actor = request.user.get_full_name() or request.user.email or str(request.user.pk)
        log_audit_event(
            request.user,
            "manual_bank_payment_rejected",
            change.student,
            (
                f"Rejected {change.request_type} request_id={change.id} "
                f"by={actor} notes={change.review_notes!r}"
            ),
            request,
        )
        return Response(
            {
                "message": f"{change.get_request_type_display()} rejected.",
                "request": serialize_change_request(change),
            }
        )


class AdminManualBankPaymentsReportView(APIView):
    """GET list of staff-posted bank reconciliation payments."""

    permission_classes = [CanViewOrApproveManualBankRequests]

    def get(self, request):
        page = _parse_page(request.query_params.get("page"))
        page_size = _parse_page_size(request.query_params.get("page_size"), default=50)
        search = (request.query_params.get("search") or "").strip()
        from_date = parse_date(request.query_params.get("from_date") or "")
        to_date = parse_date(request.query_params.get("to_date") or "")

        qs = (
            TuitionLedger.objects.filter(
                Q(schoolpay_receipt_number__startswith="BANK-")
                | Q(raw_response__source="manual_bank_reconciliation")
            )
            .select_related("student", "student__application", "user")
            .order_by("-payment_date_time", "-id")
        )
        if search:
            qs = qs.filter(
                Q(student_name__icontains=search)
                | Q(student_payment_code__icontains=search)
                | Q(student_registration_number__icontains=search)
                | Q(schoolpay_receipt_number__icontains=search)
                | Q(source_channel_transaction_id__icontains=search)
                | Q(user__email__icontains=search)
                | Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
            )
        if from_date:
            qs = qs.filter(payment_date_time__date__gte=from_date)
        if to_date:
            qs = qs.filter(payment_date_time__date__lte=to_date)

        total = qs.count()
        offset = (page - 1) * page_size
        rows = []
        for row in qs[offset : offset + page_size]:
            poster = ""
            if row.user_id:
                poster = row.user.get_full_name() or row.user.email or str(row.user_id)
            raw = row.raw_response if isinstance(row.raw_response, dict) else {}
            rows.append(
                {
                    "id": row.id,
                    "student_pk": row.student_id,
                    "student_name": row.student_name,
                    "student_id": row.student_payment_code,
                    "reg_no": row.student_registration_number,
                    "amount": str(row.amount),
                    "receipt": row.schoolpay_receipt_number,
                    "bank_reference": row.source_channel_transaction_id
                    or raw.get("bank_reference")
                    or "",
                    "bank_name": row.settlement_bank_code or raw.get("bank_name") or "",
                    "channel": row.source_payment_channel,
                    "payment_date_time": row.payment_date_time.isoformat()
                    if row.payment_date_time
                    else None,
                    "posted_by": poster,
                    "posted_by_id": row.user_id,
                    "notes": (raw.get("notes") or row.source_channel_trans_detail or "")[:500],
                    "created_at": row.created_at.isoformat() if row.created_at else None,
                }
            )

        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "results": rows,
            }
        )


class AdminAccountsClearedRegistrationReportView(APIView):
    """
    GET students cleared by Accounts for registration.
    Finance / Super Admin can view. Add export=xlsx for a full Excel download.
    """

    permission_classes = [AccountsClearedReportPermission]

    def _queryset(self, request):
        search = (request.query_params.get("search") or "").strip()
        from_date = parse_date(request.query_params.get("from_date") or "")
        to_date = parse_date(request.query_params.get("to_date") or "")
        cleared_by_id = request.query_params.get("cleared_by")

        qs = (
            AdmittedStudent.objects.filter(
                is_admitted=True,
                accounts_registration_cleared=True,
            )
            .select_related(
                "application",
                "admitted_program",
                "admitted_campus",
                "admitted_batch",
                "accounts_registration_cleared_by",
            )
            .order_by("-accounts_registration_cleared_at", "-id")
        )
        if search:
            qs = qs.filter(
                Q(student_id__icontains=search)
                | Q(reg_no__icontains=search)
                | Q(application__first_name__icontains=search)
                | Q(application__last_name__icontains=search)
                | Q(admitted_program__name__icontains=search)
            )
        if from_date:
            qs = qs.filter(accounts_registration_cleared_at__date__gte=from_date)
        if to_date:
            qs = qs.filter(accounts_registration_cleared_at__date__lte=to_date)
        if cleared_by_id:
            try:
                qs = qs.filter(accounts_registration_cleared_by_id=int(cleared_by_id))
            except (TypeError, ValueError):
                pass
        return qs

    def _row(self, s: AdmittedStudent) -> dict:
        app = s.application
        name = ""
        if app:
            name = f"{app.first_name or ''} {app.last_name or ''}".strip()
        clearer = s.accounts_registration_cleared_by
        clearer_name = ""
        if clearer:
            clearer_name = clearer.get_full_name() or clearer.email or str(clearer.pk)
        return {
            "id": s.id,
            "student_id": s.student_id or "",
            "reg_no": s.reg_no or "",
            "student_name": name or "—",
            "programme": s.admitted_program.name if s.admitted_program_id else "",
            "campus": s.admitted_campus.name if s.admitted_campus_id else "",
            "intake": s.admitted_batch.name if s.admitted_batch_id else "",
            "cleared_at": s.accounts_registration_cleared_at.isoformat()
            if s.accounts_registration_cleared_at
            else None,
            "cleared_by": clearer_name,
            "cleared_by_id": s.accounts_registration_cleared_by_id,
            "notes": (s.accounts_registration_clearance_notes or "")[:500],
        }

    def get(self, request):
        qs = self._queryset(request)
        export = (request.query_params.get("export") or "").strip().lower()
        if export in ("xlsx", "excel", "1", "true"):
            return self._excel(qs)

        page = _parse_page(request.query_params.get("page"))
        page_size = _parse_page_size(request.query_params.get("page_size"), default=50)
        total = qs.count()
        offset = (page - 1) * page_size
        rows = [self._row(s) for s in qs[offset : offset + page_size]]
        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "results": rows,
            }
        )

    def _excel(self, qs):
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        headers = [
            "Cleared at",
            "Student",
            "Student ID",
            "Reg no",
            "Programme",
            "Campus",
            "Intake",
            "Cleared by",
            "Notes",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts cleared"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e1b4b", fill_type="solid")
        thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[get_column_letter(col)].width = max(14, min(36, len(header) + 4))
        for r_i, s in enumerate(qs.iterator(chunk_size=500), 2):
            row = self._row(s)
            values = [
                (row["cleared_at"] or "")[:19].replace("T", " "),
                row["student_name"],
                row["student_id"],
                row["reg_no"],
                row["programme"],
                row["campus"],
                row["intake"],
                row["cleared_by"],
                row["notes"],
            ]
            for c_i, value in enumerate(values, 1):
                cell = ws.cell(row=r_i, column=c_i, value=value)
                cell.border = thin
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        buf = BytesIO()
        wb.save(buf)
        filename = f"accounts_cleared_registration_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class AdminTuitionLedgerTransactionsView(APIView):
    """GET /api/payments/admin/tuition_ledger/transactions"""

    permission_classes = [FinanceModuleAdminPermission]

    def get(self, request):
        page = _parse_page(request.query_params.get("page"))
        page_size = _parse_page_size(request.query_params.get("page_size"))
        search = request.query_params.get("search", "")
        status = (request.query_params.get("status") or "").strip().lower()
        source = (request.query_params.get("source") or "").strip().lower()
        from_date = parse_date(request.query_params.get("from_date") or "")
        to_date = parse_date(request.query_params.get("to_date") or "")
        cohort = _ledger_cohort_params(request)

        qs = (
            StudentTuitionPayment.objects.select_related(
                "student",
                "student__admitted_program",
                "student__admitted_batch",
                "student__application",
                "fee_head",
                "fee_plan_rule__fee_head",
            )
            .filter(_transaction_search_filter(search))
            .order_by("-paid_at", "-created_at")
        )
        qs = _apply_transaction_cohort_filters(qs, cohort)

        if status:
            qs = qs.filter(status=status)
        if source:
            qs = qs.filter(source=source)
        if from_date:
            qs = qs.filter(
                Q(paid_at__date__gte=from_date)
                | Q(paid_at__isnull=True, created_at__date__gte=from_date)
            )
        if to_date:
            qs = qs.filter(
                Q(paid_at__date__lte=to_date)
                | Q(paid_at__isnull=True, created_at__date__lte=to_date)
            )

        total = qs.count()
        offset = (page - 1) * page_size
        rows = [_transaction_row(payment) for payment in qs[offset : offset + page_size]]

        return Response(
            {
                "filters": cohort,
                "results": rows,
                "page": page,
                "page_size": page_size,
                "total": total,
            }
        )


class SendCommitmentFeeReminderView(APIView):
    """
    POST /api/payments/admin/tuition_ledger/send_commitment_reminders

    Quickly counts unpaid-commitment students and queues a Celery background job.
    Returns immediately so nginx/gunicorn do not time out on large cohorts.
    """

    permission_classes = [FinanceModuleAdminPermission]

    def post(self, request):
        params = request.query_params
        data = request.data if hasattr(request, "data") else {}

        def _param(key: str):
            if key in params and params.get(key) not in (None, ""):
                return params.get(key)
            return data.get(key) if isinstance(data, dict) else None

        cohort = {
            "batch_id": _parse_int(_param("batch_id") or _param("intake_id")),
            "program_id": _parse_int(_param("program_id")),
            "campus_id": _parse_int(_param("campus_id")),
            "program_batch_id": _parse_int(_param("program_batch_id")),
            "academic_year": (str(_param("academic_year") or "").strip() or None),
            "intake": (str(_param("intake") or "").strip() or None),
        }

        try:
            result = queue_bulk_commitment_reminders(cohort)
        except Exception as exc:
            logger.exception("Failed to queue commitment reminders")
            return Response(
                {
                    "detail": (
                        f"Failed to queue reminders. Ensure Celery/Redis is running. ({exc})"
                    )
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )

        queued = int(result.get("queued") or 0)
        detail = result.get("detail") or (
            f"{queued} payment reminder(s) queued for background delivery."
        )

        return Response(
            {
                "detail": detail,
                "status": result.get("status") or "queued",
                "queued": queued,
                "sent": queued,
                "failed": int(result.get("failed") or 0),
                "eligible": int(result.get("eligible") or queued),
                "skipped_met": int(result.get("skipped_met") or 0),
                "skipped_no_email": int(result.get("skipped_no_email") or 0),
                "commitment_threshold": float(
                    result.get("commitment_threshold") or COMMITMENT_FEE_THRESHOLD
                ),
                "filters": result.get("filters") or cohort,
                "task_id": result.get("task_id"),
            },
            status=status.HTTP_202_ACCEPTED,
        )
