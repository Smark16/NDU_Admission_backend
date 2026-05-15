#!/usr/bin/env python3
"""Convert terminal CSV export (e.g. from manage.py shell) to a formatted .xlsx file."""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)


def _collect_csv_block(text_lines: list[str]) -> list[str]:
    """Keep header + data rows from a pasted terminal export."""
    out: list[str] = []
    in_block = False
    for raw in text_lines:
        line = raw.strip()
        if not line or line.startswith("<"):
            continue
        if line.startswith("# TOTAL ROWS"):
            break
        if line.startswith("application_id,"):
            out = [line]
            in_block = True
            continue
        if in_block and line[0].isdigit():
            out.append(line)
    return out


def extract_csv_lines_from_jsonl(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    for line in text.splitlines():
        if not line.strip():
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if obj.get("role") != "user":
            continue
        content = obj.get("message", {}).get("content", "")
        if isinstance(content, list):
            for part in content:
                if part.get("type") == "text" and "application_id,first_name" in part.get(
                    "text", ""
                ):
                    block = _collect_csv_block(part["text"].splitlines())
                    if block:
                        return block
    return []


def parse_csv_rows(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    if not lines:
        return [], []
    reader = csv.reader(io.StringIO("\n".join(lines)))
    fieldnames = next(reader)
    rows = list(reader)
    return fieldnames, rows


def write_xlsx(path: Path, fieldnames: list[str], rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Affected applicants"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 14,
        2: 16,
        3: 16,
        4: 28,
        5: 16,
        6: 14,
        7: 18,
        8: 32,
        9: 14,
        10: 36,
        11: 55,
        12: 12,
    }
    for col, width in widths.items():
        if col <= len(fieldnames):
            ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        help="CSV file (UTF-8). If omitted, reads from stdin.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("affected_applicants_april_may_2026.xlsx"),
    )
    parser.add_argument(
        "--from-transcript",
        type=Path,
        help="Extract CSV from a Cursor agent transcript .jsonl file",
    )
    args = parser.parse_args()

    if args.from_transcript:
        lines = extract_csv_lines_from_jsonl(args.from_transcript)
        if not lines:
            print("No CSV lines found in transcript.", file=sys.stderr)
            raise SystemExit(1)
    elif args.input and args.input.exists():
        raw = args.input.read_text(encoding="utf-8", errors="replace")
        lines = _collect_csv_block(raw.splitlines())
    else:
        raw = sys.stdin.read()
        lines = _collect_csv_block(raw.splitlines())

    if not lines:
        print("No CSV input.", file=sys.stderr)
        raise SystemExit(1)

    fieldnames, rows = parse_csv_rows(lines)
    write_xlsx(args.output, fieldnames, rows)
    print(f"Wrote {len(rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
