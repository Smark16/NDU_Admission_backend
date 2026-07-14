"""Programme enrollment report — list + Excel export for faculty admins."""
from __future__ import annotations

from django.db.models import Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from admissions.faculty_scope import filter_programme_enrollments_for_user
from Programs.models import StudentProgrammeEnrollment
from Programs.permissions import AcademicEnrollmentAdminPermission


def _parse_enrollment_report_params(request):
    qp = request.query_params
    return {
        "program": (qp.get("program") or "").strip(),
        "program_batch": (qp.get("program_batch") or qp.get("batch") or "").strip(),
        "faculty": (qp.get("faculty") or "").strip(),
        "campus": (qp.get("campus") or "").strip(),
        "status": (qp.get("status") or "").strip(),
        "year": (qp.get("year") or qp.get("current_year_of_study") or "").strip(),
        "term": (qp.get("term") or qp.get("current_term_number") or "").strip(),
        "academic_year": (qp.get("academic_year") or "").strip(),
        "search": (qp.get("search") or "").strip(),
    }


def _parse_pagination(request):
    try:
        page = max(1, int(request.query_params.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.query_params.get("page_size", 50))))
    except (TypeError, ValueError):
        page_size = 50
    return page, page_size


def _enrollment_report_queryset(params, user):
    qs = (
        StudentProgrammeEnrollment.objects.select_related(
            "student",
            "student__application",
            "student__admitted_campus",
            "program",
            "program__faculty",
            "program_batch",
            "curriculum_version",
            "enrolled_by",
        )
        .order_by(
            "program__faculty__name",
            "program__name",
            "program_batch__name",
            "student__reg_no",
        )
    )

    if params["program"]:
        qs = qs.filter(program_id=params["program"])
    if params["program_batch"]:
        qs = qs.filter(program_batch_id=params["program_batch"])
    if params["faculty"]:
        qs = qs.filter(program__faculty_id=params["faculty"])
    if params["campus"]:
        qs = qs.filter(student__admitted_campus_id=params["campus"])
    if params["status"]:
        qs = qs.filter(status=params["status"])
    if params["year"]:
        qs = qs.filter(current_year_of_study=params["year"])
    if params["term"]:
        qs = qs.filter(current_term_number=params["term"])
    if params["academic_year"]:
        qs = qs.filter(program_batch__academic_year=params["academic_year"])
    if params["search"]:
        term = params["search"]
        qs = qs.filter(
            Q(student__student_id__icontains=term)
            | Q(student__reg_no__icontains=term)
            | Q(student__application__first_name__icontains=term)
            | Q(student__application__last_name__icontains=term)
            | Q(student__application__middle_name__icontains=term)
            | Q(student__application__email__icontains=term)
        )

    return filter_programme_enrollments_for_user(qs, user)


def _status_counts_for_queryset(qs) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in qs.values("status").annotate(c=Count("id")):
        key = row["status"] or "unknown"
        counts[key] = row["c"]
    return counts


def _enrollment_report_row(enrollment: StudentProgrammeEnrollment) -> dict:
    student = enrollment.student
    app = getattr(student, "application", None)
    program = enrollment.program
    faculty = getattr(program, "faculty", None)
    campus = getattr(student, "admitted_campus", None)
    batch = enrollment.program_batch
    curriculum = enrollment.curriculum_version

    name = ""
    gender = ""
    email = ""
    phone = ""
    if app is not None:
        parts = [app.first_name or "", app.middle_name or "", app.last_name or ""]
        name = " ".join(p.strip() for p in parts if p and p.strip())
        gender = app.gender or ""
        email = app.email or ""
        phone = app.phone or ""

    enrolled_by = ""
    if enrollment.enrolled_by_id:
        enrolled_by = (
            enrollment.enrolled_by.get_full_name() or enrollment.enrolled_by.email or ""
        )

    return {
        "id": enrollment.id,
        "student_name": name,
        "gender": gender,
        "student_id": student.student_id or "",
        "reg_no": student.reg_no or "",
        "email": email,
        "phone": phone,
        "program_code": program.code or "",
        "program_name": program.name or "",
        "program_short": program.short_form or "",
        "faculty": faculty.name if faculty else "",
        "campus": campus.name if campus else "",
        "study_mode": student.study_mode or "",
        "academic_batch": batch.name if batch else "",
        "academic_year": batch.academic_year if batch else "",
        "year_of_study": enrollment.current_year_of_study,
        "term": enrollment.current_term_number,
        "specialization": enrollment.specialization or "",
        "status": enrollment.status,
        "status_display": enrollment.get_status_display(),
        "curriculum_version": curriculum.name if curriculum else "",
        "enrolled_at": (
            enrollment.enrolled_at.strftime("%Y-%m-%d %H:%M")
            if enrollment.enrolled_at
            else ""
        ),
        "enrolled_by": enrolled_by,
        "notes": (enrollment.notes or "").strip(),
    }


def _enrollment_report_excel_row(row: dict) -> list:
    return [
        row["student_name"],
        row["gender"],
        row["student_id"],
        row["reg_no"],
        row["program_code"],
        row["program_name"],
        row["faculty"],
        row["campus"],
        row["study_mode"],
        row["academic_batch"],
        row["academic_year"],
        row["year_of_study"],
        row["term"],
        row["specialization"],
        row["status_display"],
        row["curriculum_version"],
        row["enrolled_at"],
        row["enrolled_by"],
        row["email"],
        row["phone"],
        row["notes"],
    ]


def _filter_summary_blurb(params) -> str:
    parts = []
    if params["academic_year"]:
        parts.append(f"Academic year {params['academic_year']}")
    if params["status"]:
        parts.append(f"Status: {params['status']}")
    if params["year"]:
        parts.append(f"Year of study {params['year']}")
    if params["term"]:
        parts.append(f"Term {params['term']}")
    if params["search"]:
        parts.append(f'Search: "{params["search"]}"')
    return " · ".join(parts) if parts else "All programme enrollments (within your faculty scope)"


EXCEL_HEADERS = [
    "STUDENT NAME",
    "GENDER",
    "STUDENT ID",
    "REG NO",
    "PROGRAM CODE",
    "PROGRAMME",
    "FACULTY",
    "CAMPUS",
    "STUDY MODE",
    "ACADEMIC BATCH",
    "ACADEMIC YEAR",
    "YEAR",
    "TERM",
    "SPECIALIZATION",
    "STATUS",
    "CURRICULUM",
    "ENROLLED AT",
    "ENROLLED BY",
    "EMAIL",
    "PHONE",
    "NOTES",
]


class EnrollmentReportListView(APIView):
    """JSON programme enrollment report for on-screen display."""

    permission_classes = [IsAuthenticated, AcademicEnrollmentAdminPermission]

    def get(self, request):
        params = _parse_enrollment_report_params(request)
        page, page_size = _parse_pagination(request)
        qs = _enrollment_report_queryset(params, request.user)

        total = qs.count()
        status_counts = _status_counts_for_queryset(qs)
        offset = (page - 1) * page_size
        enrollments = qs[offset : offset + page_size]
        results = [_enrollment_report_row(e) for e in enrollments]

        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size if page_size else 0,
                "filter_summary": _filter_summary_blurb(params),
                "status_counts": status_counts,
                "results": results,
            },
            status=200,
        )


class EnrollmentReportExcelView(APIView):
    """Excel export for programme enrollment report."""

    permission_classes = [IsAuthenticated, AcademicEnrollmentAdminPermission]

    def get(self, request):
        params = _parse_enrollment_report_params(request)
        qs = _enrollment_report_queryset(params, request.user)

        n_cols = len(EXCEL_HEADERS)
        wb = Workbook()
        ws = wb.active
        ws.title = "Programme enrollments"

        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=11)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", fill_type="solid")
        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        t1 = ws.cell(row=1, column=1, value="STUDENT PROGRAMME ENROLLMENT REPORT")
        t1.font = title_font
        t1.alignment = Alignment(horizontal="center", vertical="center")

        total = qs.count()
        sub = _filter_summary_blurb(params)
        if total:
            sub = f"{sub} · {total} record(s)"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        t2 = ws.cell(row=2, column=1, value=sub)
        t2.font = subtitle_font
        t2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_row = 4
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                12, min(28, len(str(header)) + 2)
            )

        for enrollment in qs.iterator(chunk_size=250):
            ws.append(_enrollment_report_excel_row(_enrollment_report_row(enrollment)))

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        fname = "programme_enrollment_report.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response
