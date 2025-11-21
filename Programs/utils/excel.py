# utils/excel.py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Optional, Dict, Any


def create_workbook(
    headers: List[str],
    rows: List[List[Any]] = None,
    sheet_name: str = "Sheet1",
    header_bg: str = "4F81BD",
    dropdowns: Optional[Dict[int, List[str]]] = None,  # column_index: list of options
    freeze_header: bool = True,
    instructions: str = None,
) -> Workbook:
    """
    Creates a styled Excel workbook.
    
    Args:
        headers: List of column headers
        rows: Data rows (optional)
        sheet_name: Name of the sheet
        header_bg: Hex color for header background
        dropdowns: {column_index: ["Option1", "Option2", ...]} for dropdowns (1-indexed!)
        freeze_header: Freeze first row
        instructions: Optional text to show in row 1 (merged)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # === Styles ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    # === Optional: Instructions Row ===
    start_row = 2
    if instructions:
        ws.append([instructions])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1).font = Font(color="FF0000", italic=True, size=12)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        start_row = 2

    # === Write Headers ===
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

        # Auto column width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(header)) + 4)

    # === Write Data Rows ===
    if rows:
        for row in rows:
            ws.append(row)

    # === Apply borders & alignment to data area ===
    data_min_row = start_row + 1
    data_max_row = ws.max_row
    for row in ws.iter_rows(min_row=data_min_row, max_row=data_max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # === Add Dropdowns (Data Validation) ===
    if dropdowns:
        for col_idx, options in dropdowns.items():
            if not options:
                continue
            formula = f'"{",".join(map(str, options))}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(f"{get_column_letter(col_idx)}{data_min_row}:{get_column_letter(col_idx)}1000")
            ws.add_data_validation(dv)

    # === Freeze header ===
    if freeze_header:
        ws.freeze_panes = ws[f"A{start_row + 1}"]  # Freeze below header

    return wb