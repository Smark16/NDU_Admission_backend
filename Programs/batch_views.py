import calendar as _calendar
import csv
from datetime import date, datetime, timedelta
import io

from django.db import transaction
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.db.utils import ProgrammingError
from django.http import HttpResponse
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from .permissions import ProgramSchedulingAPIPermission
from rest_framework.response import Response
from rest_framework.views import APIView

from .batch_offer_defaults import resolve_program_batch_offer_dates
from .models import (
    CourseCatalogUnit,
    CourseUnit,
    Program,
    ProgramBatch,
    ProgramCurriculumVersion,
    Semester,
    resolve_program_default_curriculum_version,
)
from .utils.excel import create_workbook


class _BatchUnavailableMixin:
    """Returns a clear 503 when the ProgramBatch table has not been migrated yet."""
    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except ProgrammingError:
            return Response(
                {'detail': 'Batch management is not available on this server.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )


class CreateBatchView(_BatchUnavailableMixin, APIView):
    """Create a program batch (academic year level)."""
    permission_classes = [ProgramSchedulingAPIPermission]

    def post(self, request, program_id):
        try:
            try:
                program = Program.objects.get(id=program_id)
            except Program.DoesNotExist:
                return Response(
                    {'detail': f'Program with id {program_id} not found'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            name = request.data.get('name', '').strip()
            academic_year = request.data.get('academic_year', '').strip()
            start_date = request.data.get('start_date', '').strip()
            end_date = request.data.get('end_date', '').strip()
            curriculum_version_id = request.data.get('curriculum_version')

            if not name:
                return Response({'detail': 'Batch name is required'}, status=status.HTTP_400_BAD_REQUEST)
            if not academic_year:
                return Response(
                    {'detail': 'Academic year is required (e.g. 2024/2025)'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not start_date:
                return Response({'detail': 'Start date is required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = None
                if end_date:
                    end = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if end < start:
                        return Response(
                            {'detail': 'End date must be after start date'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
            except ValueError as e:
                return Response(
                    {'detail': f'Invalid date format: {str(e)}. Use YYYY-MM-DD format.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            offer_start_s = (request.data.get('offer_start_date') or '').strip()
            offer_end_s = (request.data.get('offer_end_date') or '').strip()
            offer_start = None
            offer_end = None
            if offer_start_s or offer_end_s:
                if not offer_start_s or not offer_end_s:
                    return Response(
                        {
                            'detail': (
                                'offer_start_date and offer_end_date must both be set, '
                                'or both left empty to use the admission intake offer window.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                try:
                    offer_start = datetime.strptime(offer_start_s, '%Y-%m-%d').date()
                    offer_end = datetime.strptime(offer_end_s, '%Y-%m-%d').date()
                    if offer_end < offer_start:
                        return Response(
                            {'detail': 'Offer end date must be on or after offer start date'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                except ValueError as e:
                    return Response(
                        {'detail': f'Invalid offer date format: {str(e)}. Use YYYY-MM-DD format.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            offer_start, offer_end = resolve_program_batch_offer_dates(
                start_date=start,
                end_date=end,
                offer_start_date=offer_start,
                offer_end_date=offer_end,
            )

            if ProgramBatch.objects.filter(program=program, name=name).exists():
                return Response(
                    {'detail': f'Batch with name "{name}" already exists for this program'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            curriculum_version = None
            if curriculum_version_id:
                try:
                    curriculum_version = ProgramCurriculumVersion.objects.get(
                        pk=int(curriculum_version_id), program=program
                    )
                except (ValueError, ProgramCurriculumVersion.DoesNotExist):
                    return Response(
                        {'detail': 'curriculum_version is invalid for this programme.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            else:
                curriculum_version = resolve_program_default_curriculum_version(program)

            try:
                with transaction.atomic():
                    batch = ProgramBatch.objects.create(
                        program=program,
                        curriculum_version=curriculum_version,
                        name=name,
                        academic_year=academic_year,
                        start_date=start,
                        end_date=end,
                        offer_start_date=offer_start,
                        offer_end_date=offer_end,
                        is_active=True,
                    )
            except Exception as e:
                return Response(
                    {'detail': f'Error creating batch: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # Same behaviour as batch bulk upload: generate term rows from programme
            # calendar_type / min_years. Separate transaction so the batch is kept if
            # semester generation fails.
            semesters_created = 0
            semester_generation_warning = None
            try:
                with transaction.atomic():
                    semesters_created = _auto_create_semesters(batch, program)
            except Exception as exc:
                semester_generation_warning = (
                    f'Batch was created but semesters could not be auto-generated: {exc}'
                )

            payload = {
                'id': batch.id,
                'name': batch.name,
                'academic_year': batch.academic_year,
                'program_id': program.id,
                'program_name': program.name,
                'curriculum_version': {
                    'id': curriculum_version.id if curriculum_version else None,
                    'name': curriculum_version.name if curriculum_version else None,
                },
                'start_date': batch.start_date.isoformat(),
                'end_date': batch.end_date.isoformat() if batch.end_date else None,
                'offer_start_date': batch.offer_start_date.isoformat() if batch.offer_start_date else None,
                'offer_end_date': batch.offer_end_date.isoformat() if batch.offer_end_date else None,
                'is_offer_active': batch.is_offer_active,
                'semesters_created': semesters_created,
                'message': 'Batch created successfully',
            }
            if semester_generation_warning:
                payload['semester_generation_warning'] = semester_generation_warning

            return Response(payload, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CreateSemesterView(_BatchUnavailableMixin, APIView):
    """Create a semester for a program batch."""
    permission_classes = [ProgramSchedulingAPIPermission]

    def post(self, request, batch_id):
        try:
            try:
                batch = ProgramBatch.objects.select_related('program').get(id=batch_id, is_active=True)
            except ProgramBatch.DoesNotExist:
                return Response({'detail': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)

            name = request.data.get('name', '').strip()
            start_date = request.data.get('start_date', '').strip()
            end_date = request.data.get('end_date', '').strip()
            order = request.data.get('order', 1)
            year_of_study_raw = request.data.get('year_of_study')
            term_number_raw = request.data.get('term_number')

            if not name:
                return Response({'detail': 'Semester name is required'}, status=status.HTTP_400_BAD_REQUEST)
            if not start_date:
                return Response({'detail': 'Start date is required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = None
                if end_date:
                    end = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if end < start:
                        return Response(
                            {'detail': 'End date must be after start date'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
            except ValueError as e:
                return Response(
                    {'detail': f'Invalid date format: {str(e)}. Use YYYY-MM-DD format.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                order = int(order)
                if order < 1:
                    raise ValueError('Order must be positive')
            except (ValueError, TypeError):
                return Response({'detail': 'Order must be a positive integer'}, status=status.HTTP_400_BAD_REQUEST)

            max_allowed = int(batch.program.min_years or 0) * int(batch.program.max_terms_per_year or 0)
            if max_allowed > 0 and order > max_allowed:
                return Response(
                    {
                        'detail': (
                            f'Cannot create term {order}. This programme allows at most '
                            f'{max_allowed} semesters/terms ({batch.program.min_years} year(s) x '
                            f'{batch.program.max_terms_per_year} term(s) per year).'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if Semester.objects.filter(program_batch=batch, order=order).exists():
                return Response(
                    {'detail': f'Semester with order {order} already exists for this batch'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # --- curriculum position (optional) ---
            year_of_study = None
            term_number = None
            if year_of_study_raw is not None or term_number_raw is not None:
                try:
                    year_of_study = int(year_of_study_raw)
                    term_number = int(term_number_raw)
                except (ValueError, TypeError):
                    return Response(
                        {'detail': 'year_of_study and term_number must both be integers when provided.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                program = batch.program
                if year_of_study < 1 or year_of_study > program.max_years:
                    return Response(
                        {'detail': f'year_of_study must be between 1 and {program.max_years}.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                max_terms = program.max_terms_per_year
                if term_number not in range(1, max_terms + 1):
                    return Response(
                        {'detail': f'term_number must be between 1 and {max_terms} for a {program.calendar_type}-based programme.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if Semester.objects.filter(
                    program_batch=batch,
                    year_of_study=year_of_study,
                    term_number=term_number,
                ).exists():
                    return Response(
                        {'detail': f'A semester at Year {year_of_study} Term {term_number} already exists in this batch.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            try:
                with transaction.atomic():
                    semester = Semester.objects.create(
                        program_batch=batch,
                        name=name,
                        start_date=start,
                        end_date=end,
                        order=order,
                        year_of_study=year_of_study,
                        term_number=term_number,
                        is_active=True,
                    )

                    return Response(
                        {
                            'id': semester.id,
                            'name': semester.name,
                            'batch_id': batch.id,
                            'batch_name': batch.name,
                            'order': semester.order,
                            'year_of_study': semester.year_of_study,
                            'term_number': semester.term_number,
                            'start_date': semester.start_date.isoformat(),
                            'end_date': semester.end_date.isoformat() if semester.end_date else None,
                            'message': 'Semester created successfully',
                        },
                        status=status.HTTP_201_CREATED,
                    )

            except Exception as e:
                return Response(
                    {'detail': f'Error creating semester: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CreateSubjectView(_BatchUnavailableMixin, APIView):
    """Create a course unit for a program batch/semester."""
    permission_classes = [ProgramSchedulingAPIPermission]

    def post(self, request, batch_id):
        try:
            try:
                batch = ProgramBatch.objects.select_related('program').get(id=batch_id, is_active=True)
            except ProgramBatch.DoesNotExist:
                return Response({'detail': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)

            semester_id = request.data.get('semester')
            course_unit_id = request.data.get('course_unit_id')
            name = ''
            code = ''
            credit_units = None
            catalog_source = None

            if course_unit_id is not None and course_unit_id != '':
                try:
                    catalog_source = CourseCatalogUnit.objects.get(pk=int(course_unit_id))
                except (CourseCatalogUnit.DoesNotExist, ValueError, TypeError):
                    return Response({'detail': 'Catalog course unit not found'}, status=status.HTTP_404_NOT_FOUND)
                name = catalog_source.title.strip()
                code = catalog_source.code.strip()
                credit_units = catalog_source.credit_units
            else:
                name = request.data.get('name', '').strip()
                code = request.data.get('code', '').strip()
                credit_units = request.data.get('credit_units')

            if not name:
                return Response({'detail': 'Course unit name is required'}, status=status.HTTP_400_BAD_REQUEST)
            if not code:
                return Response({'detail': 'Course unit code is required'}, status=status.HTTP_400_BAD_REQUEST)

            semester = None
            if semester_id:
                try:
                    semester = Semester.objects.get(id=semester_id, program_batch=batch, is_active=True)
                except Semester.DoesNotExist:
                    return Response({'detail': 'Semester not found for this batch'}, status=status.HTTP_404_NOT_FOUND)

            if semester:
                if CourseUnit.objects.filter(code=code, semester=semester).exists():
                    return Response(
                        {'detail': f'Course unit with code "{code}" already exists in this semester'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            try:
                with transaction.atomic():
                    course_unit_data = {
                        'name': name,
                        'code': code,
                        'program_batch': batch,
                        'is_active': True,
                    }
                    if catalog_source is not None:
                        course_unit_data['catalog_unit'] = catalog_source

                    if semester:
                        course_unit_data['semester'] = semester

                    if credit_units:
                        try:
                            course_unit_data['credit_units'] = float(credit_units)
                        except (ValueError, TypeError):
                            pass

                    course_unit = CourseUnit.objects.create(**course_unit_data)

                    return Response(
                        {
                            'id': course_unit.id,
                            'name': course_unit.name,
                            'code': course_unit.code,
                            'credit_units': float(course_unit.credit_units) if course_unit.credit_units else None,
                            'batch_id': batch.id,
                            'semester_id': semester.id if semester else None,
                            'semester_name': semester.name if semester else None,
                            'message': 'Course unit created successfully',
                        },
                        status=status.HTTP_201_CREATED,
                    )

            except Exception as e:
                return Response(
                    {'detail': f'Error creating course unit: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ListProgramBatchesView(_BatchUnavailableMixin, APIView):
    """List all program batches for a program."""
    permission_classes = [ProgramSchedulingAPIPermission]

    def get(self, request, program_id):
        try:
            program = Program.objects.get(id=program_id)
            batches = ProgramBatch.objects.filter(program=program).order_by('-start_date', 'name')

            batches_data = []
            for batch in batches:
                sems = batch.semesters.all().order_by('order', 'start_date', 'id')
                batches_data.append(
                    {
                        'id': batch.id,
                        'name': batch.name,
                        'academic_year': batch.academic_year or '',
                        'program_id': program.id,
                        'program_name': program.name,
                        'curriculum_version': {
                            'id': batch.curriculum_version.id if batch.curriculum_version_id else None,
                            'name': batch.curriculum_version.name if batch.curriculum_version_id else None,
                        },
                        'start_date': batch.start_date.isoformat(),
                        'end_date': batch.end_date.isoformat() if batch.end_date else None,
                        'offer_start_date': batch.offer_start_date.isoformat() if batch.offer_start_date else None,
                        'offer_end_date': batch.offer_end_date.isoformat() if batch.offer_end_date else None,
                        'is_offer_active': batch.is_offer_active,
                        'is_active': batch.is_active,
                        'created_at': batch.created_at.isoformat(),
                        'semester_count': sems.count(),
                        'course_unit_count': batch.course_units.count(),
                        'semesters': [
                            {
                                'id': s.id,
                                'name': s.name,
                                'order': s.order,
                                'year_of_study': s.year_of_study,
                                'term_number': s.term_number,
                                'is_curriculum_positioned': s.year_of_study is not None and s.term_number is not None,
                                'start_date': s.start_date.isoformat(),
                                'end_date': s.end_date.isoformat() if s.end_date else None,
                                'is_active': s.is_active,
                            }
                            for s in sems
                        ],
                    }
                )

            return Response(
                {
                    'program': {
                        'id': program.id,
                        'name': program.name,
                        'short_form': program.short_form,
                        'code': program.code,
                    },
                    'batches': batches_data,
                    'total': len(batches_data),
                },
                status=status.HTTP_200_OK,
            )

        except Program.DoesNotExist:
            return Response(
                {'detail': f'Program with id {program_id} not found'},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UpdateProgramBatchView(_BatchUnavailableMixin, APIView):
    """Update a program batch."""
    permission_classes = [ProgramSchedulingAPIPermission]

    def put(self, request, batch_id):
        try:
            try:
                batch = ProgramBatch.objects.select_related('program').get(id=batch_id)
            except ProgramBatch.DoesNotExist:
                return Response({'detail': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)

            name = request.data.get('name', '').strip()
            academic_year = request.data.get('academic_year')
            start_date = request.data.get('start_date', '').strip()
            end_date = request.data.get('end_date', '').strip()
            is_active = request.data.get('is_active')
            curriculum_version_id = request.data.get('curriculum_version')

            if academic_year is not None:
                batch.academic_year = str(academic_year).strip()

            if name:
                if ProgramBatch.objects.filter(program=batch.program, name=name).exclude(id=batch_id).exists():
                    return Response(
                        {'detail': f'Batch with name "{name}" already exists for this program'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                batch.name = name

            if start_date:
                try:
                    batch.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    return Response(
                        {'detail': 'Invalid start date format. Use YYYY-MM-DD format.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if end_date:
                try:
                    parsed_end = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if parsed_end < batch.start_date:
                        return Response(
                            {'detail': 'End date must be after start date'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    batch.end_date = parsed_end
                except ValueError:
                    return Response(
                        {'detail': 'Invalid end date format. Use YYYY-MM-DD format.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if is_active is not None:
                batch.is_active = bool(is_active)

            if curriculum_version_id is not None:
                if str(curriculum_version_id).strip() == "":
                    batch.curriculum_version = resolve_program_default_curriculum_version(batch.program)
                else:
                    try:
                        batch.curriculum_version = ProgramCurriculumVersion.objects.get(
                            pk=int(curriculum_version_id), program=batch.program
                        )
                    except (ValueError, ProgramCurriculumVersion.DoesNotExist):
                        return Response(
                            {'detail': 'curriculum_version is invalid for this programme.'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

            if 'offer_start_date' in request.data or 'offer_end_date' in request.data:
                offer_start_s = (request.data.get('offer_start_date') or '').strip()
                offer_end_s = (request.data.get('offer_end_date') or '').strip()
                if not offer_start_s and not offer_end_s:
                    batch.offer_start_date = None
                    batch.offer_end_date = None
                elif not offer_start_s or not offer_end_s:
                    return Response(
                        {
                            'detail': (
                                'offer_start_date and offer_end_date must both be set, '
                                'or both cleared to inherit from the admission intake.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                else:
                    try:
                        batch.offer_start_date = datetime.strptime(offer_start_s, '%Y-%m-%d').date()
                        batch.offer_end_date = datetime.strptime(offer_end_s, '%Y-%m-%d').date()
                        if batch.offer_end_date < batch.offer_start_date:
                            return Response(
                                {'detail': 'Offer end date must be on or after offer start date'},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                    except ValueError:
                        return Response(
                            {'detail': 'Invalid offer date format. Use YYYY-MM-DD format.'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

            batch.save()

            return Response(
                {
                    'id': batch.id,
                    'name': batch.name,
                    'academic_year': batch.academic_year or '',
                    'program_id': batch.program.id,
                    'program_name': batch.program.name,
                    'curriculum_version': {
                        'id': batch.curriculum_version.id if batch.curriculum_version_id else None,
                        'name': batch.curriculum_version.name if batch.curriculum_version_id else None,
                    },
                    'start_date': batch.start_date.isoformat(),
                    'end_date': batch.end_date.isoformat() if batch.end_date else None,
                    'offer_start_date': batch.offer_start_date.isoformat() if batch.offer_start_date else None,
                    'offer_end_date': batch.offer_end_date.isoformat() if batch.offer_end_date else None,
                    'is_offer_active': batch.is_offer_active,
                    'is_active': batch.is_active,
                    'message': 'Batch updated successfully',
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DeleteProgramBatchView(_BatchUnavailableMixin, APIView):
    """Delete a program batch when no students are actively enrolled in that cohort.

    Auto-generated semesters and course units are removed with the batch (CASCADE).
    Pending/withdrawn programme enrollments on this cohort are removed automatically.
    Blocks when any enrollment has status enrolled, suspended, or completed (PROTECT).
    """
    permission_classes = [ProgramSchedulingAPIPermission]

    _BLOCKING_ENROLLMENT_STATUSES = ('enrolled', 'suspended', 'completed')

    def delete(self, request, batch_id):
        try:
            from Programs.models import StudentProgrammeEnrollment

            try:
                batch = ProgramBatch.objects.select_related('program').get(id=batch_id)
            except ProgramBatch.DoesNotExist:
                return Response({'detail': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)

            status_rows = (
                StudentProgrammeEnrollment.objects.filter(program_batch=batch)
                .values('status')
                .annotate(count=Count('id'))
            )
            status_breakdown = {row['status']: row['count'] for row in status_rows}

            blocking_count = sum(
                status_breakdown.get(s, 0) for s in self._BLOCKING_ENROLLMENT_STATUSES
            )
            if blocking_count > 0:
                enrolled_only = status_breakdown.get('enrolled', 0)
                return Response(
                    {
                        'detail': (
                            f'Cannot delete batch "{batch.name}": {blocking_count} student '
                            f'placement record(s) still use this cohort '
                            f'({enrolled_only} fully enrolled). Move or withdraw them first, '
                            f'or set is_active=false instead of deleting.'
                        ),
                        'enrolled_student_count': enrolled_only,
                        'blocking_placement_count': blocking_count,
                        'enrollment_status_breakdown': status_breakdown,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            batch_name = batch.name
            program_name = batch.program.name
            semester_count = batch.semesters.count()
            course_unit_count = batch.course_units.count()
            intended_count = batch.intended_admissions.count()

            with transaction.atomic():
                cleared, _ = (
                    StudentProgrammeEnrollment.objects.filter(program_batch=batch)
                    .exclude(status__in=self._BLOCKING_ENROLLMENT_STATUSES)
                    .delete()
                )
                try:
                    batch.delete()
                except ProtectedError:
                    remaining = list(
                        StudentProgrammeEnrollment.objects.filter(program_batch=batch)
                        .values_list('status', flat=True)
                    )
                    return Response(
                        {
                            'detail': (
                                f'Cannot delete batch "{batch_name}": linked student records '
                                f'still reference this cohort.'
                            ),
                            'remaining_enrollment_statuses': remaining,
                            'enrollment_status_breakdown': status_breakdown,
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            return Response(
                {
                    'message': f'Batch "{batch_name}" for program "{program_name}" deleted successfully',
                    'semesters_removed': semester_count,
                    'course_units_removed': course_unit_count,
                    'pending_enrollments_cleared': cleared,
                    'intended_admissions_cleared': intended_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {'detail': f'Unexpected error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ---------------------------------------------------------------------------
# Semester auto-generation
# ---------------------------------------------------------------------------

def _add_months(d: date, n: int) -> date:
    """Return d + n calendar months, clamped to the last day of that month."""
    month = d.month + n
    year = d.year + (month - 1) // 12
    month = ((month - 1) % 12) + 1
    day = min(d.day, _calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _auto_create_semesters(batch: "ProgramBatch", program: "Program") -> int:
    """
    Auto-create the full minimum-duration term structure for a new ProgramBatch.

    Term count = program.min_years × terms_per_year, where:
      semester  calendar → 2 terms/year, ~6 months each
      trimester calendar → 3 terms/year, ~4 months each

    Each Semester is named "Year Y Semester T" / "Year Y Trimester T" and has
    year_of_study and term_number populated so they align with the curriculum.

    Dates are generated as sequential month-offsets from batch.start_date so
    they span the full minimum programme duration regardless of batch.end_date.

    Returns the total number of Semester rows created.
    """
    cal_type = getattr(program, "calendar_type", None) or "semester"
    terms_per_year = 3 if cal_type == "trimester" else 2
    label = "Trimester" if cal_type == "trimester" else "Semester"
    months_each = 4 if cal_type == "trimester" else 6

    raw_min = getattr(program, "min_years", None)
    min_years = max(int(raw_min), 1) if raw_min else 1

    start = batch.start_date
    order = 1  # absolute sequence number across all years (unique per batch)

    for year in range(1, min_years + 1):
        for term in range(1, terms_per_year + 1):
            offset = order - 1
            term_start = _add_months(start, offset * months_each)
            term_end   = _add_months(start, order * months_each)

            Semester.objects.create(
                program_batch=batch,
                name=f"Year {year} {label} {term}",
                order=order,
                year_of_study=year,
                term_number=term,
                start_date=term_start,
                end_date=term_end,
            )
            order += 1

    return order - 1  # total created


# ---------------------------------------------------------------------------
# Batch Bulk Upload
# ---------------------------------------------------------------------------

# Template columns
_BATCH_TEMPLATE_HEADERS = [
    "batch_id",         # filled for existing cohorts — do not change when updating
    "program_code",
    "program_name",     # reference only — ignored on upload
    "batch_name",
    "academic_year",
    "start_date",
    "end_date",
    "offer_start_date",
    "offer_end_date",
    "is_active",
]


def _batch_template_date_str(value) -> str:
    if not value:
        return ""
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _batch_row_from_model(batch: ProgramBatch) -> list:
    prog = batch.program
    return [
        str(batch.id),
        prog.code or "",
        prog.name or "",
        batch.name,
        batch.academic_year or "",
        _batch_template_date_str(batch.start_date),
        _batch_template_date_str(batch.end_date),
        _batch_template_date_str(batch.offer_start_date),
        _batch_template_date_str(batch.offer_end_date),
        "TRUE" if batch.is_active else "FALSE",
    ]


def _empty_batch_template_row(prog: Program) -> list:
    return [
        "",
        prog.code or "",
        prog.name or "",
        "",
        "",
        "",
        "",
        "",
        "",
        "TRUE",
    ]


def _batch_template_instructions(*, program_count: int, cohort_count: int, blank_count: int, campus_scope_label: str = "") -> str:
    scope_prefix = f"{campus_scope_label}: " if campus_scope_label else ""
    return (
        f"{scope_prefix}{program_count} programme(s): {cohort_count} existing cohort row(s), "
        f"{blank_count} blank row(s) for programmes without a cohort. "
        "Columns offer_start_date and offer_end_date control the admit dropdown (admission offer window). "
        "Leave both offer columns blank to copy from cohort start_date/end_date. "
        "Update mode: rows with batch_id — edit any dates including offer_start_date and offer_end_date."
    )


def _build_batch_template_rows(all_programs, batches_by_program: dict) -> list:
    sample_rows: list = []
    for prog in sorted(all_programs, key=lambda p: ((p.code or "").lower(), (p.name or "").lower())):
        prog_batches = batches_by_program.get(prog.id) or []
        if prog_batches:
            for batch in sorted(prog_batches, key=lambda b: b.name.lower()):
                sample_rows.append(_batch_row_from_model(batch))
        else:
            sample_rows.append(_empty_batch_template_row(prog))
    return sample_rows


def _finalize_batch_offer_dates(
    *,
    start_date: date,
    end_date: date | None,
    offer_start_date: date | None,
    offer_end_date: date | None,
    has_explicit_offer: bool,
    update_existing: bool,
    existing: ProgramBatch | None,
) -> tuple[date, date]:
    if (
        update_existing
        and existing is not None
        and not has_explicit_offer
        and existing.offer_start_date is not None
        and existing.offer_end_date is not None
    ):
        return existing.offer_start_date, existing.offer_end_date
    return resolve_program_batch_offer_dates(
        start_date=start_date,
        end_date=end_date,
        offer_start_date=offer_start_date,
        offer_end_date=offer_end_date,
    )


def _resolve_existing_program_batch(program, batch_name: str, batch_id_cell: str):
    """Match cohort by batch_id (preferred) or batch_name (case-insensitive)."""
    bid = (batch_id_cell or "").strip()
    if bid:
        try:
            pk = int(float(bid))
        except (TypeError, ValueError):
            return None, "invalid_batch_id"
        batch = ProgramBatch.objects.filter(pk=pk, program=program).first()
        if batch:
            return batch, None
        return None, "batch_id_not_found"

    name = (batch_name or "").strip()
    if not name:
        return None, "missing_batch_name"
    batch = ProgramBatch.objects.filter(program=program, name__iexact=name).first()
    if batch:
        return batch, None
    return None, "not_found"


class BatchTemplateDownloadView(_BatchUnavailableMixin, APIView):
    """GET /api/program/batches/template — downloadable Excel template.

    Query params:
      program_ids — optional comma-separated programme primary keys. When set,
        the workbook contains one sample row (and dropdown options) only for
        those programmes.
      campus_id — optional campus primary key. When set (and program_ids omitted),
        includes all active programmes offered at that campus (e.g. Main Campus).
      When both are omitted, all programmes are included (legacy).
    """
    permission_classes = [ProgramSchedulingAPIPermission]

    def get(self, request):
        raw_ids = (request.GET.get("program_ids") or "").strip()
        raw_campus_id = (request.GET.get("campus_id") or "").strip()
        campus_scope_label = ""

        if not raw_ids:
            if raw_campus_id:
                try:
                    campus_pk = int(raw_campus_id)
                except ValueError:
                    return Response(
                        {'detail': f'Invalid campus_id: {raw_campus_id!r}.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                from accounts.models import Campus

                campus = Campus.objects.filter(pk=campus_pk).first()
                if not campus:
                    return Response(
                        {'detail': f'Campus id {campus_pk} not found.'},
                        status=status.HTTP_404_NOT_FOUND,
                    )
                campus_scope_label = campus.name
                all_programs = list(
                    Program.objects.filter(
                        campuses__id=campus_pk,
                        is_active=True,
                    )
                    .distinct()
                    .order_by('code')
                )
            else:
                all_programs = list(Program.objects.order_by('code'))
        else:
            id_list: list[int] = []
            for part in raw_ids.split(","):
                part = part.strip()
                if not part:
                    continue
                try:
                    id_list.append(int(part))
                except ValueError:
                    return Response(
                        {'detail': f'Invalid program_ids value: {part!r} (use comma-separated integers).'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            if not id_list:
                return Response(
                    {'detail': 'program_ids must list at least one numeric programme id.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Preserve order, dedupe
            unique_ids = list(dict.fromkeys(id_list))
            all_programs = list(Program.objects.filter(id__in=unique_ids).order_by('code'))
            found_ids = {p.id for p in all_programs}
            missing = [i for i in unique_ids if i not in found_ids]
            if missing:
                return Response(
                    {'detail': f'Unknown program id(s): {missing}'},
                    status=status.HTTP_404_NOT_FOUND,
                )

        program_codes = [p.code for p in all_programs if p.code]
        program_ids = [p.id for p in all_programs]

        existing_batches = list(
            ProgramBatch.objects.filter(program_id__in=program_ids)
            .select_related("program")
            .order_by("program__code", "name")
        )
        batches_by_program: dict[int, list] = {}
        for batch in existing_batches:
            batches_by_program.setdefault(batch.program_id, []).append(batch)

        sample_rows = _build_batch_template_rows(all_programs, batches_by_program)

        cohort_count = len(existing_batches)
        program_count = len(all_programs)
        blank_count = program_count - len(batches_by_program)

        instructions = _batch_template_instructions(
            program_count=program_count,
            cohort_count=cohort_count,
            blank_count=blank_count,
            campus_scope_label=campus_scope_label,
        )

        out_format = (request.GET.get("format") or "xlsx").strip().lower()
        if out_format == "csv":
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(_BATCH_TEMPLATE_HEADERS)
            for row in sample_rows:
                writer.writerow(row)
            response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
            suffix = campus_scope_label.replace(" ", "_") if campus_scope_label else "all_programmes"
            response["Content-Disposition"] = (
                f'attachment; filename="batch_upload_template_{suffix}.csv"'
            )
            return response

        wb = create_workbook(
            headers=_BATCH_TEMPLATE_HEADERS,
            rows=sample_rows,
            sheet_name="Batch Upload",
            header_bg="3E397B",
            dropdowns={
                2: program_codes,       # col 2 = program_code
                10: ["TRUE", "FALSE"],  # col 10 = is_active
            },
            instructions=instructions,
        )

        # Add a second sheet: full program reference list (code + name + short_form)
        ref_ws = wb.create_sheet(title="Programs Reference")
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="3E397B", end_color="3E397B", fill_type="solid")
        ref_ws.append(["program_code", "program_name", "short_form", "is_active"])
        for col_idx, width in enumerate([20, 50, 20, 12], start=1):
            cell = ref_ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            from openpyxl.utils import get_column_letter
            ref_ws.column_dimensions[get_column_letter(col_idx)].width = width

        for prog in all_programs:
            ref_ws.append([
                prog.code or "",
                prog.name or "",
                prog.short_form or "",
                "Yes" if prog.is_active else "No",
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="batch_upload_template.xlsx"'
        return response


class BatchBulkUploadView(_BatchUnavailableMixin, APIView):
    """POST /api/program/batches/bulk_upload — create ProgramBatch rows from file."""
    permission_classes = [ProgramSchedulingAPIPermission]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        try:
            import pandas as pd
        except ImportError:
            return Response(
                {"detail": "pandas is not installed on the server."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        try:
            if filename.endswith(".csv"):
                df = pd.read_csv(file_obj, dtype=str)
            elif filename.endswith((".xlsx", ".xls")):
                df = pd.read_excel(file_obj, dtype=str, skiprows=1)  # skip instructions row
            else:
                return Response(
                    {"detail": "Unsupported file type. Upload .csv or .xlsx/.xls."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as exc:
            return Response(
                {"detail": f"Could not read file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Normalise column names
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
        for alias, canonical in (
            ("offer_start", "offer_start_date"),
            ("offer_end", "offer_end_date"),
            ("admission_offer_start", "offer_start_date"),
            ("admission_offer_end", "offer_end_date"),
            ("admission_offer_start_date", "offer_start_date"),
            ("admission_offer_end_date", "offer_end_date"),
        ):
            if alias in df.columns and canonical not in df.columns:
                df = df.rename(columns={alias: canonical})

        is_excel_upload = filename.endswith((".xlsx", ".xls"))
        data_row_offset = 3 if is_excel_upload else 2  # excel: instructions + header

        required_cols = {"program_code", "batch_name", "start_date"}
        missing = required_cols - set(df.columns)
        if missing:
            return Response(
                {"detail": f"Missing required columns: {', '.join(sorted(missing))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Drop completely empty rows
        df = df.dropna(how="all")

        # Program lookup caches
        program_map: dict = {
            p.code.strip(): p
            for p in Program.objects.filter(is_active=True)
            if p.code
        }
        short_form_map: dict = {
            p.short_form.strip().upper(): p
            for p in Program.objects.filter(is_active=True)
            if p.short_form
        }

        raw_update = request.data.get("update_existing")
        if raw_update is None:
            raw_update = request.POST.get("update_existing", "false")
        update_existing = str(raw_update).lower() in ("true", "1", "yes")

        errors: list = []
        created_count = 0
        updated_count = 0
        total_semesters_created = 0
        total = len(df)

        def _parse_date(raw):
            """
            Normalise a cell value to a Python date.

            Accepts:
              - datetime.date / datetime.datetime  (already parsed by pandas/openpyxl)
              - pandas Timestamp
              - "YYYY-MM-DD"
              - "YYYY-MM-DD HH:MM:SS" (Excel datetime exported as string)
              - Any ISO-8601-ish string that pandas can parse

            Returns (date, None) on success or (None, error_string) on failure.
            """
            import datetime as _dt
            import pandas as _pd

            if raw is None or (isinstance(raw, float) and _pd.isna(raw)):
                return None, "empty"

            # Already a date-like object
            if isinstance(raw, _dt.datetime):
                return raw.date(), None
            if isinstance(raw, _dt.date):
                return raw, None

            # pandas Timestamp
            try:
                if isinstance(raw, _pd.Timestamp):
                    return raw.date(), None
            except Exception:
                pass

            # String normalisation
            s = str(raw).strip()
            if not s:
                return None, "empty"

            # Try common explicit formats first (fast path)
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
                        "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).date(), None
                except ValueError:
                    pass

            # Fallback: let pandas infer the format
            try:
                return _pd.to_datetime(s, dayfirst=False).date(), None
            except Exception:
                pass

            return None, s  # return the raw string so the caller can use it in the error message

        for idx, row in df.iterrows():
            row_num = int(idx) + data_row_offset

            def cell(col: str) -> str:
                val = row.get(col, "")
                try:
                    if pd.isna(val):
                        return ""
                except TypeError:
                    pass
                return str(val).strip()

            def raw_val(col):
                """Return the raw (un-stringified) cell value for date parsing."""
                val = row.get(col, "")
                try:
                    if pd.isna(val):
                        return None
                except TypeError:
                    pass
                return val

            batch_id_cell = cell("batch_id")
            prog_code   = cell("program_code")
            batch_name  = cell("batch_name")
            acad_year   = cell("academic_year")
            is_active_s = cell("is_active").upper()

            if not prog_code:
                errors.append(f"Row {row_num}: program_code is required.")
                continue
            if not batch_name and not batch_id_cell:
                errors.append(
                    f"Row {row_num}: batch_name or batch_id is required."
                )
                continue

            start_raw = raw_val("start_date")
            if start_raw is None or str(start_raw).strip() == "":
                errors.append(f"Row {row_num}: start_date is required.")
                continue

            program = program_map.get(prog_code) or short_form_map.get(prog_code.upper())
            if not program:
                errors.append(
                    f"Row {row_num}: Program code '{prog_code}' not found. "
                    f"Check it matches an existing program code exactly."
                )
                continue

            start_date, start_err = _parse_date(start_raw)
            if start_date is None:
                errors.append(
                    f"Row {row_num}: start_date '{start_err}' is not a recognised date. "
                    f"Expected YYYY-MM-DD."
                )
                continue

            end_date = None
            end_raw = raw_val("end_date")
            if end_raw is not None and str(end_raw).strip() != "":
                end_date, end_err = _parse_date(end_raw)
                if end_date is None:
                    errors.append(
                        f"Row {row_num}: end_date '{end_err}' is not a recognised date. "
                        f"Expected YYYY-MM-DD."
                    )
                    continue
                if end_date <= start_date:
                    errors.append(f"Row {row_num}: end_date must be after start_date.")
                    continue

            offer_start_date = None
            offer_end_date = None
            os_raw = raw_val("offer_start_date")
            oe_raw = raw_val("offer_end_date")
            has_os = os_raw is not None and str(os_raw).strip() != ""
            has_oe = oe_raw is not None and str(oe_raw).strip() != ""
            if has_os ^ has_oe:
                errors.append(
                    f"Row {row_num}: offer_start_date and offer_end_date must both be set or both left blank."
                )
                continue
            if has_os:
                offer_start_date, ose = _parse_date(os_raw)
                if offer_start_date is None:
                    errors.append(
                        f"Row {row_num}: offer_start_date '{ose}' is not a recognised date. "
                        f"Expected YYYY-MM-DD."
                    )
                    continue
                offer_end_date, oee = _parse_date(oe_raw)
                if offer_end_date is None:
                    errors.append(
                        f"Row {row_num}: offer_end_date '{oee}' is not a recognised date. "
                        f"Expected YYYY-MM-DD."
                    )
                    continue
                if offer_end_date < offer_start_date:
                    errors.append(
                        f"Row {row_num}: offer_end_date must be on or after offer_start_date."
                    )
                    continue

            is_active = is_active_s not in ("FALSE", "0", "NO", "F")

            existing, match_err = _resolve_existing_program_batch(
                program, batch_name, batch_id_cell,
            )

            try:
                offer_start_date, offer_end_date = _finalize_batch_offer_dates(
                    start_date=start_date,
                    end_date=end_date,
                    offer_start_date=offer_start_date,
                    offer_end_date=offer_end_date,
                    has_explicit_offer=has_os and has_oe,
                    update_existing=update_existing,
                    existing=existing,
                )
            except ValueError as exc:
                errors.append(f"Row {row_num}: {exc}")
                continue

            if update_existing:
                if match_err == "invalid_batch_id":
                    errors.append(f"Row {row_num}: batch_id '{batch_id_cell}' is not a valid number.")
                    continue
                if match_err == "batch_id_not_found":
                    errors.append(
                        f"Row {row_num}: batch_id '{batch_id_cell}' not found for program "
                        f"'{program.code}'."
                    )
                    continue
                if not existing:
                    available = list(
                        ProgramBatch.objects.filter(program=program)
                        .order_by("name")
                        .values_list("name", flat=True)[:10]
                    )
                    hint = ""
                    if available:
                        hint = f" Existing batch_name values: {', '.join(repr(n) for n in available)}."
                    errors.append(
                        f"Row {row_num}: No cohort matched for program '{program.code}' "
                        f"(batch_name={batch_name!r}, batch_id={batch_id_cell or '—'}). "
                        f"Download Batch template again and edit dates only — do not rename the cohort."
                        f"{hint}"
                    )
                    continue
                # ── Update mode: patch the existing batch (never create) ──
                try:
                    with transaction.atomic():
                        existing.start_date = start_date
                        existing.end_date = end_date
                        existing.academic_year = acad_year
                        existing.is_active = is_active
                        existing.offer_start_date = offer_start_date
                        existing.offer_end_date = offer_end_date
                        existing.save(update_fields=[
                            "start_date", "end_date", "academic_year", "is_active",
                            "offer_start_date", "offer_end_date",
                        ])
                    updated_count += 1
                except Exception as exc:
                    errors.append(f"Row {row_num}: Could not update batch '{batch_name}' — {exc}")
                continue  # no semester re-creation on update

            if update_existing:
                errors.append(
                    f"Row {row_num}: Update mode — row was not applied (internal). "
                    f"Contact support if this persists."
                )
                continue

            if existing:
                errors.append(
                    f"Row {row_num}: Batch '{existing.name}' already exists for "
                    f"program '{program.code}' (id={existing.id}). Skipped. "
                    f"Use 'Update batches' to change dates, not Upload batches."
                )
                continue

            # ── Create mode: new batch (own committed transaction) ──
            try:
                with transaction.atomic():
                    batch = ProgramBatch.objects.create(
                        program=program,
                        name=batch_name,
                        academic_year=acad_year,
                        start_date=start_date,
                        end_date=end_date,
                        offer_start_date=offer_start_date,
                        offer_end_date=offer_end_date,
                        is_active=is_active,
                    )
                created_count += 1
            except Exception as exc:
                errors.append(f"Row {row_num}: Could not create batch — {exc}")
                continue

            # ── Step 2: auto-create semesters (separate transaction so a failure
            #            here does NOT roll back the batch that was just saved) ──
            try:
                with transaction.atomic():
                    sems = _auto_create_semesters(batch, program)
                    total_semesters_created += sems
            except Exception as exc:
                errors.append(
                    f"Row {row_num}: Batch '{batch_name}' created but semesters could "
                    f"not be auto-generated — {exc}"
                )

        processed = created_count + updated_count
        return Response(
            {
                "total":              total,
                "created":            created_count,
                "updated":            updated_count,
                "failed":             total - processed,
                "semesters_created":  total_semesters_created,
                "errors":             errors,
            },
            status=status.HTTP_200_OK,
        )


class AutoCreateSemestersView(_BatchUnavailableMixin, APIView):
    """
    POST /api/program/batches/auto_create_semesters

    Finds every ProgramBatch that currently has zero Semester rows and
    auto-generates the full minimum-duration term structure for it using
    the program's calendar_type and min_years.

    Optional body param:
      program_id (int) — limit to batches of one specific program.

    Returns:
      {
        "batches_processed": N,
        "semesters_created": M,
        "skipped_already_have_semesters": K,
        "errors": [...]
      }
    """
    permission_classes = [ProgramSchedulingAPIPermission]

    def post(self, request):
        program_id = request.data.get("program_id")

        qs = ProgramBatch.objects.select_related("program").prefetch_related("semesters")
        if program_id:
            qs = qs.filter(program_id=program_id)

        batches_processed = 0
        semesters_created = 0
        skipped = 0
        errors = []

        for batch in qs:
            if batch.semesters.exists():
                skipped += 1
                continue
            try:
                with transaction.atomic():
                    n = _auto_create_semesters(batch, batch.program)
                    semesters_created += n
                    batches_processed += 1
            except Exception as exc:
                errors.append(
                    f"Batch '{batch.name}' ({batch.program.code}): {exc}"
                )

        return Response(
            {
                "batches_processed":              batches_processed,
                "semesters_created":              semesters_created,
                "skipped_already_have_semesters": skipped,
                "errors":                         errors,
            },
            status=status.HTTP_200_OK,
        )
