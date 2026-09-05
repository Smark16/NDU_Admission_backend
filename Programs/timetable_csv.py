"""
TimetableSession CSV template, parse, validate, and apply.

CSV columns (header required):
  course_code*   — one code or comma-separated (BIT1107,BCS1107)
  programmes     — informational (full programme names); ignored on import
  day†           — Mon…Sun or 1…7
  start_time* / end_time*  — HH:MM
  start_date / end_date†   — recurring range (defaults to semester dates)
  session_date   — one-off (when set, day/range derived from it)
  venue_code / room_label  — registered venue preferred for on-campus
  session_type   — lecture | tutorial | practical
  delivery_mode  — on_campus | online | hybrid
  teaching_section — stream code e.g. I / II
  shared_key     — same key → link units into SharedTeachingOffering
  lecturer_email — comma-separated staff emails
  is_published   — default false
  notes
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q

from Programs.models import CourseUnit, Semester, TeachingSection, TimetableSession, Venue
from Programs.shared_teaching import (
    course_code_number,
    create_shared_offering_from_course_units,
    normalize_course_code,
    study_mode_for_course_unit,
)
from Programs.teaching_sections import section_covers_batch
from Programs.timetable_utils import parse_delivery_mode, validate_session_scheduling

IMPORT_TAG = "[TT-CSV]"

TEMPLATE_COLUMNS = [
    "course_code",
    "programmes",
    "day",
    "start_time",
    "end_time",
    "start_date",
    "end_date",
    "session_date",
    "venue_code",
    "room_label",
    "session_type",
    "delivery_mode",
    "teaching_section",
    "shared_key",
    "lecturer_email",
    "is_published",
    "notes",
]


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%H:%M")
        return value.date().isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    # Excel sometimes gives floats for whole numbers
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    # Excel time as fraction of day is rare if openpyxl returns datetime; leave as-is
    return text


def spreadsheet_bytes_to_csv_text(data: bytes, filename: str = "") -> str:
    """
    Accept UTF-8 CSV or Excel (.xlsx / .xls) and return CSV text for apply_import.

    Excel: first worksheet, first row = headers (same columns as the worksheet).
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel support requires openpyxl on the server.") from exc
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_out: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows_out.append([_cell_str(c) for c in row])
        finally:
            wb.close()
        if not rows_out:
            raise ValueError("Excel sheet is empty.")
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in rows_out:
            writer.writerow(row)
        return buf.getvalue()

    # CSV / plain text
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("Could not decode file — use UTF-8 CSV or Excel (.xlsx).") from exc


def worksheet_csv_to_xlsx_bytes(
    csv_text: str,
    *,
    venue_codes: list[str] | None = None,
) -> bytes:
    """
    Convert worksheet/template CSV text into a downloadable .xlsx workbook.

    Adds Excel data-validation dropdowns (CSV cannot have dropdowns):
      day, session_type, delivery_mode, is_published, and venue_code when codes are provided.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise ValueError("Excel export requires openpyxl on the server.") from exc

    reader = csv.reader(io.StringIO(csv_text))
    rows = list(reader)
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    header_fill = PatternFill("solid", fgColor="1B3A5C")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="FFF3CD")

    headers: list[str] = []
    max_row = 1
    for r_idx, row in enumerate(rows, start=1):
        max_row = r_idx
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                headers.append(_norm_col(val))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True)
            elif (val or "").startswith("#"):
                cell.fill = section_fill
                cell.font = Font(italic=True, color="664D03")

    # A=course_code B=programmes C=day …
    widths = {
        "A": 28,
        "B": 48,
        "C": 8,
        "D": 10,
        "E": 10,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 16,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 16,
        "O": 28,
        "P": 12,
        "Q": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    # Dropdown source lists on a hidden sheet (Excel only — not CSV)
    lists = wb.create_sheet("_Lists")
    day_vals = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    session_vals = ["lecture", "tutorial", "practical"]
    delivery_vals = ["on_campus", "online", "hybrid"]
    published_vals = ["false", "true"]
    venues = [
        (c or "").strip()
        for c in (venue_codes or [])
        if (c or "").strip()
    ]
    if not venues:
        try:
            venues = list(
                Venue.objects.filter(is_active=True)
                .exclude(code="")
                .order_by("code")
                .values_list("code", flat=True)[:400]
            )
        except Exception:
            venues = []

    list_cols = {
        "day": day_vals,
        "session_type": session_vals,
        "delivery_mode": delivery_vals,
        "is_published": published_vals,
        "venue_code": venues,
    }
    for col_i, (name, values) in enumerate(list_cols.items(), start=1):
        lists.cell(row=1, column=col_i, value=name)
        for r_i, val in enumerate(values, start=2):
            lists.cell(row=r_i, column=col_i, value=val)
    lists.sheet_state = "hidden"

    def _col_index(field: str) -> int | None:
        try:
            return headers.index(field) + 1  # 1-based
        except ValueError:
            return None

    def _add_list_dv(field: str, list_col: int, n_values: int) -> None:
        if n_values < 1:
            return
        col = _col_index(field)
        if not col:
            return
        end_row = max(max_row + 200, 500)
        letter = get_column_letter(list_col)
        formula = f"'_Lists'!${letter}$2:${letter}${n_values + 1}"
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False = show the arrow in Excel
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Pick a value from the {field} dropdown.",
        )
        dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{end_row}")
        ws.add_data_validation(dv)

    _add_list_dv("day", 1, len(day_vals))
    _add_list_dv("session_type", 2, len(session_vals))
    _add_list_dv("delivery_mode", 3, len(delivery_vals))
    _add_list_dv("is_published", 4, len(published_vals))
    _add_list_dv("venue_code", 5, len(venues))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

DAY_ALIASES = {
    "1": 1,
    "mon": 1,
    "monday": 1,
    "2": 2,
    "tue": 2,
    "tues": 2,
    "tuesday": 2,
    "3": 3,
    "wed": 3,
    "wednesday": 3,
    "4": 4,
    "thu": 4,
    "thur": 4,
    "thurs": 4,
    "thursday": 4,
    "5": 5,
    "fri": 5,
    "friday": 5,
    "6": 6,
    "sat": 6,
    "saturday": 6,
    "7": 7,
    "sun": 7,
    "sunday": 7,
}

SESSION_TYPES = {c[0] for c in TimetableSession.SESSION_TYPE_CHOICES}


@dataclass
class CsvRowError:
    row: int
    course_code: str
    reason: str

    def as_dict(self) -> dict:
        return {
            "row": self.row,
            "course_code": self.course_code,
            "reason": self.reason,
        }


@dataclass
class PreparedSession:
    row_num: int
    course_unit: CourseUnit
    teaching_section: TeachingSection | None
    day_of_week: int
    session_date: date | None
    start_date: date | None
    end_date: date | None
    start_time: time
    end_time: time
    venue: Venue | None
    room_label: str
    session_type: str
    delivery_mode: str
    is_published: bool
    notes: str
    shared_key: str
    lecturer_ids: list[int] = field(default_factory=list)


@dataclass
class ImportResult:
    created: int = 0
    shared_offerings: int = 0
    dry_run: bool = False
    strict: bool = True
    errors: list[CsvRowError] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    session_ids: list[int] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "created": self.created,
            "shared_offerings": self.shared_offerings,
            "dry_run": self.dry_run,
            "strict": self.strict,
            "errors": [e.as_dict() for e in self.errors],
            "warnings": self.warnings,
            "session_ids": self.session_ids,
        }


def _norm_col(name: str) -> str:
    return (name or "").strip().lower().replace(" ", "_")


def _parse_bool(value: Any, default: bool = False) -> bool | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ("1", "true", "yes", "y", "t"):
        return True
    if text in ("0", "false", "no", "n", "f"):
        return False
    return None


def _parse_date(value: str, label: str) -> date:
    text = (value or "").strip()
    if not text:
        raise ValueError(f"{label} is required.")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{label} must be YYYY-MM-DD (got {text!r}).")


def _parse_time(value: str, label: str) -> time:
    text = (value or "").strip()
    if not text:
        raise ValueError(f"{label} is required.")
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"{label} must be HH:MM (got {text!r}).")


def _parse_day(value: str) -> int:
    key = (value or "").strip().lower()
    if not key:
        raise ValueError("day is required for a recurring class.")
    day = DAY_ALIASES.get(key)
    if day is None:
        raise ValueError(f"day must be Mon–Sun or 1–7 (got {value!r}).")
    return day


def _split_codes(raw: str) -> list[str]:
    parts = re.split(r"[,/;|]+", raw or "")
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        code = p.strip()
        if not code:
            continue
        key = normalize_course_code(code)
        if key in seen:
            continue
        seen.add(key)
        out.append(code)
    return out


def _split_emails(raw: str) -> list[str]:
    parts = re.split(r"[,;]+", raw or "")
    return [p.strip().lower() for p in parts if p.strip()]


def csv_template_text(*, semester: Semester | None = None) -> str:
    """Empty template with header + one commented example row as a second header comment."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_COLUMNS)
    start = ""
    end = ""
    if semester:
        if semester.start_date:
            start = semester.start_date.isoformat()
        if semester.end_date:
            end = semester.end_date.isoformat()
    writer.writerow(
        [
            "BIT1107,BCS1107",
            "Bachelor of Information Technology (BIT1107) | Bachelor of Computer Science (BCS1107)",
            "Mon",
            "08:30",
            "11:30",
            start or "2026-08-17",
            end or "2026-12-18",
            "",
            "D1-1",
            "",
            "lecture",
            "on_campus",
            "",
            "prog-y1-fri",
            "lecturer@ndejjeuniversity.ac.ug",
            "false",
            "Example — replace codes/venues then delete this row",
        ]
    )
    return buf.getvalue()


def _semester_course_units(semester: Semester) -> list[CourseUnit]:
    return list(
        CourseUnit.objects.filter(semester_id=semester.pk, is_active=True)
        .select_related(
            "semester",
            "semester__program_batch",
            "semester__program_batch__program",
            "program_batch",
            "program_batch__program",
            "shared_teaching_offering",
            "catalog_unit",
        )
        .prefetch_related(
            "lecturers",
            "semester__program_batch__program__campuses",
            "program_batch__program__campuses",
        )
    )


def _peer_units_same_year(semester: Semester, code: str) -> list[CourseUnit]:
    """Find active units with the same normalized code in the same academic year."""
    batch = semester.program_batch
    ay = (getattr(batch, "academic_year", None) or "").strip()
    norm = normalize_course_code(code)
    if not norm:
        return []
    qs = (
        CourseUnit.objects.filter(is_active=True)
        .exclude(semester_id=semester.pk)
        .select_related(
            "semester",
            "semester__program_batch",
            "semester__program_batch__program",
            "program_batch",
            "program_batch__program",
            "shared_teaching_offering",
            "catalog_unit",
        )
        .prefetch_related(
            "lecturers",
            "semester__program_batch__program__campuses",
            "program_batch__program__campuses",
        )
    )
    if ay:
        qs = qs.filter(
            Q(program_batch__academic_year__iexact=ay)
            | Q(semester__program_batch__academic_year__iexact=ay)
        )
    matches = []
    for cu in qs.iterator(chunk_size=200):
        if normalize_course_code(cu.code) == norm:
            matches.append(cu)
            if len(matches) >= 12:
                break
    return matches


def _resolve_course_units(
    semester: Semester,
    codes: list[str],
    *,
    local_units: list[CourseUnit],
) -> tuple[list[CourseUnit], list[str]]:
    by_norm: dict[str, list[CourseUnit]] = {}
    for cu in local_units:
        key = normalize_course_code(cu.code)
        by_norm.setdefault(key, []).append(cu)

    found: list[CourseUnit] = []
    missing: list[str] = []
    seen_ids: set[int] = set()

    for code in codes:
        key = normalize_course_code(code)
        local = by_norm.get(key) or []
        if local:
            for cu in local:
                if cu.id not in seen_ids:
                    seen_ids.add(cu.id)
                    found.append(cu)
            continue
        peers = _peer_units_same_year(semester, code)
        if peers:
            for cu in peers:
                if cu.id not in seen_ids:
                    seen_ids.add(cu.id)
                    found.append(cu)
            continue
        missing.append(code)
    return found, missing


def _resolve_venue(
    semester: Semester,
    venue_code: str,
    room_label: str,
) -> tuple[Venue | None, str, str | None]:
    """Return (venue, room_label, error)."""
    code = (venue_code or "").strip()
    label = (room_label or "").strip()
    if not code:
        return None, label, None

    batch = semester.program_batch
    program = batch.program if batch else None
    campus_ids = list(program.campuses.values_list("id", flat=True)) if program else []

    qs = Venue.objects.filter(is_active=True).filter(
        Q(code__iexact=code) | Q(name__iexact=code)
    )
    if campus_ids:
        scoped = qs.filter(campus_id__in=campus_ids)
        if scoped.exists():
            qs = scoped
    venues = list(qs[:5])
    if not venues:
        if label:
            return None, label or code, None
        return None, code, None
    if len(venues) > 1:
        return None, label, (
            f'venue_code "{code}" matches multiple rooms — use a unique code or room_label.'
        )
    return venues[0], label, None


def _resolve_section(course_unit: CourseUnit, section_code: str) -> TeachingSection | None:
    code = (section_code or "").strip()
    if not code:
        return None
    batch_id = course_unit.program_batch_id
    if batch_id is None and course_unit.semester_id:
        sem = getattr(course_unit, "semester", None)
        batch_id = getattr(sem, "program_batch_id", None) if sem else None
    if not batch_id:
        raise ValueError("Course unit has no cohort for teaching sections.")
    section = (
        TeachingSection.objects.filter(is_active=True, code__iexact=code)
        .filter(Q(program_batch_id=batch_id) | Q(is_shared=True))
        .first()
    )
    if section is None or not section_covers_batch(section, batch_id):
        raise ValueError(f'Teaching section "{code}" not found on this cohort.')
    return section


def _resolve_lecturers(emails: list[str]) -> tuple[list[int], list[str]]:
    if not emails:
        return [], []
    User = get_user_model()
    found_ids: list[int] = []
    missing: list[str] = []
    for email in emails:
        user = (
            User.objects.filter(email__iexact=email, is_active=True)
            .filter(Q(is_staff=True) | Q(is_lecturer=True))
            .first()
        )
        if user is None:
            missing.append(email)
        else:
            found_ids.append(user.id)
    return found_ids, missing


def _tag_notes(notes: str) -> str:
    text = (notes or "").strip()
    if text.startswith(IMPORT_TAG):
        return text[:255]
    if not text:
        return IMPORT_TAG
    return f"{IMPORT_TAG} {text}"[:255]


def prepare_import(
    semester: Semester,
    text: str,
    *,
    strict: bool = True,
) -> tuple[list[PreparedSession], list[CsvRowError], list[str]]:
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is empty or has no header row.")

    col_aliases = {
        "course_code": "course_code",
        "code": "course_code",
        "course": "course_code",
        "day": "day",
        "day_of_week": "day",
        "weekday": "day",
        "start_time": "start_time",
        "end_time": "end_time",
        "start_date": "start_date",
        "end_date": "end_date",
        "session_date": "session_date",
        "date": "session_date",
        "venue_code": "venue_code",
        "venue": "venue_code",
        "room_code": "venue_code",
        "room_label": "room_label",
        "room": "room_label",
        "session_type": "session_type",
        "type": "session_type",
        "delivery_mode": "delivery_mode",
        "mode": "delivery_mode",
        "teaching_section": "teaching_section",
        "section": "teaching_section",
        "stream": "teaching_section",
        "shared_key": "shared_key",
        "share_key": "shared_key",
        "lecturer_email": "lecturer_email",
        "lecturer_emails": "lecturer_email",
        "is_published": "is_published",
        "published": "is_published",
        "notes": "notes",
        "note": "notes",
    }
    resolved: dict[str, str] = {}
    for raw in reader.fieldnames:
        key = col_aliases.get(_norm_col(raw))
        if key and key not in resolved:
            resolved[key] = raw

    if "course_code" not in resolved:
        raise ValueError('CSV must include a "course_code" column.')
    if "start_time" not in resolved or "end_time" not in resolved:
        raise ValueError('CSV must include "start_time" and "end_time" columns.')

    local_units = _semester_course_units(semester)
    prepared: list[PreparedSession] = []
    errors: list[CsvRowError] = []
    warnings: list[str] = []

    default_start = semester.start_date
    default_end = semester.end_date

    for row_num, raw in enumerate(reader, start=2):
        def cell(key: str) -> str:
            col = resolved.get(key)
            if not col:
                return ""
            return (raw.get(col) or "").strip()

        course_raw = cell("course_code")
        if not course_raw or course_raw.startswith("#"):
            continue

        codes = _split_codes(course_raw)
        if not codes:
            errors.append(CsvRowError(row_num, course_raw, "course_code is required."))
            continue

        units, missing = _resolve_course_units(semester, codes, local_units=local_units)
        if missing:
            errors.append(
                CsvRowError(
                    row_num,
                    course_raw,
                    f"No active course unit for code(s): {', '.join(missing)}. "
                    "Create offerings on the batch (or matching academic year) first.",
                )
            )
            if strict or not units:
                continue

        try:
            start_time = _parse_time(cell("start_time"), "start_time")
            end_time = _parse_time(cell("end_time"), "end_time")
        except ValueError as exc:
            errors.append(CsvRowError(row_num, course_raw, str(exc)))
            continue
        if end_time <= start_time:
            errors.append(CsvRowError(row_num, course_raw, "end_time must be after start_time."))
            continue

        session_date = None
        start_date = None
        end_date = None
        day = 0
        try:
            if cell("session_date"):
                session_date = _parse_date(cell("session_date"), "session_date")
                day = session_date.weekday() + 1
                start_date = session_date
                end_date = session_date
            else:
                day = _parse_day(cell("day")) if cell("day") else 0
                if not day:
                    raise ValueError("day is required when session_date is blank.")
                start_date = (
                    _parse_date(cell("start_date"), "start_date")
                    if cell("start_date")
                    else default_start
                )
                end_date = (
                    _parse_date(cell("end_date"), "end_date")
                    if cell("end_date")
                    else default_end
                )
                if not start_date or not end_date:
                    raise ValueError(
                        "start_date and end_date are required (or set them on the semester)."
                    )
                if end_date < start_date:
                    raise ValueError("end_date must be on or after start_date.")
        except ValueError as exc:
            errors.append(CsvRowError(row_num, course_raw, str(exc)))
            continue

        session_type = (cell("session_type") or "lecture").strip().lower()
        if session_type not in SESSION_TYPES:
            errors.append(
                CsvRowError(
                    row_num,
                    course_raw,
                    f"session_type must be one of: {', '.join(sorted(SESSION_TYPES))}.",
                )
            )
            continue

        try:
            delivery_mode = parse_delivery_mode(cell("delivery_mode") or "on_campus")
        except ValueError as exc:
            errors.append(CsvRowError(row_num, course_raw, str(exc)))
            continue

        published = _parse_bool(cell("is_published"), default=False)
        if published is None:
            errors.append(CsvRowError(row_num, course_raw, "is_published must be yes/no or 1/0."))
            continue

        venue, room_label, venue_err = _resolve_venue(
            semester, cell("venue_code"), cell("room_label")
        )
        if venue_err:
            errors.append(CsvRowError(row_num, course_raw, venue_err))
            continue
        if delivery_mode != "online" and not venue and not room_label:
            warnings.append(
                f"Row {row_num}: no venue_code/room_label — slot saved as TBA "
                "(publish will require a registered room for on-campus/hybrid)."
            )

        lecturer_ids, missing_emails = _resolve_lecturers(_split_emails(cell("lecturer_email")))
        if missing_emails:
            msg = f"Unknown lecturer email(s): {', '.join(missing_emails)}."
            if strict:
                errors.append(CsvRowError(row_num, course_raw, msg))
                continue
            warnings.append(f"Row {row_num}: {msg} Continuing without them.")

        section_code = cell("teaching_section")
        shared_key = (cell("shared_key") or "").strip()
        notes = _tag_notes(cell("notes"))

        for cu in units:
            try:
                section = _resolve_section(cu, section_code) if section_code else None
            except ValueError as exc:
                errors.append(CsvRowError(row_num, cu.code or course_raw, str(exc)))
                if strict:
                    break
                continue

            prepared.append(
                PreparedSession(
                    row_num=row_num,
                    course_unit=cu,
                    teaching_section=section,
                    day_of_week=day,
                    session_date=session_date,
                    start_date=start_date,
                    end_date=end_date,
                    start_time=start_time,
                    end_time=end_time,
                    venue=venue,
                    room_label=room_label,
                    session_type=session_type,
                    delivery_mode=delivery_mode,
                    is_published=bool(published),
                    notes=notes,
                    shared_key=shared_key,
                    lecturer_ids=list(lecturer_ids),
                )
            )

    return prepared, errors, warnings


def apply_import(
    semester: Semester,
    text: str,
    *,
    strict: bool = True,
    dry_run: bool = False,
) -> ImportResult:
    result = ImportResult(dry_run=dry_run, strict=strict)
    try:
        prepared, errors, warnings = prepare_import(semester, text, strict=strict)
    except ValueError as exc:
        result.errors.append(CsvRowError(0, "", str(exc)))
        return result

    result.errors.extend(errors)
    result.warnings.extend(warnings)

    if strict and result.errors:
        return result
    if not prepared:
        if not result.errors:
            result.errors.append(CsvRowError(0, "", "No valid data rows to import."))
        return result

    # Pre-validate scheduling on unsaved instances
    valid_prepared: list[PreparedSession] = []
    for item in prepared:
        session = TimetableSession(
            course_unit=item.course_unit,
            teaching_section=item.teaching_section,
            day_of_week=item.day_of_week,
            session_date=item.session_date,
            start_date=item.start_date,
            end_date=item.end_date,
            start_time=item.start_time,
            end_time=item.end_time,
            venue=item.venue,
            room_label=item.room_label,
            session_type=item.session_type,
            delivery_mode=item.delivery_mode,
            notes=item.notes,
            is_published=item.is_published,
        )
        validation = validate_session_scheduling(session, require_venue=False)
        if not validation.ok:
            result.errors.append(
                CsvRowError(
                    item.row_num,
                    item.course_unit.code or "",
                    "; ".join(validation.errors),
                )
            )
            continue
        result.warnings.extend(
            f"Row {item.row_num} ({item.course_unit.code}): {w}" for w in validation.warnings
        )
        try:
            session.full_clean()
        except Exception as exc:  # noqa: BLE001
            result.errors.append(
                CsvRowError(item.row_num, item.course_unit.code or "", str(exc))
            )
            continue
        valid_prepared.append(item)

    if strict and result.errors:
        # Drop creations when any hard error (including validation) in strict mode
        return result

    if dry_run:
        result.created = len(valid_prepared)
        # Estimate shared offerings
        keys = {p.shared_key for p in valid_prepared if p.shared_key}
        for key in keys:
            unit_ids = {p.course_unit.id for p in valid_prepared if p.shared_key == key}
            if len(unit_ids) >= 2:
                result.shared_offerings += 1
        return result

    created_sessions: list[TimetableSession] = []
    with transaction.atomic():
        for item in valid_prepared:
            session = TimetableSession(
                course_unit=item.course_unit,
                teaching_section=item.teaching_section,
                day_of_week=item.day_of_week,
                session_date=item.session_date,
                start_date=item.start_date,
                end_date=item.end_date,
                start_time=item.start_time,
                end_time=item.end_time,
                venue=item.venue,
                room_label=item.room_label,
                session_type=item.session_type,
                delivery_mode=item.delivery_mode,
                notes=item.notes,
                is_published=item.is_published,
            )
            session.save()
            created_sessions.append(session)
            if item.lecturer_ids:
                try:
                    from Programs.section_lecturers import assign_lecturers_to_section

                    assign_lecturers_to_section(
                        item.course_unit,
                        item.lecturer_ids,
                        teaching_section=item.teaching_section,
                        mode="add",
                    )
                except Exception as exc:  # noqa: BLE001
                    result.warnings.append(
                        f"Row {item.row_num}: session saved but lecturers not updated: {exc}"
                    )

        # Group shared_key → SharedTeachingOffering
        by_key: dict[str, list[PreparedSession]] = {}
        for item in valid_prepared:
            if item.shared_key:
                by_key.setdefault(item.shared_key, []).append(item)

        batch = semester.program_batch
        ay = (getattr(batch, "academic_year", None) or "").strip() if batch else ""

        for key, items in by_key.items():
            unit_ids = list({p.course_unit.id for p in items})
            if len(unit_ids) < 2:
                result.warnings.append(
                    f'shared_key "{key}" has fewer than 2 course units — skipped STO link.'
                )
                continue
            # Skip if already all on same STO
            existing = {
                CourseUnit.objects.filter(pk=uid)
                .values_list("shared_teaching_offering_id", flat=True)
                .first()
                for uid in unit_ids
            }
            if len(existing) == 1 and None not in existing:
                continue
            try:
                create_shared_offering_from_course_units(
                    course_unit_ids=unit_ids,
                    academic_year_label=ay,
                    notes=f"{IMPORT_TAG} shared_key={key}",
                    parent_course_unit_id=items[0].course_unit.id,
                )
                result.shared_offerings += 1
            except ValueError as exc:
                result.warnings.append(f'shared_key "{key}": {exc}')

    result.created = len(created_sessions)
    result.session_ids = [s.id for s in created_sessions]
    return result


# ── Worksheet export (DB → pre-filled CSV for humans to schedule) ─────────────


def _cu_batch(cu: CourseUnit):
    return cu.program_batch or getattr(cu.semester, "program_batch", None)


def _cu_program(cu: CourseUnit):
    batch = _cu_batch(cu)
    return batch.program if batch else None


def _prog_name(cu: CourseUnit) -> str:
    prog = _cu_program(cu)
    if not prog:
        return "?"
    name = (prog.name or "").strip()
    if name:
        return name
    return (prog.short_form or prog.code or str(prog.id)).strip() or "?"


def _programmes_cell(units: list[CourseUnit]) -> str:
    """Full programme names for the worksheet (ignored on import)."""
    parts: list[str] = []
    seen: set[str] = set()
    for cu in units:
        code = (cu.code or "").strip()
        name = _prog_name(cu)
        label = f"{name} ({code})" if code else name
        if label in seen:
            continue
        seen.add(label)
        parts.append(label)
    return " | ".join(parts)


def _comment_row(text: str) -> list[str]:
    return [text] + [""] * (len(TEMPLATE_COLUMNS) - 1)


def _cu_lecturer_emails(cu: CourseUnit) -> str:
    emails = []
    for lec in cu.lecturers.all():
        email = (getattr(lec, "email", None) or "").strip()
        if email:
            emails.append(email)
    return ",".join(emails)


def _section_codes_for_unit(cu: CourseUnit) -> list[str]:
    """Return teaching-section codes when the cohort has parallel streams; else ['']."""
    from Programs.teaching_sections import list_sections_for_batch

    batch = _cu_batch(cu)
    if not batch:
        return [""]
    sections = list_sections_for_batch(batch.pk)
    codes = [
        (s.get("code") or "").strip()
        for s in sections
        if (s.get("code") or "").strip() and not s.get("is_default")
    ]
    # Only emit per-stream rows when there are real parallel streams (2+).
    if len(codes) >= 2:
        return codes
    return [""]


def _blank_schedule_row(
    *,
    course_codes: str,
    start_date: str,
    end_date: str,
    programmes: str = "",
    teaching_section: str = "",
    shared_key: str = "",
    lecturer_email: str = "",
    notes: str = "",
    session_type: str = "lecture",
) -> list[str]:
    return [
        course_codes,
        programmes,
        "",  # day — staff fills
        "",  # start_time
        "",  # end_time
        start_date,
        end_date,
        "",  # session_date
        "",  # venue_code
        "",  # room_label
        session_type,
        "on_campus",
        teaching_section,
        shared_key,
        lecturer_email,
        "false",
        notes[:240],
    ]


def _collect_worksheet_units(
    *,
    semester: Semester | None = None,
    faculty_id: int | None = None,
    academic_year: str = "",
    study_mode: str = "",
    campus_id: int | None = None,
    expand: str = "faculty",
) -> tuple[list[CourseUnit], dict]:
    """
    Collect active course units for the worksheet.

    expand:
      - none: only the given semester's units (semester required)
      - faculty: same faculty + academic year (+ study_mode if set)
      - ay: same academic year only
    """
    meta: dict[str, Any] = {
        "faculty_id": faculty_id,
        "academic_year": academic_year,
        "study_mode": study_mode,
        "campus_id": campus_id,
        "expand": expand,
        "semester_id": semester.pk if semester else None,
    }

    qs = (
        CourseUnit.objects.filter(is_active=True, semester__isnull=False)
        .select_related(
            "semester",
            "semester__program_batch",
            "semester__program_batch__program",
            "semester__program_batch__program__faculty",
            "program_batch",
            "program_batch__program",
            "program_batch__program__faculty",
            "shared_teaching_offering",
            "catalog_unit",
        )
        .prefetch_related(
            "lecturers",
            "program_batch__program__campuses",
            "semester__program_batch__program__campuses",
        )
    )

    if semester is not None and (expand or "faculty") == "none":
        units = list(qs.filter(semester_id=semester.pk))
        batch = semester.program_batch
        meta["academic_year"] = (getattr(batch, "academic_year", None) or academic_year or "").strip()
        if batch and batch.program_id:
            meta["faculty_id"] = batch.program.faculty_id
            meta["study_mode"] = study_mode or study_mode_for_course_unit(
                CourseUnit(program_batch=batch, semester=semester)
            )
        return units, meta

    # Derive defaults from semester when expanding
    if semester is not None:
        batch = semester.program_batch
        if batch:
            if not academic_year:
                academic_year = (batch.academic_year or "").strip()
            if not faculty_id and batch.program_id:
                faculty_id = batch.program.faculty_id
            if not study_mode:
                # Infer from this batch's programme naming
                probe = CourseUnit(program_batch=batch, semester=semester)
                study_mode = study_mode_for_course_unit(probe)
        meta["academic_year"] = academic_year
        meta["faculty_id"] = faculty_id
        meta["study_mode"] = study_mode

    if academic_year:
        qs = qs.filter(
            Q(program_batch__academic_year__iexact=academic_year)
            | Q(semester__program_batch__academic_year__iexact=academic_year)
        )
    if faculty_id and expand != "ay":
        qs = qs.filter(
            Q(program_batch__program__faculty_id=faculty_id)
            | Q(semester__program_batch__program__faculty_id=faculty_id)
        )
    if campus_id:
        qs = qs.filter(
            Q(program_batch__program__campuses__id=campus_id)
            | Q(semester__program_batch__program__campuses__id=campus_id)
        ).distinct()

    units = list(qs.order_by("code", "id"))
    if study_mode and study_mode.lower() not in ("", "all", "any", "other"):
        preferred = study_mode.strip()
        units = [
            u
            for u in units
            if study_mode_for_course_unit(u) == preferred
            or study_mode_for_course_unit(u) == "Other"
        ]
    return units, meta


def build_timetable_worksheet(
    *,
    semester: Semester | None = None,
    faculty_id: int | None = None,
    academic_year: str = "",
    study_mode: str = "",
    campus_id: int | None = None,
    expand: str = "faculty",
) -> str:
    """
    Pre-filled CSV worksheet from the database.

    Sections (comment rows start with # and are skipped on import):
      A — Already Shared Teaching (STO)
      B — Cross-cutting candidates (catalog flag / shared paper number)
      C — Programme-only units (and engineering stream rows when sections exist)

    Day / time / venue left blank for staff to schedule.
    """
    units, meta = _collect_worksheet_units(
        semester=semester,
        faculty_id=faculty_id,
        academic_year=academic_year,
        study_mode=study_mode,
        campus_id=campus_id,
        expand=expand,
    )

    # Date defaults: prefer the open semester, else first unit's semester
    start_date = ""
    end_date = ""
    if semester and semester.start_date:
        start_date = semester.start_date.isoformat()
    if semester and semester.end_date:
        end_date = semester.end_date.isoformat()
    if not start_date and units:
        s0 = units[0].semester
        if s0 and s0.start_date:
            start_date = s0.start_date.isoformat()
        if s0 and s0.end_date:
            end_date = s0.end_date.isoformat()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_COLUMNS)

    scope_bits = []
    if meta.get("academic_year"):
        scope_bits.append(f"AY {meta['academic_year']}")
    if meta.get("faculty_id"):
        scope_bits.append(f"faculty_id={meta['faculty_id']}")
    if meta.get("study_mode"):
        scope_bits.append(f"mode={meta['study_mode']}")
    if meta.get("campus_id"):
        scope_bits.append(f"campus_id={meta['campus_id']}")
    scope_bits.append(f"expand={meta.get('expand')}")
    writer.writerow(
        _comment_row(
            f"# WORKSHEET — fill day/start_time/end_time/venue then Upload. "
            f"Column programmes is informational. Scope: {' '.join(scope_bits)} · "
            f"{len(units)} course unit(s)"
        )
    )

    claimed: set[int] = set()

    # ── A. Existing Shared Teaching ──────────────────────────────────────────
    sto_groups: dict[int, list[CourseUnit]] = {}
    for cu in units:
        sto_id = cu.shared_teaching_offering_id
        if sto_id:
            sto_groups.setdefault(sto_id, []).append(cu)

    writer.writerow(
        _comment_row(
            "# === A. Already Shared Teaching (STO) — keep shared_key; set one day/time/room ==="
        )
    )
    if not sto_groups:
        writer.writerow(_comment_row("# (none in this scope)"))
    for sto_id, group in sorted(sto_groups.items(), key=lambda x: x[0]):
        for cu in group:
            claimed.add(cu.id)
        codes = sorted({(cu.code or "").strip() for cu in group if (cu.code or "").strip()})
        sto = group[0].shared_teaching_offering
        name = (sto.name if sto else "") or (sto.code if sto else f"STO-{sto_id}")
        emails = sorted(
            {
                e
                for cu in group
                for e in _cu_lecturer_emails(cu).split(",")
                if e
            }
        )
        writer.writerow(
            _blank_schedule_row(
                course_codes=",".join(codes),
                programmes=_programmes_cell(group),
                start_date=start_date,
                end_date=end_date,
                shared_key=f"sto-{sto_id}",
                lecturer_email=",".join(emails),
                notes=f"STO {name}",
            )
        )

    # ── B. Cross-cutting candidates ──────────────────────────────────────────
    writer.writerow(
        _comment_row(
            "# === B. Cross-cutting candidates (catalog flag / shared paper#) — review shared_key ==="
        )
    )

    # Group unclaimed units by catalog_unit_id when cross-cutting, else by paper number
    # among units whose catalog is_cross_cutting OR same paper# across 2+ programmes.
    by_catalog: dict[int, list[CourseUnit]] = {}
    by_paper: dict[str, list[CourseUnit]] = {}
    for cu in units:
        if cu.id in claimed:
            continue
        cat = cu.catalog_unit
        if cat and cat.is_cross_cutting and cat.id:
            by_catalog.setdefault(cat.id, []).append(cu)
            continue
        paper = course_code_number(cu.code)
        if len(paper) >= 3:
            by_paper.setdefault(paper, []).append(cu)

    cross_rows = 0
    for cat_id, group in sorted(by_catalog.items(), key=lambda x: x[0]):
        if len(group) < 1:
            continue
        # Always emit cross-cutting catalog papers (even solo) so HOD sees them
        for cu in group:
            claimed.add(cu.id)
        codes = sorted({(cu.code or "").strip() for cu in group if (cu.code or "").strip()})
        cat = group[0].catalog_unit
        title = (cat.title if cat else "") or (cat.code if cat else "")
        note_extra = (cat.cross_cutting_note if cat else "") or ""
        emails = sorted(
            {e for cu in group for e in _cu_lecturer_emails(cu).split(",") if e}
        )
        key = f"cross-cat-{cat_id}"
        if len(codes) < 2:
            key = ""  # solo — no STO yet; still flagged in notes
        writer.writerow(
            _blank_schedule_row(
                course_codes=",".join(codes),
                programmes=_programmes_cell(group),
                start_date=start_date,
                end_date=end_date,
                shared_key=key,
                lecturer_email=",".join(emails),
                notes=(
                    f"CROSS-CUTTING catalog · {title}"
                    + (f" · {note_extra}" if note_extra else "")
                    + (
                        ""
                        if len(codes) >= 2
                        else " · single programme in scope — add peers or schedule alone"
                    )
                ),
            )
        )
        cross_rows += 1

    for paper, group in sorted(by_paper.items(), key=lambda x: x[0]):
        # Only multi-programme paper numbers (not already claimed via catalog)
        remaining = [cu for cu in group if cu.id not in claimed]
        prog_ids = {
            (_cu_program(cu).id if _cu_program(cu) else None) for cu in remaining
        }
        prog_ids.discard(None)
        if len(remaining) < 2 or len(prog_ids) < 2:
            continue
        for cu in remaining:
            claimed.add(cu.id)
        codes = sorted({(cu.code or "").strip() for cu in remaining if (cu.code or "").strip()})
        emails = sorted(
            {e for cu in remaining for e in _cu_lecturer_emails(cu).split(",") if e}
        )
        writer.writerow(
            _blank_schedule_row(
                course_codes=",".join(codes),
                programmes=_programmes_cell(remaining),
                start_date=start_date,
                end_date=end_date,
                shared_key=f"cross-paper-{paper}",
                lecturer_email=",".join(emails),
                notes=(
                    f"Same paper #{paper} across {len(prog_ids)} programmes — "
                    "confirm Shared Teaching"
                ),
            )
        )
        cross_rows += 1

    if cross_rows == 0:
        writer.writerow(_comment_row("# (none in this scope)"))

    # ── C. Programme-only ────────────────────────────────────────────────────
    writer.writerow(
        _comment_row("# === C. Programme-only offerings (one row per unit / stream) ===")
    )
    remaining = [cu for cu in units if cu.id not in claimed]
    remaining.sort(
        key=lambda cu: (
            (_cu_program(cu).short_form or "") if _cu_program(cu) else "",
            cu.code or "",
            cu.id,
        )
    )
    if not remaining:
        writer.writerow(
            _comment_row("# (none left — all units are shared or cross-cutting)")
        )
    for cu in remaining:
        prog = _cu_program(cu)
        batch = _cu_batch(cu)
        short = (prog.short_form or prog.code or "") if prog else ""
        batch_name = (batch.name if batch else "") or ""
        mode = study_mode_for_course_unit(cu)
        base_note = f"{batch_name} · {mode} · {cu.name or ''}".strip(" ·")
        emails = _cu_lecturer_emails(cu)
        # Prefer semester dates for this unit when exporting multi-semester scope
        sd = start_date
        ed = end_date
        if cu.semester and cu.semester.start_date:
            sd = cu.semester.start_date.isoformat()
        if cu.semester and cu.semester.end_date:
            ed = cu.semester.end_date.isoformat()
        for sec in _section_codes_for_unit(cu):
            note = base_note
            if sec:
                note = f"{base_note} · stream {sec}"
            writer.writerow(
                _blank_schedule_row(
                    course_codes=(cu.code or "").strip(),
                    programmes=_programmes_cell([cu]),
                    start_date=sd,
                    end_date=ed,
                    teaching_section=sec,
                    lecturer_email=emails,
                    notes=note,
                )
            )

    return buf.getvalue()
