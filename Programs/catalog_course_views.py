"""CRUD + bulk upload for CourseCatalogUnit (shared academic catalog; no programme/semester)."""

from __future__ import annotations

from decimal import Decimal

from django.db.models import Q
from rest_framework import generics, status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from django.db.utils import ProgrammingError
from .models import CourseCatalogUnit
from .serializers import CourseCatalogUnitSerializer


class CourseCatalogUnitListCreateView(generics.ListCreateAPIView):
    queryset = CourseCatalogUnit.objects.none()
    serializer_class = CourseCatalogUnitSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        try:
            qs = CourseCatalogUnit.objects.all().order_by("code", "title")
        except ProgrammingError:
            return CourseCatalogUnit.objects.none()
        search = (self.request.query_params.get("search") or "").strip()
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(title__icontains=search))
        active = self.request.query_params.get("is_active")
        if active is not None and active != "":
            if str(active).lower() in ("1", "true", "yes"):
                qs = qs.filter(is_active=True)
            elif str(active).lower() in ("0", "false", "no"):
                qs = qs.filter(is_active=False)
        return qs


class CourseCatalogUnitDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CourseCatalogUnit.objects.all()
    serializer_class = CourseCatalogUnitSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({"detail": "Course catalog unit deleted."}, status=status.HTTP_200_OK)


class BulkUploadCourseCatalogUnitsView(APIView):
    """
    Bulk upload CourseCatalogUnit rows from .xlsx or .csv.

    Expected columns (case-insensitive; spaces/underscores tolerated):
      - code (required)
      - title (required)
      - description
      - credit_units (required)
      - lecture_hours
      - practical_hours
      - tutorial_hours
      - contact_hours
      - is_active

    Optional form field:
      - update_existing: true/false (default false)
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @staticmethod
    def _norm_col(s: str) -> str:
        return (s or "").strip().lower().replace(" ", "_")

    @staticmethod
    def _as_bool(v) -> bool | None:
        if v is None or v == "":
            return None
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in ("1", "true", "yes", "y", "t"):
            return True
        if s in ("0", "false", "no", "n", "f"):
            return False
        return None

    @staticmethod
    def _as_int(v):
        if v is None or v == "":
            return None
        try:
            return int(v)
        except Exception:
            return None

    @staticmethod
    def _as_decimal(v):
        if v is None or v == "":
            return None
        try:
            return Decimal(str(v))
        except Exception:
            return None

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        update_existing = str(request.data.get("update_existing", "false")).strip().lower() in (
            "1",
            "true",
            "yes",
        )

        try:
            import pandas as pd
        except Exception:
            return Response(
                {"detail": "Server is missing pandas dependency for bulk upload."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            name = (file.name or "").lower()
            if name.endswith(".xlsx") or name.endswith(".xls"):
                df = pd.read_excel(file, engine="openpyxl")
            else:
                df = pd.read_csv(file, sep=None, engine="python")
        except Exception as e:
            return Response(
                {"detail": f"Failed to read file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if df is None or df.empty:
            return Response({"detail": "File has no rows."}, status=status.HTTP_400_BAD_REQUEST)

        df.columns = [self._norm_col(c) for c in df.columns]

        # accept synonyms
        col_map = {
            "code": "code",
            "course_code": "code",
            "unit_code": "code",
            "title": "title",
            "name": "title",
            "course_title": "title",
            "description": "description",
            "credit_units": "credit_units",
            "credits": "credit_units",
            "credit": "credit_units",
            "lecture_hours": "lecture_hours",
            "practical_hours": "practical_hours",
            "tutorial_hours": "tutorial_hours",
            "contact_hours": "contact_hours",
            "is_active": "is_active",
            "active": "is_active",
        }

        normalized = {}
        for c in df.columns:
            if c in col_map:
                normalized[col_map[c]] = c

        missing = [c for c in ("code", "title", "credit_units") if c not in normalized]
        if missing:
            return Response(
                {"detail": f"Missing required column(s): {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        failed = 0
        errors: list[str] = []

        for idx, row in df.iterrows():
            row_num = int(idx) + 2  # 1-based + header row
            code = str(row.get(normalized["code"], "")).strip()
            title = str(row.get(normalized["title"], "")).strip()
            credit_units = self._as_decimal(row.get(normalized["credit_units"]))

            if not code or not title or credit_units is None:
                failed += 1
                errors.append(f"Row {row_num}: code, title, and credit_units are required.")
                continue

            payload = {
                "code": code,
                "title": title,
                "description": str(row.get(normalized.get("description"), "") or "").strip(),
                "credit_units": str(credit_units),
                "lecture_hours": self._as_int(row.get(normalized.get("lecture_hours"))),
                "practical_hours": self._as_int(row.get(normalized.get("practical_hours"))),
                "tutorial_hours": self._as_int(row.get(normalized.get("tutorial_hours"))),
                "contact_hours": self._as_int(row.get(normalized.get("contact_hours"))),
            }
            is_active = self._as_bool(row.get(normalized.get("is_active")))
            if is_active is not None:
                payload["is_active"] = is_active

            existing = CourseCatalogUnit.objects.filter(code__iexact=code).first()
            if existing and not update_existing:
                failed += 1
                errors.append(f"Row {row_num}: code '{code}' already exists (enable update_existing to update).")
                continue

            serializer = CourseCatalogUnitSerializer(
                instance=existing,
                data=payload,
                partial=bool(existing),
            )
            if serializer.is_valid():
                serializer.save()
                if existing:
                    updated += 1
                else:
                    created += 1
            else:
                failed += 1
                errors.append(f"Row {row_num}: {serializer.errors}")

        return Response(
            {
                "created": created,
                "updated": updated,
                "failed": failed,
                "errors": errors[:200],
                "message": "Bulk upload completed.",
            },
            status=status.HTTP_200_OK,
        )


class CourseCatalogUnitTemplateDownloadView(APIView):
    """Download an Excel template for bulk uploading catalog course units."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception as e:
            return Response(
                {"detail": f"Excel dependency missing on server: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "catalog_course_units"

        headers = [
            "code",
            "title",
            "credit_units",
            "description",
            "lecture_hours",
            "practical_hours",
            "tutorial_hours",
            "contact_hours",
            "is_active",
        ]

        ws.append(headers)
        ws.append(
            [
                "CS101",
                "Introduction to Computer Science",
                3,
                "Optional description",
                30,
                15,
                0,
                "",
                True,
            ]
        )
        ws.append(
            [
                "MAT110",
                "Basic Mathematics",
                3,
                "",
                30,
                0,
                0,
                "",
                True,
            ]
        )

        # Style header
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{ws.max_row}"

        # Column widths
        widths = {
            "A": 14,  # code
            "B": 44,  # title
            "C": 12,  # credit_units
            "D": 40,  # description
            "E": 14,  # lecture_hours
            "F": 16,  # practical_hours
            "G": 14,  # tutorial_hours
            "H": 14,  # contact_hours
            "I": 10,  # is_active
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Notes sheet
        ns = wb.create_sheet("notes")
        ns.append(["Required columns:", "code, title, credit_units"])
        ns.append(["Optional columns:", "description, lecture_hours, practical_hours, tutorial_hours, contact_hours, is_active"])
        ns.append(["Notes:", "is_active accepts true/false, 1/0, yes/no"])
        ns.append(["Update behavior:", "If upload with update_existing=true, existing rows match by code (case-insensitive)."])

        from django.http import HttpResponse
        from datetime import date

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="catalog_course_units_template_{date.today().isoformat()}.xlsx"'
        wb.save(resp)
        return resp
