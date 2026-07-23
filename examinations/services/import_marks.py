"""Import CA and exam marks from Excel or CSV (reg_no, ca_mark, exam_mark)."""
from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from Programs.models import StudentCourseUnitEnrollment

from ..models import CourseUnitResult
from .policy_resolver import resolve_assessment_policy


def _cell_value(cell):
    if cell is None:
        return None
    v = cell.value
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    return v


def _decimal_or_none(raw):
    if raw is None:
        return None
    if isinstance(raw, str) and not raw.strip():
        return None
    try:
        return Decimal(str(raw).strip())
    except (InvalidOperation, ValueError):
        raise ValidationError(f"Invalid number: {raw!r}")


def _normalize_header(h: str) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _map_headers(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        key = _normalize_header(h)
        if key in ("reg_no", "regno", "registration", "registration_number"):
            col_map["reg_no"] = idx
        elif key in ("ca", "ca_mark", "continuous_assessment", "coursework"):
            col_map["ca_mark"] = idx
        elif key in ("exam", "exam_mark", "examination"):
            col_map["exam_mark"] = idx
        # student_name / name columns are ignored on import
    if "reg_no" not in col_map:
        raise ValidationError("File must have a reg_no (or regno) column.")
    return col_map


def parse_marks_workbook(file_bytes: bytes) -> list[dict]:
    """Return rows: {reg_no, ca_mark, exam_mark} from first sheet."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=False)
    header_cells = next(rows_iter, None)
    if not header_cells:
        raise ValidationError("Empty spreadsheet.")

    headers = [str(_cell_value(c) or "") for c in header_cells]
    col_map = _map_headers(headers)

    out = []
    for cells in rows_iter:
        reg = _cell_value(cells[col_map["reg_no"]] if col_map["reg_no"] < len(cells) else None)
        if not reg:
            continue
        row = {"reg_no": str(reg).strip()}
        if "ca_mark" in col_map and col_map["ca_mark"] < len(cells):
            row["ca_mark"] = _decimal_or_none(_cell_value(cells[col_map["ca_mark"]]))
        else:
            row["ca_mark"] = None
        if "exam_mark" in col_map and col_map["exam_mark"] < len(cells):
            row["exam_mark"] = _decimal_or_none(_cell_value(cells[col_map["exam_mark"]]))
        else:
            row["exam_mark"] = None
        out.append(row)
    wb.close()
    return out


def parse_marks_csv(file_bytes: bytes) -> list[dict]:
    """Return rows from CSV: reg_no, ca_mark, exam_mark (student_name ignored)."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        raise ValidationError("Empty CSV.")
    col_map = _map_headers(headers)

    out = []
    for cells in reader:
        if not cells:
            continue
        reg_idx = col_map["reg_no"]
        reg = cells[reg_idx].strip() if reg_idx < len(cells) else ""
        if not reg:
            continue
        row: dict = {"reg_no": reg}
        if "ca_mark" in col_map and col_map["ca_mark"] < len(cells):
            row["ca_mark"] = _decimal_or_none(cells[col_map["ca_mark"]])
        else:
            row["ca_mark"] = None
        if "exam_mark" in col_map and col_map["exam_mark"] < len(cells):
            row["exam_mark"] = _decimal_or_none(cells[col_map["exam_mark"]])
        else:
            row["exam_mark"] = None
        out.append(row)
    return out


def parse_marks_upload(filename: str, file_bytes: bytes) -> list[dict]:
    """Detect CSV vs Excel from filename and parse."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_marks_csv(file_bytes)
    if name.endswith((".xlsx", ".xls")):
        return parse_marks_workbook(file_bytes)
    sample = file_bytes[:64].lstrip()
    if sample.lower().startswith(b"reg") or b"," in sample[:200]:
        return parse_marks_csv(file_bytes)
    return parse_marks_workbook(file_bytes)


def build_marks_entry_csv(course_unit) -> tuple[str, str]:
    """
    CSV template for marks entry: reg_no, student_name, ca_mark, exam_mark.
    Pre-fills existing marks when present.
    Returns (filename, csv_text).
    """
    enrollments = (
        StudentCourseUnitEnrollment.objects.filter(
            course_unit=course_unit,
            status="enrolled",
        )
        .select_related("student", "student__application", "course_result")
        .order_by("student__reg_no")
    )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["reg_no", "student_name", "ca_mark", "exam_mark"])

    for enr in enrollments:
        application = getattr(enr.student, "application", None)
        if application is not None and getattr(application, "is_revoked", False):
            continue
        result = getattr(enr, "course_result", None)
        try:
            student_name = enr.student.full_name or ""
        except Exception:
            student_name = ""
        ca = ""
        exam = ""
        if result is not None:
            if result.ca_mark is not None:
                ca = str(result.ca_mark)
            if result.exam_mark is not None:
                exam = str(result.exam_mark)
        writer.writerow(
            [
                enr.student.reg_no or "",
                student_name,
                ca,
                exam,
            ]
        )

    code = (course_unit.code or "course").replace("/", "-")
    filename = f"marks_entry_{code}.csv"
    # BOM so Excel opens UTF-8 names correctly
    return filename, "\ufeff" + buf.getvalue()


def import_marks_for_course(course_unit, rows: list[dict], *, user) -> dict:
    """Apply import rows; returns {saved, errors}."""
    if not resolve_assessment_policy(course_unit=course_unit):
        raise ValidationError("No assessment policy configured.")

    saved = []
    errors = []

    for row in rows:
        reg = row.get("reg_no", "")
        try:
            enrollment = StudentCourseUnitEnrollment.objects.select_related(
                "student",
                "student__application",
                "course_result",
                "course_unit",
                "course_unit__program_batch__program__academic_level",
            ).get(
                course_unit=course_unit,
                status="enrolled",
                student__reg_no__iexact=reg,
            )
        except StudentCourseUnitEnrollment.DoesNotExist:
            errors.append({"reg_no": reg, "detail": "Student not enrolled on this course."})
            continue

        if enrollment.student.application and enrollment.student.application.is_revoked:
            errors.append({"reg_no": reg, "detail": "Admission revoked."})
            continue

        policy = resolve_assessment_policy(enrollment=enrollment)
        if not policy:
            errors.append({"reg_no": reg, "detail": "No assessment policy configured."})
            continue

        result, _ = CourseUnitResult.objects.get_or_create(
            enrollment=enrollment,
            defaults={"policy": policy, "entered_by": user},
        )

        if result.status == CourseUnitResult.STATUS_PUBLISHED and not result.edit_unlocked:
            errors.append(
                {"reg_no": reg, "detail": "Published — request a grade change or unlock first."}
            )
            continue

        result.policy = policy
        if row.get("ca_mark") is not None:
            result.ca_mark = row["ca_mark"]
        if row.get("exam_mark") is not None:
            result.exam_mark = row["exam_mark"]
        result.entered_by = user
        if result.status == CourseUnitResult.STATUS_PUBLISHED:
            result.status = CourseUnitResult.STATUS_VERIFIED
        elif result.status != CourseUnitResult.STATUS_VERIFIED:
            result.status = CourseUnitResult.STATUS_DRAFT

        try:
            result.recompute()
            result.full_clean()
            result.save()
            saved.append(reg)
        except Exception as exc:
            errors.append({"reg_no": reg, "detail": str(exc)})

    return {"saved_count": len(saved), "saved": saved, "errors": errors}
