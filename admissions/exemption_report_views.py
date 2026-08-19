"""Course exemption report JSON + Excel."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.permissions import BasePermission, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.erp_drf_permissions import user_has_any_erp_perm
from accounts.super_admin import user_is_super_admin
from admissions.exemption_report import build_exemption_report, exemption_report_filter_options


class ExemptionReportPermission(BasePermission):
    message = "You do not have permission to view the exemption report."

    def has_permission(self, request, view):
        u = request.user
        if not u or not u.is_authenticated:
            return False
        if user_is_super_admin(u):
            return True
        if u.has_perm("admissions.view_admissionchangerequest"):
            return True
        if u.has_perm("admissions.approve_exemption_requests"):
            return True
        if u.has_perm("AdmissionReports.view_admissionreports"):
            return True
        return user_has_any_erp_perm(u, "access_reports", "access_finance")


def _parse_int(raw) -> int | None:
    try:
        value = int(str(raw or "").strip())
        return value if value > 0 else None
    except (TypeError, ValueError):
        return None


def _params(request) -> dict:
    from django.utils.dateparse import parse_date

    return {
        "campus_id": _parse_int(request.query_params.get("campus")),
        "faculty_id": _parse_int(request.query_params.get("faculty")),
        "status": (request.query_params.get("status") or "").strip().lower(),
        "from_date": parse_date((request.query_params.get("from_date") or "").strip() or "") or None,
        "to_date": parse_date((request.query_params.get("to_date") or "").strip() or "") or None,
    }


class ExemptionReportView(APIView):
    permission_classes = [IsAuthenticated, ExemptionReportPermission]

    def get(self, request):
        params = _params(request)
        data = build_exemption_report(request.user, params)
        data["filters"] = exemption_report_filter_options(request.user)
        data["applied"] = {
            "campus": params["campus_id"],
            "faculty": params["faculty_id"],
            "status": params["status"] or None,
            "from_date": params["from_date"].isoformat() if params["from_date"] else None,
            "to_date": params["to_date"].isoformat() if params["to_date"] else None,
        }
        return Response(data)


class ExemptionReportExcelView(APIView):
    permission_classes = [IsAuthenticated, ExemptionReportPermission]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        params = _params(request)
        data = build_exemption_report(request.user, params)
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e1b4b", fill_type="solid")
        thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        def write_sheet(ws, headers, rows):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin
                ws.column_dimensions[get_column_letter(col)].width = max(14, min(36, len(str(header)) + 4))
            for r_i, row in enumerate(rows, 2):
                for c_i, value in enumerate(row, 1):
                    cell = ws.cell(row=r_i, column=c_i, value=value)
                    cell.border = thin
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        t = data["totals"]
        summary = wb.active
        summary.title = "Summary"
        write_sheet(
            summary,
            ["Metric", "Count"],
            [
                ["Submitted", t["submitted"]],
                ["Pending HOD", t["pending"]],
                ["Approved", t["approved"]],
                ["Rejected", t["rejected"]],
                ["Paid, not submitted", t["paid_unsubmitted"]],
            ],
        )

        headers = [
            "Student",
            "Student ID",
            "Reg no",
            "Campus",
            "Programme",
            "Intake",
            "Status",
            "Papers",
            "Form fee paid",
            "Submitted at",
            "Submitted by",
            "Reviewed by",
            "Reviewed at",
        ]

        def row_values(r):
            return [
                r.get("name") or "—",
                r.get("student_id") or "",
                r.get("reg_no") or "",
                r.get("campus") or "—",
                r.get("program") or "—",
                r.get("intake") or "—",
                r.get("status") or "",
                r.get("papers") or 0,
                "Yes" if r.get("form_fee_paid") else "No",
                (r.get("submitted_at") or "")[:19].replace("T", " "),
                r.get("submitted_by") or "",
                r.get("reviewed_by") or "",
                (r.get("reviewed_at") or "")[:19].replace("T", " "),
            ]

        write_sheet(
            wb.create_sheet("Applications"),
            headers,
            [row_values(r) for r in data["applications"]],
        )
        write_sheet(
            wb.create_sheet("Paid not submitted"),
            headers,
            [row_values(r) for r in data["paid_unsubmitted"]],
        )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"exemption_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
