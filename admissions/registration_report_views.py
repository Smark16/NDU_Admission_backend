"""Registration report JSON + Excel."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.permissions import BasePermission, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.erp_drf_permissions import user_has_any_erp_perm
from accounts.super_admin import user_is_super_admin
from admissions.registration_report import (
    active_intake,
    build_registration_report,
    registration_report_filter_options,
)


class RegistrationReportPermission(BasePermission):
    message = "You do not have permission to view the registration report."

    def has_permission(self, request, view):
        u = request.user
        if not u or not u.is_authenticated:
            return False
        if user_is_super_admin(u):
            return True
        if u.has_perm("admissions.view_admittedstudent"):
            return True
        if u.has_perm("AdmissionReports.view_admissionreports"):
            return True
        if u.has_perm("admissions.verify_physical_documents"):
            return True
        return user_has_any_erp_perm(u, "access_reports", "access_finance")


def _parse_int(raw) -> int | None:
    try:
        value = int(str(raw or "").strip())
        return value if value > 0 else None
    except (TypeError, ValueError):
        return None


def _report_params(request) -> dict:
    from django.utils.dateparse import parse_date

    basis = (request.query_params.get("date_basis") or "verified").strip().lower()
    if basis not in ("admission", "registered", "cleared", "verified"):
        basis = "verified"

    raw_year = (request.query_params.get("academic_year") or "").strip()
    batch_id = _parse_int(request.query_params.get("batch"))
    academic_year = ""
    if raw_year.lower() == "all":
        academic_year = ""
    elif raw_year:
        academic_year = raw_year
    elif not batch_id:
        # Current academic year, all intakes — same slice as a year on the
        # verified roster (26/1 and 26/2 students sit in one 700 figure).
        intake = active_intake()
        if intake:
            academic_year = (intake.academic_year or "").strip()

    return {
        "academic_year": academic_year,
        "batch_id": batch_id,
        "admission_period": (request.query_params.get("admission_period") or "").strip(),
        "campus_id": _parse_int(request.query_params.get("campus")),
        "faculty_id": _parse_int(request.query_params.get("faculty")),
        "date_basis": basis,
        "from_date": parse_date((request.query_params.get("from_date") or "").strip() or "") or None,
        "to_date": parse_date((request.query_params.get("to_date") or "").strip() or "") or None,
    }


def _ugx(n) -> str:
    if n is None:
        return ""
    return f"{float(n):,.0f}"


class RegistrationReportView(APIView):
    permission_classes = [IsAuthenticated, RegistrationReportPermission]

    def get(self, request):
        from accounts.finance_access import user_can_view_student_finance

        params = _report_params(request)
        include_finance = user_can_view_student_finance(request.user)
        try:
            data = build_registration_report(request.user, params, include_finance=include_finance)
            data["filters"] = registration_report_filter_options(request.user)
        except Exception:
            import logging

            logging.getLogger(__name__).exception("registration report failed")
            return Response(
                {"detail": "Could not build the registration report. Check gunicorn logs."},
                status=500,
            )
        data["applied"] = {
            "academic_year": params["academic_year"] or None,
            "batch": params["batch_id"],
            "admission_period": params.get("admission_period") or None,
            "campus": params["campus_id"],
            "faculty": params["faculty_id"],
            "date_basis": params["date_basis"],
            "from_date": params["from_date"].isoformat() if params["from_date"] else None,
            "to_date": params["to_date"].isoformat() if params["to_date"] else None,
        }
        return Response(data)


class RegistrationReportExcelView(APIView):
    permission_classes = [IsAuthenticated, RegistrationReportPermission]

    def get(self, request):
        from accounts.finance_access import user_can_view_student_finance
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        params = _report_params(request)
        include_finance = user_can_view_student_finance(request.user)
        data = build_registration_report(request.user, params, include_finance=include_finance)
        export = (request.query_params.get("export") or "all").strip().lower()
        allowed = {
            "all",
            "summary",
            "reported",
            "verified",
            "enrolled",
            "cleared",
            "registered",
            "temp_passes",
            "scholarships",
        }
        if export not in allowed:
            export = "all"

        def want(*names: str) -> bool:
            return export == "all" or export in names

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e1b4b", fill_type="solid")
        thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        def write_sheet(ws, headers, rows):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin
                ws.column_dimensions[get_column_letter(col)].width = max(14, min(32, len(str(header)) + 4))
            for r_i, row in enumerate(rows, 2):
                for c_i, value in enumerate(row, 1):
                    cell = ws.cell(row=r_i, column=c_i, value=value)
                    cell.border = thin
                    cell.alignment = Alignment(vertical="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        t = data["totals"]
        summary = wb.active
        summary.title = "Summary"
        write_sheet(
            summary,
            ["Metric", "Count", "% of admitted"],
            [
                ["Admitted", t["admitted"], 100 if t["admitted"] else None],
                ["Reported (cleared + temp pass + scholarship)", t.get("reported"), t.get("reported_pct")],
                ["Verified roster", t["verified"], t["verified_pct"]],
                ["Enrolled (programme)", t.get("enrolled"), t.get("enrolled_pct")],
                ["Registered (courses)", t["registered"], t["registered_pct"]],
                ["Accounts cleared", t["cleared"], t["clearance_pct"]],
                ["Temporary passes", t["temporary_passes"], None],
                ["Scholarship students", t["scholarships"], None],
            ],
        )
        summary.auto_filter.ref = "A1:C10"
        summary["E1"] = "Applied filters"
        summary["E1"].font = Font(bold=True)
        filter_rows = [
            ("Export", export),
            ("Academic year", params["academic_year"] or "All years"),
            ("Intake batch id", params["batch_id"] or "All intakes"),
            ("Admission period", params.get("admission_period") or "—"),
            ("Campus id", params["campus_id"] or "All campuses"),
            ("Faculty id", params["faculty_id"] or "All faculties"),
            ("Date basis", params["date_basis"]),
            ("From date", params["from_date"].isoformat() if params["from_date"] else "—"),
            ("To date", params["to_date"].isoformat() if params["to_date"] else "—"),
        ]
        for i, (label, value) in enumerate(filter_rows, 2):
            summary.cell(row=i, column=5, value=label)
            summary.cell(row=i, column=6, value=str(value))
        summary.column_dimensions["E"].width = 20
        summary.column_dimensions["F"].width = 24

        if want("reported"):
            write_sheet(
                wb.create_sheet("Reported students"),
            [
                "Student",
                "Student ID",
                "Reg No",
                "Campus",
                "Programme",
                "Accounts cleared",
                "Temp pass",
                "Scholarship",
                "Courses registered",
            ],
            [
                [
                    r.get("name") or "—",
                    r.get("student_id") or "",
                    r.get("reg_no") or "",
                    r.get("campus") or "—",
                    r.get("program") or "—",
                    "Y" if r.get("accounts_cleared") else "N",
                    "Y" if r.get("temp_pass") else "N",
                    "Y" if r.get("scholarship") else "N",
                    "Y" if r.get("is_registered") else "N",
                ]
                for r in (data.get("reported_students") or [])
            ],
            )

        if want("verified"):
            write_sheet(
                wb.create_sheet("Verified roster"),
            [
                "Student",
                "Student ID",
                "Reg No",
                "Campus",
                "Programme",
                "Verified by",
                "Verified at",
                "Courses registered",
            ],
            [
                [
                    r.get("name") or "—",
                    r.get("student_id") or "",
                    r.get("reg_no") or "",
                    r.get("campus") or "—",
                    r.get("program") or "—",
                    r.get("verified_by") or "",
                    (r.get("verified_at") or "")[:19].replace("T", " "),
                    "Y" if r.get("is_registered") else "N",
                ]
                for r in (data.get("verified_students") or [])
            ],
            )

        if want("enrolled"):
            write_sheet(
                wb.create_sheet("Enrolled students"),
                [
                    "Student",
                    "Student ID",
                    "Reg No",
                    "Campus",
                    "Programme",
                    "Year of study",
                    "Term",
                    "Cohort",
                    "Courses registered",
                ],
                [
                    [
                        r.get("name") or "—",
                        r.get("student_id") or "",
                        r.get("reg_no") or "",
                        r.get("campus") or "—",
                        r.get("program") or "—",
                        r.get("year_of_study") or "",
                        r.get("term") or "",
                        r.get("cohort") or "—",
                        "Y" if r.get("is_registered") else "N",
                    ]
                    for r in (data.get("enrolled_students") or [])
                ],
            )

        if want("cleared"):
            write_sheet(
                wb.create_sheet("Cleared students"),
            [
                "Student",
                "Student ID",
                "Reg No",
                "Campus",
                "Programme",
                "Accounts cleared by",
            ],
            [
                [
                    r.get("name") or "—",
                    r.get("student_id") or "",
                    r.get("reg_no") or "",
                    r.get("campus") or "—",
                    r.get("program") or "—",
                    r.get("cleared_by") or "",
                ]
                for r in (data.get("cleared_students") or [])
            ],
            )

        if want("registered"):
            write_sheet(
                wb.create_sheet("Registered students"),
                ["Student", "Student ID", "Reg No", "Campus", "Programme"],
                [
                    [
                        r.get("name") or "—",
                        r.get("student_id") or "",
                        r.get("reg_no") or "",
                        r.get("campus") or "—",
                        r.get("program") or "—",
                    ]
                    for r in (data.get("registered_students") or [])
                ],
            )

        if want("summary"):
            write_sheet(
                wb.create_sheet("By campus"),
                ["Campus", "Admitted", "Reported", "Enrolled", "Verified", "Registered", "Cleared"],
                [
                    [
                        r.get("campus") or "—",
                        r["admitted"],
                        r.get("reported", 0),
                        r.get("enrolled", 0),
                        r.get("verified", 0),
                        r["registered"],
                        r["cleared"],
                    ]
                    for r in data["by_campus"]
                ],
            )
            write_sheet(
                wb.create_sheet("By programme"),
                ["Faculty", "Programme", "Admitted", "Reported", "Enrolled", "Verified", "Registered", "Cleared"],
                [
                    [
                        r.get("faculty") or "—",
                        r.get("program") or "—",
                        r["admitted"],
                        r.get("reported", 0),
                        r.get("enrolled", 0),
                        r.get("verified", 0),
                        r["registered"],
                        r["cleared"],
                    ]
                    for r in data["by_program"]
                ],
            )

        if want("temp_passes"):
            pass_headers = ["Name", "Student ID", "Reg No", "Campus", "Programme", "Intake", "Sponsor", "Valid until"]
            pass_rows = []
            for r in data["temporary_passes"]:
                row = [
                    r["name"],
                    r["student_id"],
                    r["reg_no"],
                    r["campus"],
                    r["program"],
                    r["intake"],
                    r["sponsor"],
                    r["valid_until"] or "Open",
                ]
                if include_finance:
                    row.append(_ugx(r.get("tuition_paid_ugx")))
                pass_rows.append(row)
            if include_finance:
                pass_headers.append("Tuition paid (UGX)")
            write_sheet(wb.create_sheet("Temporary passes"), pass_headers, pass_rows)

        if want("scholarships"):
            sch_headers = [
                "Name",
                "Student ID",
                "Reg No",
                "Campus",
                "Programme",
                "Intake",
                "Scholarship",
                "Sponsor",
                "Award amount",
            ]
            sch_rows = []
            for r in data["scholarships"]:
                row = [
                    r["name"],
                    r["student_id"],
                    r["reg_no"],
                    r["campus"],
                    r["program"],
                    r["intake"],
                    r["scholarship_name"],
                    r["sponsor"],
                    r.get("award_amount"),
                ]
                if include_finance:
                    row.append(_ugx(r.get("tuition_paid_ugx")))
                sch_rows.append(row)
            if include_finance:
                sch_headers.append("Tuition paid (UGX)")
            write_sheet(wb.create_sheet("Scholarships"), sch_headers, sch_rows)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"registration_report_{export}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
