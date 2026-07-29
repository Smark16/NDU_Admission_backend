"""Bulk import admitted students into a programme batch (academic cohort)."""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import date, datetime, timedelta
from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError as DRFValidationError

from accounts.models import Campus, User
from admissions.models import AcademicLevel, AdmittedStudent, Application, Batch
from admissions.serializers import AdmittedStudentSerializer
from admissions.utils.program_choices import sync_application_program_choices
from admissions.utils.reference import generate_reference
from admissions.utils.trigger_background_tasks import queue_admission_notification_emails
from admissions.utils.student_portal_provisioning import (
    StudentPortalProvisioningError,
    provision_student_portal_on_admission,
)
from payments.utils.school_pay_code import _schoolpay_phone
from Programs.models import Program, ProgramBatch
from Programs.specialization_rules import resolve_specialization_for_program

logger = logging.getLogger(__name__)

STUDENT_IMPORT_REQUIRED_HEADERS = [
    "first_name",
    "last_name",
    "middle_name",
    "email",
    "phone",
    "date_of_birth",
    "gender",
    "nationality",
    "reg_no",
    "study_mode",
    "student_id",
    "specialization",
    "address",
]

STUDENT_IMPORT_OPTIONAL_HEADERS = [
    "current_year_of_study",
    "current_term_number",
    "fees_paid_ugx",
    "fees_paid_reference",
    "fees_outstanding_ugx",
    "admission_fee_paid",
]

# Backward-compatible alias for API responses (required columns only).
STUDENT_IMPORT_HEADERS = STUDENT_IMPORT_REQUIRED_HEADERS

STUDENT_IMPORT_TEMPLATE_HEADERS = (
    STUDENT_IMPORT_REQUIRED_HEADERS + STUDENT_IMPORT_OPTIONAL_HEADERS
)

STUDY_MODES = frozenset({"W", "D", "DL", "DJ", "WJ"})


def _parse_optional_position(row: dict, program: Program) -> tuple[int, int] | None:
    """Parse continuing-student year/term; both columns required when either is set."""
    raw_year = (row.get("current_year_of_study") or "").strip()
    raw_term = (row.get("current_term_number") or "").strip()
    if not raw_year and not raw_term:
        return None
    if not raw_year or not raw_term:
        raise ValueError(
            "current_year_of_study and current_term_number must both be set together."
        )
    try:
        year = int(raw_year)
        term = int(raw_term)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "current_year_of_study and current_term_number must be integers."
        ) from exc
    if year < 1 or year > program.max_years:
        raise ValueError(
            f"current_year_of_study must be between 1 and {program.max_years}."
        )
    max_terms = program.max_terms_per_year
    if term not in range(1, max_terms + 1):
        raise ValueError(
            f"current_term_number must be between 1 and {max_terms} "
            f"for a {program.calendar_type}-based programme."
        )
    return year, term


def _resolve_curriculum_version(program: Program, program_batch: ProgramBatch):
    from Programs.models import resolve_program_default_curriculum_version

    if program_batch.curriculum_version_id:
        return program_batch.curriculum_version
    return resolve_program_default_curriculum_version(program)


def _upsert_programme_enrollment_from_import(
    admitted: AdmittedStudent,
    *,
    program: Program,
    program_batch: ProgramBatch,
    year_of_study: int,
    term_number: int,
    admitted_by,
    specialization: str | None = None,
) -> dict:
    from Programs.models import StudentProgrammeEnrollment
    from payments.admin_enrollment_requirements import (
        admin_programme_enrollment_activation_block,
    )

    curriculum_version = _resolve_curriculum_version(program, program_batch)
    if curriculum_version is None:
        raise ValueError(
            "No curriculum version is configured for this programme — "
            "cannot set current year/semester."
        )

    activation_block = admin_programme_enrollment_activation_block(
        admitted, target_status="enrolled"
    )
    enroll_status = "enrolled" if activation_block is None else "pending"

    enrollment, created = StudentProgrammeEnrollment.objects.get_or_create(
        student=admitted,
        defaults={
            "program": program,
            "program_batch": program_batch,
            "curriculum_version": curriculum_version,
            "current_year_of_study": year_of_study,
            "current_term_number": term_number,
            "specialization": specialization,
            "status": enroll_status,
            "enrolled_by": admitted_by if enroll_status == "enrolled" else None,
            "notes": "Bulk import — continuing student position.",
        },
    )
    if not created:
        enrollment.program = program
        enrollment.program_batch = program_batch
        enrollment.curriculum_version = curriculum_version
        enrollment.current_year_of_study = year_of_study
        enrollment.current_term_number = term_number
        if activation_block is None:
            enrollment.status = "enrolled"
            enrollment.enrolled_by = admitted_by
        elif enrollment.status == "pending":
            pass
        else:
            enrollment.status = enroll_status
        if specialization:
            enrollment.specialization = specialization
        enrollment.save()

    return {
        "enrollment_set": True,
        "enrollment_created": created,
        "enrollment_status": enrollment.status,
        "enrollment_blocked": activation_block,
        "current_year_of_study": year_of_study,
        "current_term_number": term_number,
    }


def _refresh_enrollment_status_result(admitted: AdmittedStudent, result: dict) -> None:
    """Attach final SPE status / activation block after fees + year upsert."""
    from payments.admin_enrollment_requirements import (
        admin_programme_enrollment_activation_block,
    )
    from Programs.models import StudentProgrammeEnrollment

    enrollment = (
        StudentProgrammeEnrollment.objects.filter(student_id=admitted.pk).first()
    )
    if enrollment is None:
        result.setdefault("enrollment_status", None)
        result.setdefault("enrollment_blocked", None)
        return

    result["enrollment_status"] = enrollment.status
    if enrollment.status == "enrolled":
        result["enrollment_activated"] = True
        result["enrollment_blocked"] = None
    else:
        result["enrollment_blocked"] = admin_programme_enrollment_activation_block(
            admitted, target_status="enrolled"
        )


def _apply_import_extensions(
    admitted: AdmittedStudent,
    *,
    program: Program,
    program_batch: ProgramBatch,
    row: dict,
    admitted_by,
    specialization: str | None = None,
    require_academic_position: bool = False,
) -> dict:
    """
    Continuing-student fields: academic position then legacy fee balances, then activate.

    Order matters: set year/term before fee-driven activation so SPE is not created at Y1/S1.
    """
    from admissions.student_fee_balance_import import (
        apply_legacy_fee_balances,
        row_has_legacy_fee_data,
    )
    from payments.programme_enrollment_activation import (
        activate_programme_enrollment_after_commitment_payment,
    )

    result = {
        "enrollment_set": False,
        "current_year_of_study": None,
        "current_term_number": None,
        "enrollment_status": None,
        "enrollment_blocked": None,
        "fees_paid_recorded": False,
        "fees_outstanding_recorded": False,
        "admission_fee_paid_set": False,
        "enrollment_activated": False,
        "extensions_applied": False,
    }

    position = _parse_optional_position(row, program)
    if require_academic_position and position is None:
        raise ValueError(
            "current_year_of_study and current_term_number are required "
            "for continuing-student import."
        )

    if position is not None:
        year, term = position
        enrollment_info = _upsert_programme_enrollment_from_import(
            admitted,
            program=program,
            program_batch=program_batch,
            year_of_study=year,
            term_number=term,
            admitted_by=admitted_by,
            specialization=specialization,
        )
        result.update(enrollment_info)
        result["extensions_applied"] = True

    if row_has_legacy_fee_data(row):
        fee_result = apply_legacy_fee_balances(
            admitted, row, admitted_by=admitted_by
        )
        result.update(fee_result)
        result["extensions_applied"] = True

    # Re-run activation after year/term + fees so position is preserved and status is final.
    if result["extensions_applied"]:
        activation = activate_programme_enrollment_after_commitment_payment(
            admitted, activated_by=admitted_by
        )
        if activation.get("activated") or activation.get("reason") == "already_enrolled":
            result["enrollment_activated"] = True
        _refresh_enrollment_status_result(admitted, result)

    return result


def _preflight_continuing_import(program: Program, program_batch: ProgramBatch) -> None:
    """Fail early when curriculum or semester fees are missing for the academic batch."""
    from payments.models import FeePlanRule

    curriculum_version = _resolve_curriculum_version(program, program_batch)
    if curriculum_version is None:
        raise ValueError(
            f"No curriculum version is configured for programme '{program.name}' "
            f"(batch '{program_batch.name}'). Assign a curriculum on the programme or batch "
            "before importing continuing students."
        )

    if not FeePlanRule.objects.filter(program_batch_id=program_batch.id).exists():
        raise ValueError(
            f"No semester fees are configured for academic batch '{program_batch.name}'. "
            "Set up Batch Semester Fees (tuition/functional) before importing continuing students."
        )


def _normalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", (name or "").strip().lower())


def _cell_to_import_value(val) -> str:
    """Normalize Excel/CSV cell values; dates from Excel often arrive as datetime or serial numbers."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    text = str(val).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _parse_date_of_birth(raw) -> date:
    """Accept YYYY-MM-DD, DD/MM/YYYY (common in Uganda), Excel serials, and datetime strings."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    text = str(raw or "").strip()
    if not text:
        raise ValueError("date_of_birth is required.")

    if text.endswith(".0") and text[:-2].replace(".", "", 1).isdigit():
        text = text[:-2]

    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            serial = float(text)
            if 1 <= serial <= 120000:
                excel_epoch = date(1899, 12, 30)
                return excel_epoch + timedelta(days=int(serial))
        except (ValueError, OverflowError):
            pass

    if "T" in text:
        try:
            return datetime.fromisoformat(text[:19]).date()
        except ValueError:
            pass

    date_part = text.split()[0] if " " in text else text

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%m/%d/%Y",
        "%m/%d/%y",
    ):
        try:
            return datetime.strptime(date_part, fmt).date()
        except ValueError:
            continue

    raise ValueError(
        f"date_of_birth '{text}' is not recognized. Use YYYY-MM-DD, DD/MM/YYYY, or a valid Excel date."
    )


def _parse_upload_file(uploaded_file) -> tuple[list[str], list[dict]]:
    raw = uploaded_file.read()
    if not raw:
        raise ValueError("Uploaded file is empty.")

    name = (getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel support requires openpyxl on the server.") from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel sheet is empty.")
        headers = [_normalize_header(str(c or "")) for c in header_row]
        rows = []
        for idx, cells in enumerate(rows_iter, start=2):
            if not any(c is not None and str(c).strip() for c in cells):
                continue
            row = {}
            for col_idx, key in enumerate(headers):
                if not key:
                    continue
                val = cells[col_idx] if col_idx < len(cells) else None
                row[key] = _cell_to_import_value(val)
            rows.append({"__row__": idx, **row})
        return headers, rows

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row.")
    headers = [_normalize_header(h) for h in reader.fieldnames]
    rows = []
    for idx, raw_row in enumerate(reader, start=2):
        if not any(str(v or "").strip() for v in raw_row.values()):
            continue
        row = {"__row__": idx}
        for orig, norm in zip(reader.fieldnames, headers):
            if norm:
                row[norm] = (raw_row.get(orig) or "").strip()
        rows.append(row)
    return headers, rows


CONTINUING_INTAKE_CODE = "CONTINUING"
CONTINUING_INTAKE_NAME = "Continuing / Legacy Students"


def get_or_create_continuing_admission_intake(*, created_by) -> Batch:
    """
    Dedicated admission intake for bulk-imported / continuing students.

    Kept inactive so it never appears as the live offer window. Academic placement
    still uses ``ProgramBatch``; this intake is only an admissions tag.
    """
    existing = Batch.objects.filter(code=CONTINUING_INTAKE_CODE).first()
    if existing:
        # Ensure it stays out of the live offer pool
        dirty = False
        if existing.is_active:
            existing.is_active = False
            dirty = True
        if existing.name != CONTINUING_INTAKE_NAME:
            existing.name = CONTINUING_INTAKE_NAME
            dirty = True
        if dirty:
            existing.save(update_fields=["is_active", "name", "updated_at"])
        return existing

    today = timezone.now().date()
    # Closed historical window — not an open application/offer period
    start = today.replace(year=max(today.year - 10, 2000), month=1, day=1)
    end = today.replace(year=max(today.year - 1, 2001), month=12, day=31)

    return Batch.objects.create(
        name=CONTINUING_INTAKE_NAME,
        code=CONTINUING_INTAKE_CODE,
        academic_year="",
        application_start_date=start,
        application_end_date=end,
        admission_start_date=start,
        admission_end_date=end,
        offer_start_date=None,
        offer_end_date=end,
        is_active=False,
        created_by=created_by,
    )


def _resolve_admission_intake_batch(
    admission_batch_id: int | None,
    *,
    created_by=None,
) -> Batch:
    """
    Resolve admissions intake for bulk import.

    - Explicit ``admission_batch_id`` → that intake
    - Otherwise → Continuing / Legacy intake (never the live offer intake)
    """
    if admission_batch_id is not None:
        try:
            return Batch.objects.get(pk=admission_batch_id)
        except Batch.DoesNotExist as exc:
            raise ValueError("Admission intake batch not found.") from exc

    if created_by is None:
        raise ValueError(
            "Cannot resolve Continuing intake without a user. "
            "Pass admission_batch_id or ensure the importer is authenticated."
        )
    return get_or_create_continuing_admission_intake(created_by=created_by)

def _require_columns(headers: list[str]) -> list[str]:
    missing = []
    for col in (
        "first_name",
        "last_name",
        "email",
        "phone",
        "date_of_birth",
        "gender",
        "nationality",
        "reg_no",
        "study_mode",
    ):
        if col not in headers:
            missing.append(col)
    return missing


def _link_legacy_schoolpay_code(admitted: AdmittedStudent, paycode: str) -> None:
    """Preserve an existing SchoolPay payment code from the legacy system."""
    code = paycode.strip()
    if not code:
        raise ValueError("Legacy student_id (SchoolPay payment code) cannot be empty.")
    if (
        AdmittedStudent.objects.filter(student_id=code)
        .exclude(pk=admitted.pk)
        .exists()
    ):
        raise ValueError(f"SchoolPay payment code '{code}' is already linked to another student.")
    if (
        AdmittedStudent.objects.filter(schoolpay_code=code)
        .exclude(pk=admitted.pk)
        .exists()
    ):
        raise ValueError(f"SchoolPay payment code '{code}' is already in use.")

    admitted.student_id = code
    admitted.schoolpay_code = code
    admitted.is_registered_with_schoolpay = True
    admitted.save(
        update_fields=[
            "student_id",
            "schoolpay_code",
            "is_registered_with_schoolpay",
            "updated_at",
        ]
    )


def _ensure_schoolpay_protection(
    admitted: AdmittedStudent,
    row: dict,
    *,
    register_schoolpay: bool,
) -> str:
    """
    Every imported student must have a SchoolPay payment code (student_id) so tuition
    is tied to the correct wallet. Legacy rows supply student_id in the file; new rows
    are registered via the SchoolPay API (same as direct admission).
    """
    legacy_paycode = row.get("student_id", "").strip()
    if legacy_paycode:
        _link_legacy_schoolpay_code(admitted, legacy_paycode)
        return legacy_paycode

    if not register_schoolpay:
        raise ValueError(
            "SchoolPay registration is required when student_id is not provided. "
            "Add the legacy payment code in the student_id column, or allow SchoolPay registration."
        )

    if admitted.is_registered_with_schoolpay and (admitted.student_id or "").strip():
        return admitted.student_id.strip()

    from payments.utils.school_pay_code import register_student_with_schoolpay

    result = register_student_with_schoolpay(admitted)
    if not result.get("success"):
        raise ValueError(
            f"SchoolPay registration failed: {result.get('error') or 'unknown error'}"
        )
    admitted.refresh_from_db()
    paycode = (admitted.student_id or "").strip()
    if not paycode:
        raise ValueError("SchoolPay registration succeeded but no payment code was stored.")
    return paycode


def _import_one_row(
    *,
    row: dict,
    program: Program,
    program_batch: ProgramBatch,
    admission_batch: Batch,
    campus: Campus,
    academic_level: AcademicLevel,
    admitted_by,
    register_schoolpay: bool,
    require_academic_position: bool = False,
) -> AdmittedStudent:
    first_name = row.get("first_name", "").strip()
    last_name = row.get("last_name", "").strip()
    email = row.get("email", "").strip().lower()
    phone_raw = row.get("phone", "").strip()
    phone = _schoolpay_phone(phone_raw)
    if not phone or len(re.sub(r"\D", "", phone)) != 10:
        raise ValueError(
            "phone must be a valid Ugandan mobile number (10 digits, e.g. 0701234567)."
        )
    reg_no = row.get("reg_no", "").strip()
    study_mode = row.get("study_mode", "").strip().upper()

    if not first_name or not last_name:
        raise ValueError("first_name and last_name are required.")
    if not email:
        raise ValueError("email is required.")
    if not phone:
        raise ValueError("phone is required.")
    if not reg_no:
        raise ValueError("reg_no is required.")
    if study_mode not in STUDY_MODES:
        raise ValueError(f"study_mode must be one of: {', '.join(sorted(STUDY_MODES))}.")
    if AdmittedStudent.objects.filter(reg_no=reg_no).exists():
        raise ValueError(f"reg_no '{reg_no}' is already in use.")

    try:
        dob_date = _parse_date_of_birth(row.get("date_of_birth", ""))
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    specialization = None
    spec_raw = row.get("specialization", "").strip()
    if spec_raw:
        matched, spec_err = resolve_specialization_for_program(program, spec_raw)
        if spec_err:
            raise ValueError(spec_err)
        specialization = matched

    applicant_user = User.objects.filter(email=email, is_applicant=True).first()
    if not applicant_user:
        base_username = email.split("@")[0]
        username = base_username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base_username}_{counter}"
            counter += 1
        applicant_user = User.objects.create_user(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            password="NDU@1234",
            is_applicant=True,
            allow_multi_campus_per_day=False,
            primary_campus=campus,
        )

    application = Application.objects.create(
        applicant=applicant_user,
        batch=admission_batch,
        campus=campus,
        academic_level=academic_level,
        source=Application.SOURCE_LEGACY,
        status="Admitted",
        application_reference=generate_reference(),
        first_name=first_name,
        last_name=last_name,
        middle_name=row.get("middle_name", "").strip(),
        date_of_birth=dob_date,
        gender=row.get("gender", "").strip(),
        nationality=row.get("nationality", "").strip(),
        phone=phone,
        email=email,
        address=row.get("address", "").strip(),
    )
    sync_application_program_choices(application, [program.id])

    payload = {
        "application": application.pk,
        "reg_no": reg_no,
        "admitted_program": program.pk,
        "admitted_batch": admission_batch.pk,
        "admitted_campus": campus.pk,
        "study_mode": study_mode,
        "is_admitted": True,
        "admission_date": timezone.now(),
        "admitted_by": admitted_by.pk if admitted_by else None,
        "intended_program_batch": program_batch.pk,
        "admission_notes": f"Bulk import into academic batch: {program_batch.name}.",
    }
    serializer = AdmittedStudentSerializer(data=payload)
    serializer.is_valid(raise_exception=True)
    admitted = serializer.save()

    if specialization and hasattr(admitted, "programme_enrollment"):
        try:
            spe = admitted.programme_enrollment
            spe.specialization = specialization
            spe.save(update_fields=["specialization", "updated_at"])
        except Exception:
            pass

    paycode = _ensure_schoolpay_protection(
        admitted,
        row,
        register_schoolpay=register_schoolpay,
    )

    try:
        from payments.utils.tuition_ledger_linking import relink_tuition_ledgers_for_student

        relink_tuition_ledgers_for_student(admitted)
    except Exception:
        logger.exception(
            "Failed to relink tuition ledgers after bulk import for student pk=%s",
            admitted.pk,
        )

    ext = _apply_import_extensions(
        admitted,
        program=program,
        program_batch=program_batch,
        row=row,
        admitted_by=admitted_by,
        specialization=specialization,
        require_academic_position=require_academic_position,
    )

    try:
        provision_student_portal_on_admission(admitted.id, send_credentials_email=True)
    except StudentPortalProvisioningError as exc:
        raise ValueError(str(exc)) from exc

    transaction.on_commit(
        lambda aid=admitted.id, app_id=application.id: queue_admission_notification_emails(
            aid, app_id
        )
    )
    admitted._import_paycode = paycode  # noqa: SLF001
    admitted._import_extensions = ext  # noqa: SLF001
    return admitted


def _tally_extension_counters(ext: dict, counters: dict) -> None:
    if ext.get("enrollment_set"):
        counters["enrollment_set_rows"] += 1
        if ext.get("enrollment_status") == "pending" or (
            not ext.get("enrollment_activated") and ext.get("enrollment_blocked")
        ):
            counters["enrollment_pending_rows"] += 1
            block = ext.get("enrollment_blocked")
            if block and len(counters["warnings"]) < 20:
                reg = ext.get("reg_no") or ""
                counters["warnings"].append(
                    f"{reg + ': ' if reg else ''}Year/semester set but enrollment pending — {block}"
                )
    if ext.get("fees_paid_recorded") or ext.get("fees_outstanding_recorded"):
        counters["fees_imported_rows"] += 1
    if ext.get("enrollment_activated"):
        counters["enrollment_activated_rows"] += 1


def process_student_batch_import(
    *,
    uploaded_file,
    program_batch_id: int,
    campus_id: int,
    admitted_by,
    admission_batch_id: int | None = None,
    register_schoolpay: bool = True,
    skip_existing_reg_no: bool = False,
    require_academic_position: bool = False,
) -> dict:
    try:
        program_batch = ProgramBatch.objects.select_related("program").get(pk=program_batch_id)
    except ProgramBatch.DoesNotExist:
        raise ValueError("Academic programme batch not found.")

    admission_batch = _resolve_admission_intake_batch(
        admission_batch_id,
        created_by=admitted_by,
    )

    try:
        campus = Campus.objects.get(pk=campus_id)
    except Campus.DoesNotExist:
        raise ValueError("Campus not found.")

    program = program_batch.program
    if program_batch.program_id != program.id:
        raise ValueError("Programme batch does not match programme.")

    academic_level = program.academic_level
    if academic_level is None:
        raise ValueError("Programme has no academic level configured.")

    # Continuing migration needs curriculum + semester fees before any rows are created.
    if require_academic_position:
        _preflight_continuing_import(program, program_batch)

    headers, rows = _parse_upload_file(uploaded_file)
    missing = _require_columns(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    if require_academic_position:
        for col in ("current_year_of_study", "current_term_number"):
            if col not in headers:
                raise ValueError(
                    f"Missing required continuing-student column: {col}. "
                    "Re-download the template and fill year and semester for every row."
                )
    if not rows:
        raise ValueError("No data rows found in file.")

    created = 0
    updated = 0
    skipped = 0
    failed = 0
    errors: list[str] = []
    created_students: list[dict] = []
    updated_students: list[dict] = []
    skipped_students: list[dict] = []
    counters = {
        "enrollment_set_rows": 0,
        "fees_imported_rows": 0,
        "enrollment_activated_rows": 0,
        "enrollment_pending_rows": 0,
        "warnings": [],
    }

    for row in rows:
        row_num = row.get("__row__", "?")
        reg_no = row.get("reg_no", "").strip()
        try:
            if skip_existing_reg_no and reg_no:
                existing = (
                    AdmittedStudent.objects.filter(reg_no=reg_no)
                    .select_related("application", "programme_enrollment")
                    .first()
                )
                if existing is not None:
                    batch_changed = False
                    if existing.intended_program_batch_id != program_batch.id:
                        existing.intended_program_batch = program_batch
                        existing.save(
                            update_fields=["intended_program_batch", "updated_at"]
                        )
                        AdmittedStudentSerializer._sync_programme_enrollment_batch(existing)
                        batch_changed = True

                    spec_raw = row.get("specialization", "").strip()
                    specialization = None
                    if spec_raw:
                        matched, spec_err = resolve_specialization_for_program(
                            program, spec_raw
                        )
                        if spec_err:
                            raise ValueError(spec_err)
                        specialization = matched

                    with transaction.atomic():
                        ext = _apply_import_extensions(
                            existing,
                            program=program,
                            program_batch=program_batch,
                            row=row,
                            admitted_by=admitted_by,
                            specialization=specialization,
                            require_academic_position=require_academic_position,
                        )

                    student_row = {
                        "id": existing.id,
                        "reg_no": existing.reg_no,
                        "student_id": existing.student_id,
                        "name": existing.full_name,
                        **ext,
                    }
                    if ext.get("extensions_applied"):
                        updated += 1
                        ext_for_tally = {**ext, "reg_no": existing.reg_no}
                        _tally_extension_counters(ext_for_tally, counters)
                        if batch_changed:
                            student_row["note"] = (
                                "Already in system — batch and import columns updated."
                            )
                        updated_students.append(student_row)
                    else:
                        skipped += 1
                        student_row["note"] = (
                            "Already in system — batch updated if needed."
                            if batch_changed
                            else "Already in system — no optional columns to apply."
                        )
                        skipped_students.append(student_row)
                    continue

            with transaction.atomic():
                admitted = _import_one_row(
                    row=row,
                    program=program,
                    program_batch=program_batch,
                    admission_batch=admission_batch,
                    campus=campus,
                    academic_level=academic_level,
                    admitted_by=admitted_by,
                    register_schoolpay=register_schoolpay,
                    require_academic_position=require_academic_position,
                )
            created += 1
            paycode = getattr(admitted, "_import_paycode", None) or admitted.student_id
            ext = getattr(admitted, "_import_extensions", {}) or {}
            _tally_extension_counters({**ext, "reg_no": admitted.reg_no}, counters)
            created_students.append(
                {
                    "id": admitted.id,
                    "reg_no": admitted.reg_no,
                    "student_id": paycode,
                    "schoolpay_registered": bool(admitted.is_registered_with_schoolpay),
                    "name": f"{row.get('first_name', '')} {row.get('last_name', '')}".strip(),
                    **ext,
                }
            )
        except (ValueError, DRFValidationError) as exc:
            failed += 1
            if isinstance(exc, DRFValidationError):
                detail = exc.detail
                msg = str(detail) if not isinstance(detail, dict) else "; ".join(
                    f"{k}: {v}" for k, v in detail.items()
                )
            else:
                msg = str(exc)
            errors.append(f"Row {row_num}: {msg}")
        except Exception as exc:
            failed += 1
            logger.exception("Bulk import row %s failed", row_num)
            errors.append(f"Row {row_num}: {exc}")

    return {
        "program_batch_id": program_batch.id,
        "program_batch_name": program_batch.name,
        "program_name": program.name,
        "admission_batch_id": admission_batch.id,
        "admission_batch_name": admission_batch.name,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "failed": failed,
        "enrollment_set_rows": counters["enrollment_set_rows"],
        "fees_imported_rows": counters["fees_imported_rows"],
        "enrollment_activated_rows": counters["enrollment_activated_rows"],
        "enrollment_pending_rows": counters["enrollment_pending_rows"],
        "warnings": counters["warnings"][:20],
        "require_academic_position": require_academic_position,
        "errors": errors[:100],
        "created_students": created_students[:50],
        "updated_students": updated_students[:50],
        "skipped_students": skipped_students[:50],
    }


def build_student_import_template_csv() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(STUDENT_IMPORT_TEMPLATE_HEADERS)
    writer.writerow(
        [
            "Jane",
            "Doe",
            "",
            "jane.doe@example.com",
            "0701234567",
            "2000-05-15",
            "Female",
            "Ugandan",
            "26/1/100/D/0001",
            "D",
            "1701234567",
            "",
            "Kampala",
            "3",
            "1",
            "850000",
            "LEG-2024-001",
            "450000",
            "yes",
        ]
    )
    return buf.getvalue()
