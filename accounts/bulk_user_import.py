"""Bulk ERP staff user import from Excel template (matches Add User form fields)."""
from __future__ import annotations

import io
import re
import secrets
import string
from typing import Any

from django.contrib.auth.models import Group
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from accounts.models import Campus, User
from accounts.serializers import normalize_staff_id
from accounts.tasks import celery_send_account_email

# Same essentials as Add User form (password is system-generated — not in the sheet).
# Faculty is assigned later via Edit User when needed (Dean / Admin / HOD).
HEADER_MAP = {
    "FIRST NAME *": "first_name",
    "LAST NAME *": "last_name",
    "EMAIL *": "email",
    "ROLE *": "role",
    "STAFF ID": "staff_id",
    "PHONE": "phone",
    "CAMPUS": "campus",
}

REQUIRED_FIELDS = ("first_name", "last_name", "email", "role")

TEMPLATE_HEADERS = list(HEADER_MAP.keys())

TEMPLATE_INSTRUCTIONS = (
    "* = compulsory. Use ROLE and CAMPUS dropdowns only. "
    "Password is emailed automatically — user changes it on first login."
)


def generate_temp_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    chars = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ]
    chars += [secrets.choice(alphabet) for _ in range(max(0, length - 3))]
    secrets.SystemRandom().shuffle(chars)
    return "".join(chars)


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _resolve_group(role_name: str) -> Group | None:
    name = (role_name or "").strip()
    if not name or name.lower() == "student":
        return None
    return Group.objects.filter(name__iexact=name).first()


def _resolve_campus(name: str) -> Campus | None:
    n = (name or "").strip()
    if not n:
        return None
    return (
        Campus.objects.filter(name__iexact=n).first()
        or Campus.objects.filter(code__iexact=n).first()
    )


def build_user_upload_workbook() -> Workbook:
    """
    Excel template with ROLE and CAMPUS dropdowns backed by DB lists on the Lookups sheet.
    (Sheet-range validation avoids Excel's 255-character inline-list limit.)
    """
    roles = list(
        Group.objects.exclude(name__iexact="Student")
        .order_by("name")
        .values_list("name", flat=True)
    )
    campuses = list(Campus.objects.order_by("name").values_list("name", flat=True))

    wb = Workbook()

    # --- Lookups (source for dropdowns) ---
    ws_look = wb.active
    ws_look.title = "Lookups"
    ws_look["A1"] = "ROLE"
    ws_look["B1"] = "CAMPUS"
    ws_look["A1"].font = Font(bold=True)
    ws_look["B1"].font = Font(bold=True)
    for i, role in enumerate(roles, start=2):
        ws_look.cell(row=i, column=1, value=role)
    for i, campus in enumerate(campuses, start=2):
        ws_look.cell(row=i, column=2, value=campus)
    ws_look.column_dimensions["A"].width = 36
    ws_look.column_dimensions["B"].width = 28
    role_last = max(2, len(roles) + 1)
    campus_last = max(2, len(campuses) + 1)

    # --- Users sheet ---
    ws = wb.create_sheet("Users", 0)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C1519", end_color="7C1519", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TEMPLATE_HEADERS))
    ws.cell(1, 1, TEMPLATE_INSTRUCTIONS)
    ws.cell(1, 1).font = Font(color="7C1519", italic=True, size=11)
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 28

    header_row = 2
    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(header) + 2)

    # Empty data rows — staff fill from row 3 downward (no sample data).
    for row_idx in range(3, 8):
        for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border

    data_start = 3
    data_end = 1000

    # ROLE * = column 4 — dropdown from Lookups!A
    if roles:
        dv_role = DataValidation(
            type="list",
            formula1=f"Lookups!$A$2:$A${role_last}",
            allow_blank=False,
            showDropDown=False,  # False = show dropdown arrow in Excel
            showErrorMessage=True,
            errorTitle="Invalid role",
            error="Pick a role from the dropdown (values come from the database).",
            promptTitle="Role",
            prompt="Select a role from the list.",
        )
        dv_role.add(f"D{data_start}:D{data_end}")
        ws.add_data_validation(dv_role)

    # CAMPUS = column 7 — dropdown from Lookups!B
    if campuses:
        dv_campus = DataValidation(
            type="list",
            formula1=f"Lookups!$B$2:$B${campus_last}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid campus",
            error="Pick a campus from the dropdown (values come from the database).",
            promptTitle="Campus",
            prompt="Select a campus from the list (optional).",
        )
        dv_campus.add(f"G{data_start}:G{data_end}")
        ws.add_data_validation(dv_campus)

    ws.freeze_panes = "A3"

    # Field guide
    guide = wb.create_sheet("Field guide")
    guide.append(["Column", "Required?", "Notes"])
    guide.append(["FIRST NAME *", "Compulsory", "Same as Add User form"])
    guide.append(["LAST NAME *", "Compulsory", "Same as Add User form"])
    guide.append(["EMAIL *", "Compulsory", "Login username = email (unique)"])
    guide.append(["ROLE *", "Compulsory", "Dropdown from DB roles (Groups)"])
    guide.append(["STAFF ID", "Optional", "Same as Add User form"])
    guide.append(["PHONE", "Optional", "Same as Add User form"])
    guide.append(["CAMPUS", "Optional", "Dropdown from DB campuses"])
    guide.append(
        [
            "(password)",
            "System-generated",
            "Emailed to the user; must_change_password on first login",
        ]
    )
    for col in ("A", "B", "C"):
        guide.column_dimensions[col].width = 42

    return wb


def parse_upload_file(upload) -> list[dict[str, Any]]:
    """Return list of row dicts with raw string values + _row number."""
    name = (getattr(upload, "name", "") or "").lower()
    raw = upload.read()
    rows: list[dict[str, Any]] = []

    if name.endswith(".csv"):
        import csv

        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
    else:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb["Users"] if "Users" in wb.sheetnames else wb.active
        all_rows = [
            [c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)
        ]

    if not all_rows:
        return []

    start = 0
    first = _norm_header(all_rows[0][0] if all_rows[0] else "")
    if first and "FIRST NAME" not in first and len(str(all_rows[0][0] or "")) > 40:
        start = 1

    if start >= len(all_rows):
        return []

    header_cells = [_norm_header(h) for h in all_rows[start]]
    col_index: dict[str, int] = {}
    for idx, h in enumerate(header_cells):
        for template_h, field in HEADER_MAP.items():
            if _norm_header(template_h) == h:
                col_index[field] = idx
                break
        # Accept older multi-campus header if someone still has it
        if h.startswith("CAMPUSES") and "campus" not in col_index:
            col_index["campus"] = idx

    missing = [f for f in REQUIRED_FIELDS if f not in col_index]
    if missing:
        raise ValueError(
            "Template headers missing or renamed. Download a fresh template. "
            f"Missing: {', '.join(missing)}"
        )

    for r_i, row in enumerate(all_rows[start + 1 :], start=start + 2):
        if not row or not any(str(c or "").strip() for c in row):
            continue
        item: dict[str, Any] = {"_row": r_i}
        for field, idx in col_index.items():
            item[field] = row[idx] if idx < len(row) else ""
        rows.append(item)
    return rows


def import_users_from_rows(rows: list[dict[str, Any]], *, send_email: bool = True) -> dict:
    """
    Create users quickly: one-shot lookup caches, single save per user,
    direct group/campus assignment, lightweight staff profile create.
    Emails are queued after create (Celery) so the HTTP response isn't blocked on SMTP.
    """
    from django.contrib.auth.hashers import make_password

    from accounts.role_assignment import (
        _promote_staff_portal_identity,
        primary_staff_role,
        sync_user_role_flags,
    )
    from hr.staff.models import StaffProfile

    created: list[dict] = []
    errors: list[str] = []
    emails_queued = 0
    pending_emails: list[tuple[int, str]] = []  # (user_id, password)

    # Prefetch lookups once (avoids N queries per row).
    groups_by_name = {
        g.name.lower(): g
        for g in Group.objects.exclude(name__iexact="Student")
    }
    campuses_by_key: dict[str, Campus] = {}
    for c in Campus.objects.all():
        campuses_by_key[c.name.lower()] = c
        if c.code:
            campuses_by_key[c.code.lower()] = c

    existing_emails = {
        e.lower()
        for e in User.objects.exclude(email__isnull=True)
        .exclude(email="")
        .values_list("email", flat=True)
    }
    existing_staff_ids = {
        s
        for s in User.objects.exclude(staff_id__isnull=True)
        .exclude(staff_id="")
        .values_list("staff_id", flat=True)
    }
    # Also reserve emails within this upload file
    seen_in_file: set[str] = set()
    seen_staff_in_file: set[str] = set()

    for item in rows:
        row_no = item.get("_row", "?")
        first_name = str(item.get("first_name") or "").strip()
        last_name = str(item.get("last_name") or "").strip()
        email = str(item.get("email") or "").strip().lower()
        role_raw = str(item.get("role") or "").strip()
        staff_id = normalize_staff_id(item.get("staff_id"))
        phone = str(item.get("phone") or "").strip().replace(" ", "")[:20]
        campus_raw = str(item.get("campus") or "").strip()
        if "|" in campus_raw or ";" in campus_raw or "," in campus_raw:
            campus_raw = re.split(r"[|;,]+", campus_raw)[0].strip()

        row_errors: list[str] = []
        if not first_name:
            row_errors.append("first name required")
        if not last_name:
            row_errors.append("last name required")
        if not email:
            row_errors.append("email required")
        elif email in existing_emails or email in seen_in_file:
            row_errors.append(f"email '{email}' already exists")
        if not role_raw:
            row_errors.append("role required — use the ROLE dropdown")
        elif role_raw.lower() == "student":
            row_errors.append("Student role cannot be created here")

        group = groups_by_name.get(role_raw.lower()) if role_raw else None
        if role_raw and group is None and role_raw.lower() != "student":
            row_errors.append(
                f"unknown role '{role_raw}' — pick from the ROLE dropdown (database roles)"
            )

        campus = campuses_by_key.get(campus_raw.lower()) if campus_raw else None
        if campus_raw and campus is None:
            row_errors.append(
                f"unknown campus '{campus_raw}' — pick from the CAMPUS dropdown"
            )

        if staff_id and (staff_id in existing_staff_ids or staff_id in seen_staff_in_file):
            row_errors.append(f"staff ID '{staff_id}' already in use")

        if row_errors:
            errors.append(f"Row {row_no}: {'; '.join(row_errors)}")
            continue

        password = generate_temp_password()
        try:
            with transaction.atomic():
                user = User(
                    email=email,
                    username=email,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone or None,
                    staff_id=staff_id or None,
                    role=group.name if group else role_raw,
                    is_staff=True,
                    is_active=True,
                    must_change_password=True,
                    password=make_password(password),
                )
                user.save()

                if campus:
                    user.campuses.set([campus])
                if group:
                    user.groups.set([group])
                    sync_user_role_flags(user, save=False)
                    primary = primary_staff_role(user)
                    if primary:
                        user.role = primary
                    update_fields = ["role", "is_staff", "is_lecturer"]
                    update_fields.extend(_promote_staff_portal_identity(user))
                    user.save(update_fields=list(dict.fromkeys(update_fields)))

                # Fast path: new ERP user → create StaffProfile without extra lookups.
                staff = StaffProfile.objects.create(
                    user=user,
                    first_name=first_name,
                    last_name=last_name,
                    university_email=email,
                    staff_no=staff_id or "",
                    system_login=True,
                )
                if campus:
                    staff.campus.set([campus])
        except Exception as exc:
            errors.append(f"Row {row_no}: could not create user — {exc}")
            continue

        seen_in_file.add(email)
        existing_emails.add(email)
        if staff_id:
            seen_staff_in_file.add(staff_id)
            existing_staff_ids.add(staff_id)

        pending_emails.append((user.pk, password))
        created.append(
            {
                "id": user.pk,
                "email": email,
                "name": f"{first_name} {last_name}".strip(),
                "role": group.name if group else role_raw,
                "email_queued": False,
            }
        )

    if send_email and pending_emails:
        for i, (user_id, password) in enumerate(pending_emails):
            try:
                celery_send_account_email.delay(user_id, password, True)
                emails_queued += 1
                created[i]["email_queued"] = True
            except Exception:
                errors.append(
                    f"User id {user_id}: created but email queue failed"
                )

    return {
        "created": len(created),
        "emails_queued": emails_queued,
        "created_users": created,
        "errors": errors,
    }
