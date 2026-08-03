"""Import halls/rooms/beds from Halls of Residence Excel (idempotent on room code)."""
from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import Campus

from hostel.models import Bed, Building, Floor, Hostel, Room

# Stable building codes aligned with offer-letter hall names where possible
BUILDING_CODE_MAP = {
    "bishop yokana": "YOKANA",
    "bishop yokana mukasa": "YOKANA",
    "noah's ark": "NOAHS_ARK",
    "noahs ark": "NOAHS_ARK",
    "akiibua": "AKIIBUA",
    "kakungulu": "KAKUNGULU",
    "kakungulu annex": "KAKUNGULU_ANNEX",
    "muteesa gents": "MUTEESA",
    "muteesa": "MUTEESA",
    "westbuganda": "WEST_BUGANDA",
    "west buganda": "WEST_BUGANDA",
    "west buganda annex": "WEST_BUGANDA_ANNEX",
    "wekisa": "WEKISA",
    "victoria mwaka": "VICTORIA_MWAKA",
    "luweero": "LUWEERO",
    "luwero": "LUWEERO",
}

FLOOR_SORT = {
    "basement": 0,
    "ground floor": 1,
    "groundfloor": 1,
    "level 1": 2,
    "level1": 2,
    "level 2": 3,
    "level2": 3,
    "level 3": 4,
    "level3": 4,
    "level 4": 5,
    "level4": 5,
}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _building_code(name: str, block_id: str) -> str:
    key = _norm(name).lower()
    if key in BUILDING_CODE_MAP:
        return BUILDING_CODE_MAP[key]
    # fallback from block id
    bid = _norm(block_id).upper().replace(" ", "_")
    if bid:
        return bid
    slug = re.sub(r"[^A-Z0-9]+", "_", key.upper()).strip("_")
    return slug or "UNKNOWN"


def _hostel_gender(hostel_name: str) -> str:
    n = _norm(hostel_name).lower()
    if "female" in n or "ladies" in n or "girls" in n:
        return Hostel.GENDER_FEMALE
    return Hostel.GENDER_MALE


def _hostel_code(gender: str) -> str:
    return "FEMALE" if gender == Hostel.GENDER_FEMALE else "GENTS"


def _hostel_display(gender: str) -> str:
    return "Female Hostel" if gender == Hostel.GENDER_FEMALE else "Gents Hostel"


def _resolve_campus(raw: str) -> Campus:
    name = _norm(raw).replace("Unversity", "University")
    if not name:
        name = "Main Campus"
    # Prefer MAIN code, then Main Campus, then Ndejje*
    campus = Campus.objects.filter(code__iexact="MAIN").first()
    if campus:
        return campus
    campus = Campus.objects.filter(name__iexact="Main Campus").first()
    if campus:
        return campus
    campus = Campus.objects.filter(name__icontains="Ndejje").first()
    if campus:
        return campus
    campus, _ = Campus.objects.get_or_create(
        name="Main Campus",
        defaults={"code": "MAIN", "address": "", "email": ""},
    )
    return campus


def _room_kind(cap, note) -> tuple[str, int]:
    note_l = _norm(note).lower()
    cap_s = _norm(cap)
    if note_l == "store" or "store" in note_l:
        return Room.KIND_STORE, 0
    if "common" in cap_s.lower() or "common" in note_l:
        return Room.KIND_COMMON, 0
    try:
        n = int(float(cap_s)) if cap_s else 0
    except ValueError:
        return Room.KIND_OTHER, 0
    if n <= 0:
        return Room.KIND_OTHER, 0
    return Room.KIND_BEDROOM, n


def _floor_sort(name: str) -> int:
    return FLOOR_SORT.get(_norm(name).lower(), 50)


class Command(BaseCommand):
    help = "Import hostel rooms/beds from Halls of Residence Excel (idempotent on room code)."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="",
            help="Path to Halls of Residence details-2026.xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report without writing.",
        )

    def handle(self, *args, **options):
        path = options["path"]
        if not path:
            # Default: workspace sibling of backend
            candidates = [
                Path(__file__).resolve().parents[3] / "Halls of Residence details-2026.xlsx",
                Path(__file__).resolve().parents[2].parent
                / "Halls of Residence details-2026.xlsx",
                Path.cwd() / "Halls of Residence details-2026.xlsx",
                Path.cwd().parent / "Halls of Residence details-2026.xlsx",
            ]
            for c in candidates:
                if c.is_file():
                    path = str(c)
                    break
        if not path or not Path(path).is_file():
            raise CommandError(
                "Excel file not found. Pass path to Halls of Residence details-2026.xlsx"
            )

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required") from exc

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Room Allocation"] if "Room Allocation" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]

        stats = {
            "rooms_seen": 0,
            "rooms_created": 0,
            "rooms_updated": 0,
            "beds_created": 0,
            "skipped": 0,
        }

        @transaction.atomic
        def run():
            for r in rows:
                if not r or not r[6]:
                    stats["skipped"] += 1
                    continue
                hostel_name = _norm(r[1]) or "Gents Hostel"
                block_id = _norm(r[2])
                building_name = _norm(r[3])
                floor_id = _norm(r[4])
                floor_name = _norm(r[5]) or floor_id or "LEVEL1"
                room_id = _norm(r[6])
                cap = r[7]
                room_name = _norm(r[8])
                campus_raw = _norm(r[9])
                note = _norm(r[10]) if len(r) > 10 else ""

                if not room_id:
                    stats["skipped"] += 1
                    continue

                stats["rooms_seen"] += 1
                if options["dry_run"]:
                    continue

                campus = _resolve_campus(campus_raw)
                gender = _hostel_gender(hostel_name)
                h_code = _hostel_code(gender)
                hostel, _ = Hostel.objects.update_or_create(
                    campus=campus,
                    code=h_code,
                    defaults={
                        "name": _hostel_display(gender),
                        "gender": gender,
                        "is_active": True,
                    },
                )
                b_code = _building_code(building_name, block_id)
                # Prefer longest canonical name for display
                building, created_b = Building.objects.get_or_create(
                    hostel=hostel,
                    code=b_code,
                    defaults={
                        "name": building_name or b_code,
                        "external_block_id": block_id,
                        "is_active": True,
                    },
                )
                if not created_b:
                    updates = {}
                    if building_name and len(building_name) > len(building.name or ""):
                        updates["name"] = building_name
                    if block_id and not building.external_block_id:
                        updates["external_block_id"] = block_id
                    if updates:
                        for k, v in updates.items():
                            setattr(building, k, v)
                        building.save(update_fields=list(updates.keys()) + ["updated_at"])

                f_code = floor_id or re.sub(r"[^A-Za-z0-9]+", "", floor_name) or "F1"
                floor, _ = Floor.objects.update_or_create(
                    building=building,
                    code=f_code,
                    defaults={
                        "name": floor_name,
                        "sort_order": _floor_sort(floor_name),
                    },
                )

                kind, capacity = _room_kind(cap, note)
                room, created = Room.objects.update_or_create(
                    code=room_id,
                    defaults={
                        "floor": floor,
                        "display_name": room_name or room_id,
                        "room_kind": kind,
                        "capacity": capacity,
                        "notes": note,
                        "is_active": True,
                    },
                )
                if created:
                    stats["rooms_created"] += 1
                else:
                    stats["rooms_updated"] += 1

                if kind == Room.KIND_BEDROOM and capacity > 0:
                    existing = {b.label: b for b in room.beds.all()}
                    for i in range(1, capacity + 1):
                        label = f"Bed {i}"
                        if label not in existing:
                            Bed.objects.create(
                                room=room,
                                label=label,
                                status=Bed.STATUS_AVAILABLE,
                            )
                            stats["beds_created"] += 1

        run()
        self.stdout.write(self.style.SUCCESS(f"Imported from {path}"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
