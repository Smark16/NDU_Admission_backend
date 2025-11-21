# utils/excel.py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_workbook(headers, rows, sheet_name="Report", header_bg="4F81BD"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # === Styles ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_bg, fill_type="solid")
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # === Write Header Row ===
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        # Auto-fit column width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(header)) + 2)

    # === Write Data Rows ===
    for row in rows:
        ws.append(row)

    # Apply alignment and borders to all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    ws.freeze_panes = "A2"   # Freeze header row

    return wb
