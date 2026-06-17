from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import *

from admissions.permissions import ExportVerificationRegisterPermission
from admissions.models import * 
from django.db.models.functions import Coalesce
from django.db.models import Count, F, Q, Prefetch
from collections import defaultdict
from django.db import connection
from datetime import datetime

from .utils.excel import create_workbook
from .utils.calculate_passes import calculate_pp_sp
from django.http import HttpResponse
from admissions.models import *
from admissions.utils.academic_year import get_current_academic_year
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from admissions.utils.program_choices import (
    PROGRAM_CHOICE_CONFIRMED_BY_APPLICANT,
)

def _application_full_name_upper(app):
    parts = [app.first_name or "", app.middle_name or "", app.last_name or ""]
    return " ".join(p.strip() for p in parts if p and p.strip()).upper()

def _gender_short(app):
    g = (app.gender or "").strip().upper()
    if g.startswith("F"):
        return "F"
    if g.startswith("M"):
        return "M"
    return (app.gender or "").strip()

def _admission_mode_label(app):
    if app.academic_level_id:
        return (app.academic_level.name or "").strip()
    return ""

def _direct_admission_reason(adm):
    app = getattr(adm, "application", None)
    if not app or (app.source or "") != Application.SOURCE_DIRECT:
        return ""
    notes = (adm.admission_notes or "").strip()
    if not notes:
        return ""
    prefix = "Direct admission reason:"
    for line in notes.splitlines():
        line_clean = line.strip()
        if line_clean.lower().startswith(prefix.lower()):
            return line_clean[len(prefix):].strip()
    return notes

# Create your views here.

# general overview
class GeneralOverview(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        batches = Batch.objects.all().order_by("academic_year", "name")

        results = []

        for batch in batches:
            apps = Application.objects.select_related(
                'batch', 'campus', 'academic_level', 'reviewed_by'
                ).filter(batch=batch).aggregate(
                   total_applications=Count('id'),
                   pending = Count('id', filter=Q(status="submitted")),
                   rejected = Count('id', filter=Q(status="rejected"))
                )
            admitted_qs = AdmittedStudent.objects.select_related(
                'admitted_campus', 'admitted_program', 'admitted_batch', 'admitted_by', 'application__applicant'
                ).filter(admitted_batch=batch).aggregate(
                    total_admitted=Count('id'),
                    accepted=Count('id', filter=Q(is_admitted=True)),
                )

            stats = {
                "admission_period": batch.name,  
                "academic_year": batch.academic_year, 
                "total_applications": apps['total_applications'],
                "accepted": admitted_qs['accepted'],
                "pending": apps['pending'],
                "rejected": apps['rejected'],
                "total_admitted": admitted_qs['total_admitted'],
            }

            results.append(stats)

        return Response(results, status=200)

# Admitted students grouped by Academic Year and Admission Period
class Admitted_students_by_Faculty(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        batches = Batch.objects.all().order_by("academic_year", "name")

        final_output = []

        for batch in batches:
            faculty_stats = (
                AdmittedStudent.objects
                .filter(admitted_batch=batch, is_admitted=True)
                .values(
                    academic_year=F("admitted_batch__academic_year"),
                    admission_period=F("admitted_batch__name"),
                    faculty=F("admitted_program__faculty__name")
                )
                .annotate(admitted=Count("id"))
                .order_by("faculty")
            )

            # convert queryset → list
            faculty_list = list(faculty_stats)

            # Skip if no admissions
            if not faculty_list:
                continue

            # Prepare final batch-grouped block
            final_output.append({
                "academic_year": faculty_list[0]["academic_year"],
                "admission_period": faculty_list[0]["admission_period"],
                "faculty_data": [
                    {
                        "faculty": item["faculty"],
                        "admitted": item["admitted"]
                    }
                    for item in faculty_list
                ]
            })

        return Response(final_output, status=200)

    
# Faculty Admitted Students reports

class ViewFacultyAdmissions(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        admitted_students = list(
            AdmittedStudent.objects
            .select_related("application", "admitted_program", "admitted_campus", "admitted_batch")
            .filter(is_admitted=True)
        )

        if not admitted_students:
            return Response([])

        app_ids = [adm.application_id for adm in admitted_students]

        # 1. Programs
        # program_data = defaultdict(list)
        # for prog in Program.objects.filter(application_programs__id__in=app_ids)\
        #         .values("application_programs__id", "name"):
        #     program_data[prog["application_programs__id"]].append(prog["name"])
         # ==================== 1. PROGRAM CHOICES (FIXED) ====================
        program_data = defaultdict(list)
        for choice in ApplicationProgramChoice.objects.filter(
            application_id__in=app_ids
        ).select_related("program").order_by("choice_order"):
            program_data[choice.application_id].append(choice.program.name)

        # 2. O-Level
        olevel_data = defaultdict(list)
        for res in OLevelResult.objects.filter(application_id__in=app_ids)\
                .select_related("subject")\
                .values("application_id", "subject__code", "grade"):
            olevel_data[res["application_id"]].append(f"{res['subject__code']}:{res['grade']}")

        # 3. A-Level
        alevel_for_pp_sp = defaultdict(list)
        alevel_scores = defaultdict(list)
        for res in ALevelResult.objects.filter(application_id__in=app_ids)\
                .select_related("subject")\
                .values("application_id", "subject__code", "grade"):
            app_id = res["application_id"]
            alevel_for_pp_sp[app_id].append({"subject_name": res["subject__code"], "grade": res["grade"]})
            alevel_scores[app_id].append(f"{res['subject__code']}:{res['grade']}")

        # 4. Additional Qualifications - String Only
        qualifications_data = defaultdict(list)
        for qual in AdditionalQualifications.objects.filter(application_id__in=app_ids)\
                .values(
                    "application_id",
                    "additional_qualification_institution",
                    "additional_qualification_type",
                    "additional_qualification_year",
                    "class_of_award"
                ):
            qualifications_data[qual["application_id"]].append(qual)

        # 5. Build Response
        grouped = defaultdict(lambda: {"academic_year": "", "admission_period": "", "students": []})

        for adm in admitted_students:
            app = adm.application
            batch = adm.admitted_batch
            key = f"{batch.academic_year}-{batch.name}"

            if key not in grouped:
                grouped[key] = {
                    "academic_year": batch.academic_year,
                    "admission_period": batch.name,
                    "students": [],
                }

            programs = program_data.get(app.id, [])
            course_applied_for = programs[0] if programs else ""
            other_choices = ", ".join(programs[1:]) if len(programs) > 1 else ""

            olevel_scores = "; ".join(olevel_data[app.id])
            alevel_scores_str = "; ".join(alevel_scores[app.id])

            pp, sp = calculate_pp_sp(alevel_for_pp_sp[app.id])
            principal_sub = f"{pp}PP, {sp}SP"

            # === Additional Qualifications as Strings (Safe & Dynamic) ===
            quals = qualifications_data.get(app.id, [])

            # Create readable combined string
            other_qual_parts = []
            institutions = []
            class_of_awards = []

            for q in quals:
                if q.get("additional_qualification_institution"):
                    qual_str = f"{q['additional_qualification_institution']} - {q.get('additional_qualification_type','')} ({q.get('additional_qualification_year','')}) - {q.get('class_of_award','')}"
                    other_qual_parts.append(qual_str)
                    institutions.append(q['additional_qualification_institution'])
                    class_of_awards.append(q.get('class_of_award', ''))

            other_qual_str = " | ".join(other_qual_parts) if other_qual_parts else "None"

            # Join multiple institutions and awards with commas
            institution_str = ", ".join(institutions) if institutions else ""
            class_of_award_str = ", ".join([c for c in class_of_awards if c]) if class_of_awards else ""

            # Final Student Dictionary - All Strings
            grouped[key]["students"].append({
                "id": adm.id,
                "student_names": f"{app.first_name} {app.last_name}",
                "gender": app.gender,
                "nationality": app.nationality,
                "contact_address": app.address or "",
                "course_applied_for": course_applied_for,
                "other_choices": other_choices,
                "program": adm.admitted_program.name if adm.admitted_program else "",
                "study_mode": getattr(adm, 'study_mode', ""),
                "campus": adm.admitted_campus.name if adm.admitted_campus else "",

                "olevel_school": app.olevel_school or "",
                "olevel_year": app.olevel_year or "",
                "olevel_index_number": app.olevel_index_number or "",
                "olevel_scores": olevel_scores,

                "alevel_school": app.alevel_school or "",
                "alevel_year": app.alevel_year or "",
                "alevel_index_number": app.alevel_index_number or "",
                "alevel_combination": app.alevel_combination or "",
                "alevel_scores": alevel_scores_str,
                "principal_subsidiaries": principal_sub,

                # All strings as requested
                "other_qualifications": other_qual_str,
                "institution": institution_str,           # All institutions combined
                "class_of_award": class_of_award_str,     # All class of awards combined

                "course_admitted_for": adm.admitted_program.name if adm.admitted_program else "",
                "remarks": adm.admission_notes or "",
                "payments": "PAID" if app.application_fee_paid else "NOT PAID",
                "admission_date": adm.admission_date.strftime("%Y-%m-%d") if adm.admission_date else "",
                "origin": "APPLIED ONLINE",
            })

        return Response(list(grouped.values()), status=200)
    
# excel reports
class ExportFacultyAdmissionsExcel(APIView):
    permission_classes = [IsAuthenticated, ExportVerificationRegisterPermission]

    def get(self, request):
        # --------------------------------------------------------------
        # 1. FILTERS
        # --------------------------------------------------------------
        academic_year = request.query_params.get("academic_year") or get_current_academic_year()
        admission_period = request.query_params.get("admission_period")
        campus_id = request.query_params.get("campus")
        program_id = request.query_params.get("program")
        faculty_id = request.query_params.get("faculty")
        documents_verified = (request.query_params.get("documents_verified") or "").lower()
        is_registered = (request.query_params.get("is_registered") or "").lower()

        # --------------------------------------------------------------
        # 2. MAIN QUERY
        # --------------------------------------------------------------
        qs = AdmittedStudent.objects.select_related(
            "application",
            "admitted_program",
            "admitted_program__faculty",
            "admitted_campus",
            "admitted_batch",
            # "physical_documents_verified_by",
        ).filter(is_admitted=True)

        if academic_year:
            qs = qs.filter(admitted_batch__academic_year=academic_year)
        if admission_period:
            qs = qs.filter(admitted_batch__name__icontains=admission_period)
        if campus_id:
            qs = qs.filter(admitted_campus_id=campus_id)
        if program_id:
            qs = qs.filter(admitted_program_id=program_id)
        if faculty_id:
            qs = qs.filter(admitted_program__faculty_id=faculty_id)
        if documents_verified in ("1", "true", "yes"):
            qs = qs.filter(physical_documents_verified=True)
        elif documents_verified in ("0", "false", "no"):
            qs = qs.filter(physical_documents_verified=False)
        if is_registered in ("1", "true", "yes"):
            qs = qs.filter(is_registered=True)
        elif is_registered in ("0", "false", "no"):
            qs = qs.filter(is_registered=False)

        admitted_students = list(qs)
        app_ids = [adm.application_id for adm in admitted_students]

        if not app_ids:
            # Return empty workbook
            wb = create_workbook([], [], sheet_name="Faculty Admissions")
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="faculty_admissions.xlsx"'
            wb.save(response)
            return response

        # --------------------------------------------------------------
        # 3. PROGRAMS
        # --------------------------------------------------------------
        # program_data = defaultdict(list)
        # for prog in Program.objects.filter(application_programs__id__in=app_ids)\
        #         .values("application_programs__id", "name"):
        #     program_data[prog["application_programs__id"]].append(prog["name"])
         # ==================== 1. PROGRAM CHOICES (FIXED) ====================
        program_data = defaultdict(list)
        for choice in ApplicationProgramChoice.objects.filter(
            application_id__in=app_ids
        ).select_related("program").order_by("choice_order"):
            program_data[choice.application_id].append(choice.program.name)

        # --------------------------------------------------------------
        # 4. O-LEVEL
        # --------------------------------------------------------------
        olevel_data = defaultdict(list)
        for res in OLevelResult.objects.filter(application_id__in=app_ids)\
                .select_related("subject")\
                .values("application_id", "subject__code", "grade"):
            olevel_data[res["application_id"]].append(f"{res['subject__code']}:{res['grade']}")

        # --------------------------------------------------------------
        # 5. A-LEVEL
        # --------------------------------------------------------------
        alevel_scores = defaultdict(list)
        alevel_for_pp = defaultdict(list)
        for res in ALevelResult.objects.filter(application_id__in=app_ids)\
                .select_related("subject")\
                .values("application_id", "subject__name", "grade"):
            app_id = res["application_id"]
            alevel_scores[app_id].append(f"{res['subject__name']}:{res['grade']}")
            alevel_for_pp[app_id].append({
                "subject_name": res["subject__name"],
                "grade": res["grade"]
            })

        # --------------------------------------------------------------
        # 6. ADDITIONAL QUALIFICATIONS (STRINGS ONLY)
        # --------------------------------------------------------------
        qualifications_data = defaultdict(list)
        for qual in AdditionalQualifications.objects.filter(application_id__in=app_ids)\
                .values(
                    "application_id",
                    "additional_qualification_institution",
                    "additional_qualification_type",
                    "additional_qualification_year",
                    "class_of_award"
                ):
            qualifications_data[qual["application_id"]].append(qual)

        # --------------------------------------------------------------
        # 7. BUILD ROWS
        # --------------------------------------------------------------
        headers = [
            "ID",
            "STUDENT NO",
            "REG NO",
            "ACADEMIC YEAR",
            "INTAKE",
            "FACULTY",
            "STUDENT NAMES",
            "GENDER",
            "NATIONALITY",
            "ADMISSION SOURCE",
            "DIRECT ADMISSION REASON",
            "REGISTERED (Y/N)",
            # "PHYS DOCS VERIFIED",
            # "VERIFIED AT",
            # "VERIFIED BY",
            # "PHYS VERIFY NOTES",
            "CONTACT/ADDRESS",
            "COURSE APPLIED FOR",
            "OTHER COURSE CHOICES",
            "PROGRAM",
            "STUDY MODE",
            "CAMPUS",
            "O-LEVEL SCHOOL",
            "O-LEVEL YEAR",
            "O-LEVEL INDEX NO",
            "O-LEVEL SCORES",
            "A-LEVEL SCHOOL",
            "A-LEVEL YEAR",
            "A-LEVEL INDEX NO",
            "A-LEVEL COMBINATION",
            "A-LEVEL SCORES",
            "PRINCIPALS/SUBSIDIARIES",
            "OTHER QUALIFICATIONS",
            "INSTITUTION",
            "CLASS OF AWARD",
            "COURSE ADMITTED FOR",
            "REMARKS",
            "PAYMENTS",
            "DATE OF ENTRY",
            "ORIGIN",
            "ADMITTED BY"
        ]

        rows = []

        for adm in admitted_students:
            app = adm.application
            batch = adm.admitted_batch

            # Programs
            programs = program_data.get(app.id, [])
            course_applied_for = programs[0] if programs else ""
            other_choices = ", ".join(programs[1:]) if len(programs) > 1 else ""

            # O-Level & A-Level
            olevel_scores = "; ".join(olevel_data[app.id])
            alevel_scores_str = "; ".join(alevel_scores[app.id])

            # PP/SP
            pp, sp = calculate_pp_sp(alevel_for_pp[app.id])
            principal_sub = f"{pp}PP, {sp}SP"

            # === ADDITIONAL QUALIFICATIONS - STRINGS ONLY ===
            quals = qualifications_data.get(app.id, [])

            other_qual_parts = []
            institutions = []
            class_of_awards = []

            for q in quals:
                if q.get("additional_qualification_institution"):
                    qual_str = (
                        f"{q['additional_qualification_institution']} - "
                        f"{q.get('additional_qualification_type', '')} "
                        f"({q.get('additional_qualification_year', '')}) - "
                        f"{q.get('class_of_award', '')}"
                    )
                    other_qual_parts.append(qual_str.strip())
                    institutions.append(q['additional_qualification_institution'])
                    if q.get('class_of_award'):
                        class_of_awards.append(q['class_of_award'])

            other_qual_str = " | ".join(other_qual_parts) if other_qual_parts else "None"
            institution_str = ", ".join(institutions) if institutions else ""
            class_of_award_str = ", ".join(class_of_awards) if class_of_awards else ""

            faculty_name = ""
            if adm.admitted_program and adm.admitted_program.faculty:
                faculty_name = adm.admitted_program.faculty.name or ""
            verified_by = ""
            # if adm.physical_documents_verified_by_id:
            #     vb = adm.physical_documents_verified_by
            #     verified_by = (vb.get_full_name() or vb.username or "") if vb else ""
            # verified_at = ""
            # if adm.physical_documents_verified_at:
            #     verified_at = adm.physical_documents_verified_at.strftime("%Y-%m-%d %H:%M")

            # Build row (all strings)
            rows.append([
                adm.id,
                adm.student_id or "",
                adm.reg_no or "",
                batch.academic_year if batch else "",
                batch.name if batch else "",
                faculty_name,
                f"{app.first_name} {app.last_name}",
                app.gender or "",
                app.nationality or "",
                app.get_source_display() if hasattr(app, "get_source_display") else (app.source or ""),
                _direct_admission_reason(adm),
                "Y" if adm.is_registered else "N",
                # "Y" if adm.physical_documents_verified else "N",
                # verified_at,
                # verified_by,
                # (adm.physical_documents_notes or "").replace("\r\n", " ").replace("\n", " ")[:2000],
                app.address or "",
                course_applied_for,
                other_choices,
                adm.admitted_program.name if adm.admitted_program else "",
                getattr(adm, 'study_mode', ""),
                adm.admitted_campus.name if adm.admitted_campus else "",
                app.olevel_school or "",
                app.olevel_year or "",
                app.olevel_index_number or "",
                olevel_scores,
                app.alevel_school or "",
                app.alevel_year or "",
                app.alevel_index_number or "",
                app.alevel_combination or "",
                alevel_scores_str,
                principal_sub,
                other_qual_str,
                institution_str,
                class_of_award_str,
                adm.admitted_program.name if adm.admitted_program else "",
                adm.admission_notes or "",
                "PAID" if app.application_fee_paid else "NOT PAID",
                adm.admission_date.strftime("%Y-%m-%d") if adm.admission_date else "",
                "APPLIED ONLINE",
                adm.admitted_by.get_full_name() if adm.admitted_by else ""
            ])

        # --------------------------------------------------------------
        # 8. CREATE EXCEL
        # --------------------------------------------------------------
        wb = create_workbook(headers, rows, sheet_name="Faculty Admissions")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"faculty_admissions_{academic_year}_{admission_period or 'all'}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response

# Admitted students reports
class ExportAdmittedExcel(APIView):
    permission_classes = [IsAuthenticated, ExportVerificationRegisterPermission]

    def get(self, request):
        # --------------------------------------------------------------
        # 1. FILTERS
        # --------------------------------------------------------------
        academic_year = request.query_params.get("academic_year") or get_current_academic_year()
        batch_id = request.query_params.get("batch")
        admission_period = request.query_params.get("admission_period")
        campus = request.query_params.get("campus")          # Can be name OR id
        program = request.query_params.get("program")
        faculty = request.query_params.get("faculty")
        is_registered = (request.query_params.get("is_registered") or "").lower()

        # --------------------------------------------------------------
        # 2. MAIN QUERY
        # --------------------------------------------------------------
        qs = AdmittedStudent.objects.select_related(
            "application",
            "admitted_program",
            "admitted_program__faculty",
            "admitted_campus",
            "admitted_batch",
        ).filter(is_admitted=True)

        if academic_year:
            qs = qs.filter(admitted_batch__academic_year=academic_year)

        if admission_period:
            qs = qs.filter(admitted_batch__name__icontains=admission_period)

        # FIXED: Handle both campus name and campus ID
        if campus:
            if campus.isdigit():                                 
                qs = qs.filter(admitted_campus_id=int(campus))
            else:                                               
                qs = qs.filter(admitted_campus__name__iexact=campus.strip())

        if program:
            qs = qs.filter(admitted_program__name__icontains=program)

        if faculty:
            qs = qs.filter(admitted_program__faculty__name__icontains=faculty)

        if is_registered in ("1", "true", "yes"):
            qs = qs.filter(is_registered=True)
        elif is_registered in ("0", "false", "no"):
            qs = qs.filter(is_registered=False)

        if batch_id and batch_id.isdigit():
            qs = qs.filter(admitted_batch_id=int(batch_id))

        admitted_students = list(qs)

        # --------------------------------------------------------------
        # 3. BUILD ROWS
        # --------------------------------------------------------------
        headers = [
            "ID",
            "STUDENT NAMES",
            "STUDENT NO",
            "REG NO",
            "PHONE",
            "EMAIL",
            "ACADEMIC YEAR",
            "INTAKE",
            "FACULTY",
            "GENDER",
            "NATIONALITY",
            "REGISTERED (Y/N)",
            "ADMITTED PROGRAM",
            "STUDY MODE",
            "CAMPUS",
            "ORIGIN",
            "ADMITTED BY",
            "ADMISSION DATE"
        ]

        rows = []
        for adm in admitted_students:
            app = adm.application
            batch = adm.admitted_batch

            faculty_name = ""
            if adm.admitted_program and adm.admitted_program.faculty:
                faculty_name = adm.admitted_program.faculty.name or ""

            rows.append([
                adm.id,
                f"{app.first_name} {app.last_name}",
                adm.student_id or "",
                adm.reg_no or "",
                app.phone or "",
                app.email or "",
                batch.academic_year if batch else "",
                batch.name if batch else "",
                faculty_name,
                app.gender or "",
                app.nationality or "",
                "Y" if getattr(adm, 'is_registered', False) else "N",
                adm.admitted_program.name if adm.admitted_program else "",
                getattr(adm, 'study_mode', ""),
                adm.admitted_campus.name if adm.admitted_campus else "",
                "APPLIED ONLINE",
                adm.admitted_by.get_full_name() if getattr(adm, 'admitted_by', None) else "",
                adm.admission_date.strftime("%Y-%m-%d") if adm.admission_date else "",
            ])

        # --------------------------------------------------------------
        # 4. CREATE EXCEL
        # --------------------------------------------------------------
        wb = create_workbook(headers, rows, sheet_name="Admitted Students")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"admitted_students_{academic_year}_{admission_period or 'all'}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response

# Applicant students reports
class ExportApplicantsExcel(APIView):
    permission_classes = [IsAuthenticated, ExportVerificationRegisterPermission]

    def get(self, request):
        # --------------------------------------------------------------
        # 1. FILTERS (Exactly as you specified)
        # --------------------------------------------------------------
        status = request.query_params.get("status")
        gender = request.query_params.get("gender")
        academic_level = request.query_params.get("academic_level")
        batch = request.query_params.get("batch")
        campus = request.query_params.get("campus")
        program = request.query_params.get("program")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("to_date")
        choice_confirmation = request.query_params.get("choice_confirmation")
        search = (request.query_params.get("search") or "").strip()

        # --------------------------------------------------------------
        # 2. MAIN QUERY
        # --------------------------------------------------------------
        qs = (
            Application.objects.select_related(
                "academic_level",
                "batch",
                "campus",
                "applicant",
                "entered_by",
            )
            .prefetch_related(
                Prefetch(
                    "program_choices",
                    queryset=ApplicationProgramChoice.objects.select_related(
                        "program__faculty"
                    ).order_by("choice_order"),
                    to_attr="prefetched_program_choices",
                )
            )
            .filter(~Q(status__in=["draft", "Admitted", "admitted", "rejected"]))
            .order_by("-created_at")
        )

        # Apply Filters
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(application_reference__icontains=search) |
                Q(program_choices__program__name__icontains=search)
            )

        if status and status != "all":
            qs = qs.filter(status=status)

        if gender and gender != "all":
            qs = qs.filter(gender=gender)

        if academic_level and academic_level != "all":
            qs = qs.filter(academic_level__name=academic_level)

        if batch and batch != "all":
            qs = qs.filter(batch__name=batch)

        if campus and campus != "all":
            qs = qs.filter(campus__name=campus)

        if program and program != "all":
            qs = qs.filter(program_choices__program__name__icontains=program)

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # if choice_confirmation and choice_confirmation != "all":
        #     if choice_confirmation.lower() == "confirmed":
        #         qs = qs.filter(program_choices_confirmed_at__isnull=False)
        #     elif choice_confirmation.lower() == "awaiting":
        #         qs = qs.filter(program_choices_confirmed_at__isnull=True)

        if choice_confirmation and choice_confirmation != "all":
            cc = choice_confirmation.strip().lower()
            if cc == "awaiting":
                qs = qs.filter(
                    status__in=["submitted", "under_review"],
                    program_choices_confirmed_at__isnull=True,
                )
            elif cc == "confirmed":
                qs = qs.filter(
                    program_choices_confirmed_at__isnull=False,
                    program_choices_confirmed_by=PROGRAM_CHOICE_CONFIRMED_BY_APPLICANT,
                )
            elif cc == "flagged":
                from admissions.utils.program_choice_integrity import application_ids_with_suspect_program_choices

                qs = qs.filter(
                    id__in=application_ids_with_suspect_program_choices()
                )

        # --------------------------------------------------------------
        # 3. BUILD EXCEL ROWS
        # --------------------------------------------------------------
        headers = [
            "FIRST NAME",
            "LAST NAME",
            "GENDER",
            "PHONE",
            "EMAIL",
            "NATIONALITY",
            "PROGRAMMES APPLIED",
            "CAMPUS",
            "BATCH",
            "ACADEMIC LEVEL",
            "STATUS",
            "REASON",
            "APPLICATION DATE",
            "ENTRY TYPE",
        ]

        rows = []
        for app in qs:
            # Get all chosen programs as comma-separated
            programs_list = [
                choice.program.name for choice in getattr(app, 'prefetched_program_choices', [])
            ]
            programs_str = ", ".join(programs_list) if programs_list else "—"

            rows.append([
                app.first_name or "",
                app.last_name or "",
                app.gender or "",
                app.phone or "",
                app.email or "",
                app.nationality or "",
                programs_str,
                app.campus.name if app.campus else "",
                app.batch.name if app.batch else "",
                app.academic_level.name if app.academic_level else "",
                app.status or "",
                app.pending_reason or "No reason",
                app.created_at.strftime("%Y-%m-%d %H:%M") if app.created_at else "",
                "Direct Entry" if getattr(app, 'is_direct_entry', False) else "Online",
            ])

        # --------------------------------------------------------------
        # 4. CREATE EXCEL
        # --------------------------------------------------------------
        wb = create_workbook(headers, rows, sheet_name="Applicants Report")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"applicants_report_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response

class ExportFirstRegistrationReportExcel(APIView):
    """Registration-details Excel after desk verification (default: verified students only)."""

    permission_classes = [IsAuthenticated, ExportVerificationRegisterPermission]

    def get(self, request):
        academic_year = request.query_params.get("academic_year") or get_current_academic_year()
        admission_period = request.query_params.get("admission_period")
        campus_id = request.query_params.get("campus")
        program_id = request.query_params.get("program")
        faculty_id = request.query_params.get("faculty")
        documents_verified_raw = request.query_params.get("documents_verified")
        documents_verified = (documents_verified_raw or "").lower()
        is_registered = (request.query_params.get("is_registered") or "").lower()
        include_all = (request.query_params.get("include_all") or "").lower() in (
            "1",
            "true",
            "yes",
        )

        qs = (
            AdmittedStudent.objects.select_related(
                "application",
                "application__academic_level",
                "admitted_program",
                "admitted_program__faculty",
                "admitted_campus",
                "admitted_batch",
                "physical_documents_verified_by",
                "programme_enrollment",
            )
            .filter(is_admitted=True)
        )

        if academic_year:
            qs = qs.filter(admitted_batch__academic_year=academic_year)
        if admission_period:
            qs = qs.filter(admitted_batch__name__icontains=admission_period)
        if campus_id:
            qs = qs.filter(admitted_campus_id=campus_id)
        if program_id:
            qs = qs.filter(admitted_program_id=program_id)
        if faculty_id:
            qs = qs.filter(admitted_program__faculty_id=faculty_id)

        # Physical documents: default to verified-only (post–desk-check roster) unless include_all or explicit param.
        if documents_verified_raw is not None and str(documents_verified_raw).strip() != "":
            if documents_verified in ("1", "true", "yes"):
                qs = qs.filter(physical_documents_verified=True)
            elif documents_verified in ("0", "false", "no"):
                qs = qs.filter(physical_documents_verified=False)
        elif not include_all:
            qs = qs.filter(physical_documents_verified=True)

        if is_registered in ("1", "true", "yes"):
            qs = qs.filter(is_registered=True)
        elif is_registered in ("0", "false", "no"):
            qs = qs.filter(is_registered=False)

        # Subtitle for cover rows (what verification filter was applied)
        if documents_verified_raw is not None and str(documents_verified_raw).strip() != "":
            if documents_verified in ("0", "false", "no"):
                verification_blurb = "Scope: physical documents not verified only"
            else:
                verification_blurb = "Scope: physical documents verified only"
        elif include_all:
            verification_blurb = "Scope: all admitted students (desk verification not filtered)"
        else:
            verification_blurb = "Scope: physical documents verified at desk only"

        admitted_students = list(qs.order_by("application__last_name", "application__first_name"))

        headers = [
            "NAME",
            "GEN",
            "REGISTRATION. NO",
            "STUDENT ID",
            "PROGRAM",
            "FACULTY",
            "STUDY",
            "HALL",
            "CAMPUS",
            "CONTACT",
            "NAME OF NEXT OF KIN",
            "TEL. OF NEXT OF KIN",
            "YEAR",
            "INTAKE",
            "YEAR (ACADEMIC)",
            "NATIONALITY",
            "MODE OF ADMISSION",
            "DIRECT ADMISSION REASON",
            "PHYS DOCS VERIFIED",
            "VERIFIED AT",
            "VERIFIED BY",
        ]
        n_cols = len(headers)
        rows = []

        for adm in admitted_students:
            app = adm.application
            batch = adm.admitted_batch
            prog = adm.admitted_program
            faculty_name = ""
            if prog and prog.faculty_id:
                faculty_name = (prog.faculty.name or "").strip()
            program_code = ""
            if prog:
                program_code = (prog.code or prog.short_form or "").strip()
            spe = getattr(adm, "programme_enrollment", None)
            year_study = ""
            if spe is not None:
                year_study = str(spe.current_year_of_study)
            verified_by = ""
            if adm.physical_documents_verified_by_id:
                vb = adm.physical_documents_verified_by
                verified_by = (vb.get_full_name() or vb.username or "") if vb else ""
            verified_at = ""
            if adm.physical_documents_verified_at:
                verified_at = adm.physical_documents_verified_at.strftime("%Y-%m-%d %H:%M")

            rows.append(
                [
                    _application_full_name_upper(app),
                    _gender_short(app),
                    adm.reg_no or "",
                    adm.student_id or "",
                    program_code,
                    faculty_name,
                    (adm.study_mode or "").strip().upper(),
                    "",
                    (adm.admitted_campus.name if adm.admitted_campus_id else "") or "",
                    (app.phone or "").strip(),
                    (app.next_of_kin_name or "").strip(),
                    (app.next_of_kin_contact or "").strip(),
                    year_study,
                    (batch.name if batch else "") or "",
                    (batch.academic_year if batch else "") or "",
                    (app.nationality or "").strip(),
                    _admission_mode_label(app),
                    _direct_admission_reason(adm),
                    "Y" if adm.physical_documents_verified else "N",
                    verified_at,
                    verified_by,
                ]
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Verified registration"

        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=11)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", fill_type="solid")
        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        t1 = ws.cell(
            row=1,
            column=1,
            value="VERIFIED REGISTRATION ROSTER (POST–DOCUMENT CHECK)",
        )
        t1.font = title_font
        t1.alignment = Alignment(horizontal="center", vertical="center")

        sub_parts = [f"Academic year {academic_year}", verification_blurb]
        if admission_period:
            sub_parts.append(f"Intake contains: {admission_period}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        t2 = ws.cell(row=2, column=1, value=" · ".join(sub_parts))
        t2.font = subtitle_font
        t2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(28, len(str(header)) + 2))

        for row in rows:
            ws.append(row)
        data_start = header_row + 1
        for row in ws.iter_rows(
            min_row=data_start, max_row=ws.max_row, min_col=1, max_col=n_cols
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

        ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_ay = "".join(c if c.isalnum() or c in "-_" else "_" for c in academic_year)
        fname = f"verified_registration_roster_{safe_ay}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response